import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import glob, sys

def find_calc_file():
    matches = glob.glob("refer/*(calculation).xlsx")
    if not matches:
        print("No calculation template found")
        sys.exit(1)
    return matches[0]

def main():
    path = find_calc_file()
    print(f"Loading {path}...")
    wb = openpyxl.load_workbook(path, data_only=False)
    
    spec_lines = []
    spec_lines.append("# Calculation Template Spec")
    spec_lines.append("IMPLEMENTATION NOTE: Read this spec completely before writing any code. After each section is implemented, tick the verification check for that section.\n")
    spec_lines.append(f"**Source Workbook**: `{Path(path).name}`\n")
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ["Data", "master powertrain", "BEV Series Name Table"]:
            continue # skip raw data sheets
            
        ws = wb[sheet_name]
        spec_lines.append(f"## Sheet: {sheet_name}\n")
        
        formulas = []
        formatting = []
        headers = []
        
        # Scan first 20 rows and 30 columns for structure
        for r in range(1, 25):
            row_vals = []
            for c in range(1, 30):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                row_vals.append(str(val) if val is not None else "")
                
                # Check for formula
                if isinstance(val, str) and val.startswith("="):
                    formulas.append(f"Cell {cell.coordinate}: `{val}`")
                
                # Check for distinct formatting (just a sample of background colors)
                if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
                    formatting.append(f"Cell {cell.coordinate}: fgColor=#{cell.fill.fgColor.rgb}")
                    
            # Try to grab header strings
            if any(row_vals):
                headers.append(f"Row {r}: " + " | ".join([v for v in row_vals if v]))
        
        spec_lines.append("### Layout & Headers")
        spec_lines.append("```text")
        for h in headers[:8]: # top 8 rows usually contain all headers
            spec_lines.append(h)
        spec_lines.append("```\n")
        
        spec_lines.append("### Formulas")
        if formulas:
            # deduplicate similar formulas (e.g. SUM(C8:C20) down a column)
            unique_f = list(dict.fromkeys(formulas))
            for f in unique_f[:10]: # show top 10 unique formulas
                spec_lines.append(f"- {f}")
            if len(unique_f) > 10:
                spec_lines.append(f"- ... and {len(unique_f)-10} more formulas.")
        else:
            spec_lines.append("No explicit formulas found in top rows. (Likely static text or data from script.)")
        spec_lines.append("\n")
        
        spec_lines.append("### Formatting & Colors")
        if formatting:
            unique_fmt = list(dict.fromkeys(formatting))
            for fmt in unique_fmt[:10]:
                spec_lines.append(f"- {fmt}")
        else:
            spec_lines.append("No special fill colors detected.")
        spec_lines.append("\n")
        
    out_path = Path("specs") / "calculation_spec.md"
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(spec_lines))
    print(f"Spec written to {out_path}")

if __name__ == "__main__":
    main()
