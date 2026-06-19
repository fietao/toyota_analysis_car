#!/usr/bin/env python3
"""
One-time script: strip all numeric data from refer/ workbooks.
Keeps sheet names, text labels, and cell formatting.
Uses pandas for large sheets (fast), openpyxl cell-by-cell for small sheets.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

REFER = Path(__file__).resolve().parent / "refer"

FILES = [
    REFER / "202605_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - พฤษภาคม 2569 - Model.xlsx",
    REFER / "202605_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - พฤษภาคม 2569(calculation).xlsx",
]

LARGE_SHEETS = {"Data"}  # handled via pandas — much faster for 600k+ rows


def strip_large_sheet(path: Path, sheet_name: str):
    """Read sheet with pandas, blank numeric columns, write back."""
    df = pd.read_excel(str(path), sheet_name=sheet_name, header=None)
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = None
    with pd.ExcelWriter(str(path), engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    print(f"    → {sheet_name}: numeric columns blanked (pandas)")


def strip_small_sheet(ws):
    """Clear numeric values and formulas cell by cell (small sheets only)."""
    cleared = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.value = None
                cleared += 1
            elif isinstance(cell.value, str) and cell.value.strip().startswith("="):
                cell.value = None
                cleared += 1
    return cleared


def strip_workbook(path: Path):
    print(f"\nProcessing: {path.name}")

    # First pass: small sheets via openpyxl (preserves formatting)
    wb = load_workbook(str(path))
    for sheet_name in wb.sheetnames:
        if sheet_name in LARGE_SHEETS:
            print(f"    → {sheet_name}: skipping (will use pandas)")
            continue
        ws = wb[sheet_name]
        cleared = strip_small_sheet(ws)
        print(f"    → {sheet_name}: {cleared} cells cleared")
    wb.save(str(path))
    print(f"  Small sheets saved.")

    # Second pass: large sheets via pandas (fast)
    for sheet_name in LARGE_SHEETS:
        strip_large_sheet(path, sheet_name)

    print(f"  Done: {path.name}")


for f in FILES:
    if not f.exists():
        print(f"NOT FOUND: {f.name}")
        continue
    strip_workbook(f)

print("\nAll done.")
