import openpyxl
from pathlib import Path
import glob, sys

def find_calc_file():
    matches = glob.glob("refer/*(calculation).xlsx")
    if not matches:
        sys.exit(1)
    return matches[0]

def main():
    path = find_calc_file()
    print(f"Loading {path} in read-only mode...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    
    spec_lines = []
    spec_lines.append("# Calculation Template Spec")
    spec_lines.append("IMPLEMENTATION NOTE: Read this spec completely before writing any code. After each section is implemented, tick the verification check for that section.\n")
    spec_lines.append(f"**Source Workbook**: `{Path(path).name}`\n")
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ["Data", "master powertrain", "BEV Series Name Table", "Pivot"]:
            continue
            
        ws = wb[sheet_name]
        spec_lines.append(f"## Sheet: {sheet_name}\n")
        
        formulas = []
        headers = []
        
        r_idx = 0
        for row in ws.iter_rows(min_row=1, max_row=20, max_col=30, values_only=False):
            r_idx += 1
            row_vals = []
            for cell in row:
                val = cell.value
                row_vals.append(str(val) if val is not None else "")
                
                if isinstance(val, str) and val.startswith("="):
                    formulas.append(f"Cell {cell.coordinate}: `{val}`")
                    
            if any(row_vals):
                headers.append(f"Row {r_idx}: " + " | ".join([v for v in row_vals if v]))
        
        spec_lines.append("### Layout & Headers")
        spec_lines.append("```text")
        for h in headers[:8]:
            spec_lines.append(h)
        spec_lines.append("```\n")
        
        spec_lines.append("### Formulas")
        if formulas:
            unique_f = list(dict.fromkeys(formulas))
            for f in unique_f[:15]:
                spec_lines.append(f"- {f}")
        else:
            spec_lines.append("- No explicit formulas found in top rows. Values appear hardcoded.")
        spec_lines.append("\n")
        
    out_path = Path("specs") / "calculation_template_spec.md"
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(spec_lines))
    print(f"Spec written to {out_path}")

if __name__ == "__main__":
    main()
