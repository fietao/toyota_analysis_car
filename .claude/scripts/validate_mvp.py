#!/usr/bin/env python3
"""
validate_mvp.py — Runs validation checks on test_model_1.xlsx in a highly optimized way.
"""

import sys
from pathlib import Path
import openpyxl
import pandas as pd

# Force stdout to UTF-8
sys.stdout.reconfigure(encoding="utf-8")

def main():
    base = Path(__file__).resolve().parents[2]
    file_path = base / "test_model_1.xlsx"
    
    print("==================================================")
    print("      OPTIMIZED SPREADSHEET VALIDATION REPORT     ")
    print("==================================================")
    print(f"Target File: {file_path.name}")
    
    if not file_path.exists():
        print(f"ERROR: {file_path.name} not found! Please run the pipeline first.")
        sys.exit(1)
        
    print("\nLoading sheets in read-only mode...")
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    wb_formulas = openpyxl.load_workbook(str(file_path), read_only=True, data_only=False)
    
    # Check 1: Sheet Names
    print("Check 1: Sheet Names...")
    required_sheets = ["BEV by Model", "BEV by Model (2)"]
    missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
    if missing_sheets:
        print(f"  [FAIL] Missing sheets: {missing_sheets}")
        sys.exit(1)
    print("  [PASS] Required sheets exist.")

    # Check 2: Filter Headers (A1:B3)
    print("\nCheck 2: Filter Headers (A1:B3)...")
    for sname in required_sheets:
        ws = wb[sname]
        
        # Read filter cells
        a1 = ws.cell(1, 1).value
        b1 = ws.cell(1, 2).value
        a2 = ws.cell(2, 1).value
        b2 = ws.cell(2, 2).value
        a3 = ws.cell(3, 1).value
        b3 = ws.cell(3, 2).value
        
        filters = {
            "A1": ("ประเภทรถ", a1),
            "B1": ("(Multiple Items)", b1),
            "A2": ("จังหวัด", a2),
            "B2": ("(All)", b2),
            "A3": ("Powertrain", a3),
            "B3": ("BEV Major", b3)
        }
        
        errors = []
        for cell, (expected, actual) in filters.items():
            if str(actual).strip() != expected:
                errors.append(f"Cell {cell}: expected '{expected}', got '{actual}'")
        
        if errors:
            print(f"  [FAIL] Sheet '{sname}' filter mismatches:")
            for err in errors:
                print(f"    - {err}")
            sys.exit(1)
    print("  [PASS] Filter labels and values match specifications.")

    # Check 3 & 4: Data Columns & Row Totals Reconciliation
    print("\nCheck 3 & 4: Columns Alignment & Row Totals...")
    
    # ── Verify BEV by Model (Hierarchical) ──
    ws1 = wb["BEV by Model"]
    ws1_form = wb_formulas["BEV by Model"]
    
    # Validate column headers first
    assert ws1.cell(6, 14).value == "2568 Total" or "2568" in str(ws1.cell(6, 14).value), "2568 Total offset incorrect"
    assert ws1.cell(6, 20).value == "2569 Total" or "2569" in str(ws1.cell(6, 20).value), "2569 Total offset incorrect"
    assert ws1.cell(6, 21).value == "Grand Total" or ws1.cell(6, 21).value == "Grand Total Total", "Grand Total offset incorrect"
    
    ws1_mismatches = 0
    # Data rows start at 8 (row 7 has col headers, row 8 has first brand/model)
    # The bottom row is Grand Total
    bottom_row_ws1 = None
    for r in range(8, 2000): # Safe upper limit for rows
        label = ws1.cell(r, 1).value
        if label == "Grand Total" or label is None:
            bottom_row_ws1 = r
            break
            
        # Get 2568 monthly sum (columns 2 to 13)
        vals_2568 = [ws1.cell(r, col).value for col in range(2, 14)]
        sum_2568 = sum(int(v) for v in vals_2568 if v is not None)
        total_2568 = ws1.cell(r, 14).value or 0
        
        # Get 2569 monthly sum (columns 15 to 19)
        vals_2569 = [ws1.cell(r, col).value for col in range(15, 20)]
        sum_2569 = sum(int(v) for v in vals_2569 if v is not None)
        total_2569 = ws1.cell(r, 20).value or 0
        
        # Grand Total column formula verification
        gt_formula = ws1_form.cell(r, 21).value
        expected_formula = f"=N{r}+T{r}"
        
        if sum_2568 != total_2568:
            print(f"    Row {r} ({label}): 2568 sum {sum_2568} != total {total_2568}")
            ws1_mismatches += 1
        if sum_2569 != total_2569:
            print(f"    Row {r} ({label}): 2569 sum {sum_2569} != total {total_2569}")
            ws1_mismatches += 1
        if str(gt_formula).strip().upper() != expected_formula:
            print(f"    Row {r} ({label}): Grand Total formula mismatch! Expected: '{expected_formula}', Got: '{gt_formula}'")
            ws1_mismatches += 1
            
    if ws1_mismatches == 0:
        print("  [PASS] 'BEV by Model' row totals & formula requirements match.")
    else:
        print(f"  [FAIL] 'BEV by Model' had {ws1_mismatches} errors.")
        sys.exit(1)

    # ── Verify BEV by Model (2) (Flat) ──
    ws2 = wb["BEV by Model (2)"]
    ws2_form = wb_formulas["BEV by Model (2)"]
    
    assert ws2.cell(6, 15).value == "2568 Total" or "2568" in str(ws2.cell(6, 15).value), "BEV(2) 2568 Total offset incorrect"
    assert ws2.cell(6, 21).value == "2569 Total" or "2569" in str(ws2.cell(6, 21).value), "BEV(2) 2569 Total offset incorrect"
    assert ws2.cell(6, 22).value == "Grand Total", "BEV(2) Grand Total offset incorrect"

    ws2_mismatches = 0
    bottom_row_ws2 = None
    for r in range(8, 2000):
        model_label = ws2.cell(r, 1).value
        brand_label = ws2.cell(r, 2).value
        if model_label == "Grand Total" or model_label is None:
            bottom_row_ws2 = r
            break
            
        vals_2568 = [ws2.cell(r, col).value for col in range(3, 15)]
        sum_2568 = sum(int(v) for v in vals_2568 if v is not None)
        total_2568 = ws2.cell(r, 15).value or 0
        
        vals_2569 = [ws2.cell(r, col).value for col in range(16, 21)]
        sum_2569 = sum(int(v) for v in vals_2569 if v is not None)
        total_2569 = ws2.cell(r, 21).value or 0
        
        gt_formula = ws2_form.cell(r, 22).value
        expected_formula = f"=O{r}+U{r}"
        
        if sum_2568 != total_2568:
            print(f"    Row {r} ({model_label} / {brand_label}): 2568 sum {sum_2568} != total {total_2568}")
            ws2_mismatches += 1
        if sum_2569 != total_2569:
            print(f"    Row {r} ({model_label} / {brand_label}): 2569 sum {sum_2569} != total {total_2569}")
            ws2_mismatches += 1
        if str(gt_formula).strip().upper() != expected_formula:
            print(f"    Row {r} ({model_label}): Grand Total formula mismatch! Expected: '{expected_formula}', Got: '{gt_formula}'")
            ws2_mismatches += 1

    if ws2_mismatches == 0:
        print("  [PASS] 'BEV by Model (2)' row totals & formula requirements match.")
    else:
        print(f"  [FAIL] 'BEV by Model (2)' had {ws2_mismatches} errors.")
        sys.exit(1)

    # Check 5: Grand Totals Reconciliation
    print("\nCheck 5: Grand Totals Reconciliation...")
    
    # Calculate Grand Total from static column totals on the bottom row
    # Bottom row Grand Total = 2568 Total + 2569 Total
    gt_val_sheet1 = int(ws1.cell(bottom_row_ws1, 14).value or 0) + int(ws1.cell(bottom_row_ws1, 20).value or 0)
    gt_val_sheet2 = int(ws2.cell(bottom_row_ws2, 15).value or 0) + int(ws2.cell(bottom_row_ws2, 21).value or 0)
            
    print(f"  Grand Total in 'BEV by Model'    : {gt_val_sheet1:,} (sum of year-total columns)")
    print(f"  Grand Total in 'BEV by Model (2)': {gt_val_sheet2:,} (sum of year-total columns)")
    
    if gt_val_sheet1 != gt_val_sheet2:
        print("  [FAIL] Grand Totals do not match between sheet 1 and sheet 2!")
        sys.exit(1)
        
    wb.close()
    wb_formulas.close()
    
    # Reconcile with Data sheet using fast Pandas read (only read 3 columns)
    print("  Loading Data sheet using fast Pandas engine...")
    df_data = pd.read_excel(str(file_path), sheet_name="Data", header=5, usecols=["ปี", "Powertrain", "รุ่นรถ", "จำนวนรถ"])
    
    # Filter conditions: Powertrain == "BEV Major", รุ่นรถ is not null, and Year is 2568 or 2569
    df_filtered = df_data[
        (df_data["Powertrain"] == "BEV Major") & 
        (df_data["รุ่นรถ"].notna()) & 
        (df_data["รุ่นรถ"].astype(str).str.strip() != "") & 
        (df_data["รุ่นรถ"].astype(str).str.strip().str.lower() != "nan") &
        (df_data["ปี"].isin([2568, 2569]))
    ]
    data_sum = int(df_filtered["จำนวนรถ"].sum())
    
    print(f"  Sum of 'จำนวนรถ' in Data sheet   : {data_sum:,} (filtered by Powertrain == 'BEV Major' and model is not null)")
    
    if gt_val_sheet1 != data_sum:
        print(f"  [FAIL] Reconciliation mismatch! Pivot Total is {gt_val_sheet1:,} but Data sheet filtered sum is {data_sum:,}.")
        sys.exit(1)
        
    print("  [PASS] Grand Totals reconcile perfectly across sheets and base Data!")
    print("\n==================================================")
    print("      ALL VALIDATION CHECKS COMPLETED SUCCESS     ")
    print("==================================================")

if __name__ == "__main__":
    main()
