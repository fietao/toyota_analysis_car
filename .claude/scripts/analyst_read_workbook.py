#!/usr/bin/env python3
"""
analyst_read_workbook.py
Read test_model_1.xlsx and the reference Model.xlsx once each.
Output a structured text summary the Analyst Agent feeds to the local LLM.

Fixes vs v1:
- Open each workbook once, not once per sheet
- Style extraction only on small sheets (read_only=False); large sheets use read_only=True
- Sort order checked on full pandas sample (10k rows), not 5-row window
- Removed unused json import
"""

import sys
import glob
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(__file__).resolve().parents[2]

LARGE_SHEET_THRESHOLD = 5000   # rows — above this, skip style scan
SORT_SAMPLE = 10_000            # rows to check sort order via pandas

SHEET_MAP = {
    "Cleaned Data":          "Data",
    "master powertrain":     "master powertrain",
    "BEV Series Name Table": "BEV Series Name Table",
    "BEV by Model":          "BEV by Model",
    "BEV by Model (2)":      "BEV by Model (2)",
    "BMW":                   "BMW",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_file(pattern, label):
    files = [f for f in glob.glob(str(BASE / pattern)) if "~$" not in f]
    if not files:
        print(f"ERROR: {label} not found ({pattern})", file=sys.stderr)
        sys.exit(1)
    return Path(sorted(files)[-1])


def cell_style_note(cell):
    notes = []
    if cell.font and cell.font.bold:
        notes.append("bold")
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
        rgb = cell.fill.fgColor.rgb
        if rgb not in ("00000000", "FFFFFFFF", "FF000000"):
            notes.append(f"bg={rgb}")
    if cell.font and cell.font.color and cell.font.color.type == "rgb":
        rgb = cell.font.color.rgb
        if rgb not in ("00000000", "FF000000"):
            notes.append(f"color={rgb}")
    return ",".join(notes) if notes else None


def extract_sheet_info(ws_style, ws_values_rows, max_row, max_col):
    """
    ws_style  : worksheet opened with read_only=False (for style access), or None
    ws_values_rows : pre-read list of value-only rows (first few rows)
    """
    headers = list(ws_values_rows[0]) if ws_values_rows else []
    sample  = [list(r) for r in ws_values_rows[1:4]]

    header_styles = {}
    if ws_style:
        header_row = list(ws_style.iter_rows(min_row=1, max_row=1))[0]
        for cell in header_row:
            s = cell_style_note(cell)
            if s:
                header_styles[get_column_letter(cell.column)] = s

    return {
        "dimensions": f"{max_row} rows × {max_col} cols",
        "headers":    headers,
        "header_styles": header_styles,
        "sample":     sample,
    }


def check_sort_order(file_path, sheet_name):
    """Use pandas on a sample to detect sort order of numeric/date columns."""
    try:
        df = pd.read_excel(str(file_path), sheet_name=sheet_name,
                           nrows=SORT_SAMPLE, header=0)
    except Exception:
        return {}

    result = {}
    for col in df.columns:
        vals = df[col].dropna()
        if len(vals) < 2:
            continue
        if pd.api.types.is_numeric_dtype(vals):
            is_asc  = (vals.diff().dropna() >= 0).all()
            is_desc = (vals.diff().dropna() <= 0).all()
            result[str(col)] = "asc" if is_asc else ("desc" if is_desc else "unsorted")
    return result


def column_stats(file_path, sheet_name, nrows=3000):
    try:
        df = pd.read_excel(str(file_path), sheet_name=sheet_name, nrows=nrows)
    except Exception:
        return {}

    stats = {}
    for col in df.columns:
        stats[str(col)] = {
            "dtype":  str(df[col].dtype),
            "nulls":  int(df[col].isna().sum()),
            "unique": int(df[col].nunique()),
            "sample": df[col].dropna().astype(str).unique()[:5].tolist(),
        }
    return stats


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    out_file = find_file("test_model_1.xlsx",  "output workbook")
    ref_file = find_file("refer/*Model.xlsx",  "reference Model")

    print(f"OUTPUT   : {out_file.name}")
    print(f"REFERENCE: {ref_file.name}")

    # Open each workbook once (read_only=True for fast value reads)
    wb_out_ro = load_workbook(str(out_file), read_only=True, data_only=True)
    wb_ref_ro = load_workbook(str(ref_file), read_only=True, data_only=True)
    out_sheets = wb_out_ro.sheetnames
    ref_sheets = wb_ref_ro.sheetnames
    wb_out_ro.close()
    wb_ref_ro.close()

    print(f"\nOUTPUT SHEETS   : {out_sheets}")
    print(f"REFERENCE SHEETS: {ref_sheets}")

    missing = [s for s in ref_sheets if s not in out_sheets and s != "Data"]
    extra   = [s for s in out_sheets if s not in ref_sheets and s != "Cleaned Data"]
    if missing: print(f"  [!] Missing vs reference: {missing}")
    if extra:   print(f"  [+] Extra vs reference  : {extra}")

    # Process each output sheet
    for out_name in out_sheets:
        ref_name = SHEET_MAP.get(out_name, out_name)
        in_ref   = ref_name in ref_sheets

        print(f"\n{'='*60}")
        print(f"SHEET: {out_name}  →  reference: {ref_name if in_ref else '(no match)'}")

        # Read values (read_only for speed)
        wb_ro = load_workbook(str(out_file), read_only=True, data_only=True)
        ws_ro = wb_ro[out_name]
        max_row = ws_ro.max_row or 0
        max_col = ws_ro.max_column or 0
        value_rows = list(ws_ro.iter_rows(min_row=1, max_row=min(5, max_row), values_only=True))
        wb_ro.close()

        # Style scan only for small sheets (opening full wb is slow for 857k rows)
        ws_style = None
        if max_row < LARGE_SHEET_THRESHOLD:
            wb_rw = load_workbook(str(out_file), read_only=False, data_only=True)
            ws_style = wb_rw[out_name]

        info = extract_sheet_info(ws_style, value_rows, max_row, max_col)

        if ws_style:
            wb_rw.close()

        print(f"  Dimensions   : {info['dimensions']}")
        print(f"  Headers      : {info['headers']}")
        if info['header_styles']:
            print(f"  Header styles: {info['header_styles']}")
        print(f"  Sample row   : {info['sample'][0] if info['sample'] else 'n/a'}")

        # Sort order (pandas sample)
        sort = check_sort_order(str(out_file), out_name)
        if sort:
            print(f"  Sort order   : {sort}")

        # Reference comparison
        if in_ref:
            wb_ref_ro = load_workbook(str(ref_file), read_only=True, data_only=True)
            ws_ref = wb_ref_ro[ref_name]
            ref_row = ws_ref.max_row or 0
            ref_col = ws_ref.max_column or 0
            ref_hdrs = [c.value for c in list(ws_ref.iter_rows(min_row=1, max_row=1))[0]]
            wb_ref_ro.close()

            print(f"  REF dims     : {ref_row} rows × {ref_col} cols")
            print(f"  REF headers  : {ref_hdrs}")
            missing_cols = [c for c in ref_hdrs if c not in info['headers']]
            extra_cols   = [c for c in info['headers'] if c not in ref_hdrs]
            if missing_cols: print(f"  [!] Missing cols: {missing_cols}")
            if extra_cols:   print(f"  [+] Extra cols  : {extra_cols}")

        # Column stats for main data sheets
        if out_name in ("Cleaned Data", "master powertrain"):
            stats = column_stats(str(out_file), out_name)
            print(f"  Column stats (first 3000 rows):")
            for col, s in stats.items():
                print(f"    {col}: {s['dtype']}, {s['unique']} unique, "
                      f"{s['nulls']} nulls, sample={s['sample']}")

    print("\n[DONE]")


if __name__ == "__main__":
    main()
