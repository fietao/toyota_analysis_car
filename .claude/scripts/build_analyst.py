"""
Builds the monthly analyst report from the model output.

Input:  test_model_1.xlsx  (Cleaned Data sheet)
        *- Model.xlsx       (Data sheet for BEV Major, master powertrain, BEV Series Name Table)
Output: YYYYMM_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - เดือนXXXX(analyst).xlsx
"""

import glob, sys, os
from pathlib import Path
import pandas as pd
import xlsxwriter

def _find_root():
    p = Path(__file__).resolve()
    for parent in p.parents:
        if (parent / "CLAUDE.md").exists():
            return parent
    raise RuntimeError("Could not find project root (no CLAUDE.md)")

BASE          = _find_root()
MODEL_PATTERN = str(BASE / "refer" / "*- Model.xlsx")
CLEANED_FILE  = BASE / "test_model_1.xlsx"

THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
THAI_TO_NUM  = {v: k for k, v in THAI_MONTHS.items()}
MONTH_ORDER  = [THAI_MONTHS[i] for i in range(1, 13)]
MONTH_EN     = {
    "มกราคม":"Jan","กุมภาพันธ์":"Feb","มีนาคม":"Mar","เมษายน":"Apr",
    "พฤษภาคม":"May","มิถุนายน":"Jun","กรกฎาคม":"Jul","สิงหาคม":"Aug",
    "กันยายน":"Sep","ตุลาคม":"Oct","พฤศจิกายน":"Nov","ธันวาคม":"Dec",
}
INCLUDE_RY = {'1','2','3','6','9','10','11'}


# ── File helpers ──────────────────────────────────────────────────────────────
def find_file(pattern, label):
    matches = [m for m in glob.glob(pattern) if not Path(m).name.startswith("~$")]
    if not matches:
        print(f"ERROR: No {label}: {pattern}"); sys.exit(1)
    return Path(max(matches, key=os.path.getmtime))


def filter_ry(df):
    """Keep only ประเภทรถ รย.1,2,3,6,9,10,11."""
    codes = df["ประเภทรถ"].str.extract(r"รย\.(\d+)")[0]
    return df[codes.isin(INCLUDE_RY)].copy()


def safe_div(a, b):
    return (a / b) if b else None


def _v(counts, yr, m):
    return int(counts.get((yr, m), 0))


def _ytd(counts, yr, months):
    return sum(_v(counts, yr, m) for m in months)


# ── Row value builder ─────────────────────────────────────────────────────────
def make_row_vals(row_c, ref_c, prev_year, curr_year, curr_months, mode):
    """
    row_c   : dict (year, month) -> int   for this row
    ref_c   : dict (year, month) -> int   for the denominator (grand total or parent)
    mode    : 'powertrain' | 'rank' | 'brand_pt' | 'bev_parent' | 'bev_child' | 'bev_rank'
    Returns list of cell values.
    """
    curr_month = curr_months[-1]
    cidx       = MONTH_ORDER.index(curr_month)
    before_m   = MONTH_ORDER[:cidx]      # e.g. Jan-Apr for May
    after_m    = MONTH_ORDER[cidx + 1:]  # e.g. Jun-Dec for May
    n_curr     = len(curr_months)

    vals = []

    # ── Prev year ──────────────────────────────────────────────────────────────
    for m in before_m:
        vals.append(_v(row_c, prev_year, m) or None)

    pc  = _v(row_c, prev_year, curr_month)
    rpc = _v(ref_c, prev_year, curr_month)
    vals += [pc or None, safe_div(pc, rpc)]

    py  = _ytd(row_c, prev_year, curr_months)
    rpy = _ytd(ref_c, prev_year, curr_months)
    vals += [py or None, safe_div(py, rpy)]

    for m in after_m:
        vals.append(_v(row_c, prev_year, m) or None)

    pf  = _ytd(row_c, prev_year, MONTH_ORDER)
    rpf = _ytd(ref_c, prev_year, MONTH_ORDER)
    vals += [pf or None, safe_div(pf, rpf)]

    # ── Curr year ──────────────────────────────────────────────────────────────
    for m in curr_months:
        vals.append(_v(row_c, curr_year, m) or None)

    cc  = _v(row_c, curr_year, curr_month)
    rcc = _v(ref_c, curr_year, curr_month)
    cs  = safe_div(cc, rcc)
    ps  = safe_div(pc, rpc)
    vals.append(cs)

    prev_m    = MONTH_ORDER[cidx - 1] if cidx > 0 else None
    mom_v     = _v(row_c, curr_year, prev_m) if prev_m else 0
    cy        = _ytd(row_c, curr_year, curr_months)
    rcy       = _ytd(ref_c, curr_year, curr_months)

    if mode == "powertrain":
        vals.append(safe_div(cc - mom_v, mom_v))
        vals.append(safe_div(cc - pc, pc))
        vals += [cy or None, safe_div(cy, rcy)]

    elif mode == "rank":
        diff = (cs - ps) if (cs is not None and ps is not None) else None
        vals.append(diff)

    elif mode in ("brand_pt", "bev_parent", "bev_child", "bev_rank"):
        diff = (cs - ps) if (cs is not None and ps is not None) else None
        vals.append(diff)
        vals.append(safe_div(cc - mom_v, mom_v))
        vals.append(safe_div(cc - pc, pc))
        vals.append(cy or None)

    return vals


# ── Header writer ─────────────────────────────────────────────────────────────
def write_header(ws, row0, title, pt_label, pt_val, row_label,
                 prev_year, curr_year, curr_months, extras, fmt_h, second_label_col=None):
    """Write rows 0..6 of a sheet header. row0 = first Excel row (0-indexed)."""
    curr_month = curr_months[-1]
    cidx       = MONTH_ORDER.index(curr_month)
    before_m   = MONTH_ORDER[:cidx]
    after_m    = MONTH_ORDER[cidx + 1:]
    n_curr     = len(curr_months)
    n_after    = len(after_m)

    # Column offsets after the label col(s)
    label_cols = 2 if second_label_col else 1
    c0 = label_cols  # first data column

    # Year-span column for 2568 (starts at c0)
    prev_span = len(before_m) + 2 + 2 + n_after + 2  # +2 for curr_month+share, +2 for ytd+share, +2 for year+share

    ws.write(row0, 0, title)
    ws.write(row0 + 1, 0, f"ประเภทรถ : {pt_label}")
    ws.write(row0 + 2, 0, f"Powertrain : {pt_val}")

    # Row +3: year spans
    r3 = row0 + 3
    ws.write(r3, 0, row_label, fmt_h)
    if second_label_col:
        ws.write(r3, 1, second_label_col, fmt_h)
    ws.write(r3, c0, prev_year, fmt_h)
    ws.write(r3, c0 + prev_span, curr_year, fmt_h)

    # Row +4: period sub-headers
    r4 = row0 + 4
    col = c0
    for m in before_m:
        ws.write(r4, col, MONTH_EN[m], fmt_h); col += 1
    ws.write(r4, col, MONTH_EN[curr_month], fmt_h); col += 1  # curr month label
    col += 1  # share blank
    ws.write(r4, col, f"Jan-{MONTH_EN[curr_month]}", fmt_h); col += 1
    col += 1  # share blank
    for m in after_m:
        ws.write(r4, col, MONTH_EN[m], fmt_h); col += 1
    ws.write(r4, col, f"{prev_year} Total", fmt_h); col += 2  # +share
    for m in curr_months:
        ws.write(r4, col, MONTH_EN[m], fmt_h); col += 1
    for extra in extras:
        lbl = {"diff": "Diff", "growth_mom": f"Growth vs {MONTH_EN[MONTH_ORDER[cidx-1]]} {curr_year}" if cidx > 0 else "Growth MoM",
               "growth_yoy": f"Growth vs {MONTH_EN[curr_month]} {prev_year}", "ytd": f"Jan-{MONTH_EN[curr_month]} {curr_year}",
               "ytd_share": f"{curr_year} Total Share"}.get(extra, extra)
        ws.write(r4, col, lbl, fmt_h); col += 1

    # Row +5: Units/Share labels
    r5 = row0 + 5
    col = c0
    for _ in before_m:
        ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Share", fmt_h); col += 1
    ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Share", fmt_h); col += 1
    for _ in after_m:
        ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Share", fmt_h); col += 1
    for _ in curr_months:
        ws.write(r5, col, "Units", fmt_h); col += 1
    ws.write(r5, col, "Share", fmt_h); col += 1
    extra_labels = {"diff": "Diff", "growth_mom": "Growth MoM", "growth_yoy": "Growth YoY",
                    "ytd": "Units", "ytd_share": "Share"}
    for extra in extras:
        ws.write(r5, col, extra_labels.get(extra, extra), fmt_h); col += 1

    return row0 + 6  # first data row


# ── Write one data row ────────────────────────────────────────────────────────
def write_data_row(ws, row, labels, vals, fmt=None):
    for c, lbl in enumerate(labels):
        if lbl is not None:
            if fmt:
                ws.write(row, c, lbl, fmt)
            else:
                ws.write(row, c, lbl)
    for c, v in enumerate(vals, start=len(labels)):
        if v is not None:
            ws.write(row, c, v)


# ── Sheet: Pivot ──────────────────────────────────────────────────────────────
def build_pivot(wb, df, prev_year, curr_year, curr_months, fmt_h):
    ws = wb.add_worksheet("Pivot")
    ws.write(0, 0, "ประเภทรถ"); ws.write(0, 1, "(Multiple Items)")
    ws.write(1, 0, "ยี่ห้อรถ2"); ws.write(1, 1, "(All)")
    ws.write(2, 0, "Sum of จำนวนรถ"); ws.write(2, 1, "Column Labels")

    row4_c = [None]  # col 0 blank in row 4
    row5_c = ["Row Labels"]
    for m in MONTH_ORDER:
        row4_c.append(prev_year if m == MONTH_ORDER[0] else None)
        row5_c.append(MONTH_EN[m])
    row4_c.append(None); row5_c.append(f"{prev_year} Total")
    for m in curr_months:
        row4_c.append(curr_year if m == curr_months[0] else None)
        row5_c.append(MONTH_EN[m])
    row4_c.append(None); row5_c.append(f"{curr_year} Total")
    row4_c.append("Grand Total"); row5_c.append(None)

    for c, v in enumerate(row4_c):
        if v is not None: ws.write(3, c, v, fmt_h)
    for c, v in enumerate(row5_c):
        if v is not None: ws.write(4, c, v, fmt_h)

    pts = ["ICE","BEV","HEV","PHEV"]
    counts_all = df.groupby(["ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    counts_pt  = df.groupby(["Powertrain","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()

    def row_for(c_dict):
        vals = []
        for m in MONTH_ORDER:
            vals.append(_v(c_dict, prev_year, m) or None)
        py = _ytd(c_dict, prev_year, MONTH_ORDER)
        vals.append(py or None)
        for m in curr_months:
            vals.append(_v(c_dict, curr_year, m) or None)
        cy = _ytd(c_dict, curr_year, curr_months)
        vals.append(cy or None)
        vals.append((py + cy) or None)
        return vals

    data_row = 5
    for pt in pts:
        c_dict = {(yr, m): v for (k, yr, m), v in counts_pt.items() if k == pt}
        vals = row_for(c_dict)
        ws.write(data_row, 0, pt)
        for c, v in enumerate(vals, 1):
            if v is not None: ws.write(data_row, c, v)
        data_row += 1

    gt_vals = row_for(counts_all)
    ws.write(data_row, 0, "Grand Total", fmt_h)
    for c, v in enumerate(gt_vals, 1):
        if v is not None: ws.write(data_row, c, v)
    print("      Pivot done")


# ── Sheet: 1.Reg by Powertrain ────────────────────────────────────────────────
def build_reg_powertrain(wb, df, prev_year, curr_year, prev_prev_year, curr_months, fmt_h):
    ws = wb.add_worksheet("1.Reg by Powertrain")
    extras = ["growth_mom","growth_yoy","ytd","ytd_share"]
    data_row = write_header(ws, 0, "Registration by Powertrain",
                            "รย.1,2,3,6,9,10,11", "ALL", "Powertrain",
                            prev_year, curr_year, curr_months, extras, fmt_h)

    pts = ["ICE","BEV","HEV","PHEV"]
    grand_c = df.groupby(["ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    pt_c    = df.groupby(["Powertrain","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()

    def row_c(pt):
        return {(yr, m): v for (k, yr, m), v in pt_c.items() if k == pt}

    # Grand Total row
    vals = make_row_vals(grand_c, grand_c, prev_year, curr_year, curr_months, "powertrain")
    write_data_row(ws, data_row, ["Grand Total"], vals, fmt_h)
    data_row += 1

    for pt in pts:
        rc = row_c(pt)
        vals = make_row_vals(rc, grand_c, prev_year, curr_year, curr_months, "powertrain")
        write_data_row(ws, data_row, [pt], vals)
        data_row += 1

    # Secondary section: Jan-CurrMonth YoY vs prev_prev_year
    data_row += 1
    labels_sec = ["Grand Total"] + pts
    all_sec = [grand_c] + [row_c(pt) for pt in pts]
    cidx = MONTH_ORDER.index(curr_months[-1])
    # Secondary section: Jan-May share of Full Year (both for prev_year).
    # Label goes at Jan-May Units col; ratio goes at Jan-May Share col.
    # Col layout (0-based): 0=label, 1..cidx=before months,
    # cidx+1=curr_month Units, cidx+2=curr_month Share, cidx+3=Jan-May Units, cidx+4=Jan-May Share
    col_label = 1 + cidx + 2   # Jan-May Units column
    col_value = col_label + 1  # Jan-May Share column
    for lbl, rc in zip(labels_sec, all_sec):
        ytd_prev  = _ytd(rc, prev_year, curr_months)         # Jan-May prev_year
        full_prev = _ytd(rc, prev_year, MONTH_ORDER)         # Full year prev_year
        ratio     = safe_div(ytd_prev, full_prev)            # = Jan-May / Full Year
        ws.write(data_row, col_label, lbl)
        if ratio is not None:
            ws.write(data_row, col_value, ratio)
        data_row += 1

    print("      1.Reg by Powertrain done")


# ── Generic brand sheet ───────────────────────────────────────────────────────
def build_brand_sheet(wb, df, sheet_name, title, pt_filter,
                      prev_year, curr_year, curr_months, fmt_h, mode):
    ws = wb.add_worksheet(sheet_name)
    pt_label = pt_filter if pt_filter != "ALL" else "ALL"
    extras = ["diff"] if mode == "rank" else ["diff","growth_mom","growth_yoy","ytd"]
    data_row = write_header(ws, 0, title, "รย.1,2,3,6,9,10,11", pt_label, "Brand",
                            prev_year, curr_year, curr_months, extras, fmt_h)

    sub = df if pt_filter == "ALL" else df[df["Powertrain"] == pt_filter]
    grand_c = sub.groupby(["ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    brand_c = sub.groupby(["ยี่ห้อรถ2","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()

    brands = sorted(
        sub["ยี่ห้อรถ2"].dropna().unique(),
        key=lambda b: -_ytd({(yr,m):v for (k,yr,m),v in brand_c.items() if k==b},
                             curr_year, curr_months)
    )

    vals = make_row_vals(grand_c, grand_c, prev_year, curr_year, curr_months, mode)
    write_data_row(ws, data_row, ["Grand Total"], vals, fmt_h)
    data_row += 1

    for b in brands:
        rc = {(yr,m):v for (k,yr,m),v in brand_c.items() if k==b}
        vals = make_row_vals(rc, grand_c, prev_year, curr_year, curr_months, mode)
        write_data_row(ws, data_row, [b], vals)
        data_row += 1

    print(f"      {sheet_name} done")


# ── Sheet: 7.BEV by Model ─────────────────────────────────────────────────────
def build_bev_by_model(wb, bev, prev_year, curr_year, curr_months, fmt_h):
    ws = wb.add_worksheet("7.BEV by Model")
    extras = ["diff","growth_mom","growth_yoy","ytd"]
    data_row = write_header(ws, 0, "Registration of Major BEV by Model",
                            "รย.1,2,3,6,9,10,11", "BEV Major", "Brand/Model",
                            prev_year, curr_year, curr_months, extras, fmt_h)

    grand_c  = bev.groupby(["ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    brand_c  = bev.groupby(["ยี่ห้อรถ2","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    model_c  = bev.groupby(["ยี่ห้อรถ2","รุ่นรถ2","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()

    def brand_ytd(b):
        return _ytd({(yr,m):v for (k,yr,m),v in brand_c.items() if k==b}, curr_year, curr_months)

    brands = sorted(bev["ยี่ห้อรถ2"].dropna().unique(), key=lambda b: -brand_ytd(b))

    vals = make_row_vals(grand_c, grand_c, prev_year, curr_year, curr_months, "bev_parent")
    write_data_row(ws, data_row, ["Grand Total"], vals, fmt_h)
    data_row += 1

    for brand in brands:
        brc = {(yr,m):v for (k,yr,m),v in brand_c.items() if k==brand}
        vals = make_row_vals(brc, grand_c, prev_year, curr_year, curr_months, "bev_parent")
        write_data_row(ws, data_row, [brand], vals, fmt_h)
        data_row += 1

        models = sorted(
            bev[bev["ยี่ห้อรถ2"]==brand]["รุ่นรถ2"].dropna().unique(),
            key=lambda m2: -_ytd({(yr,m):v for (k1,k2,yr,m),v in model_c.items() if k1==brand and k2==m2},
                                  curr_year, curr_months)
        )
        for mdl in models:
            mrc = {(yr,m):v for (k1,k2,yr,m),v in model_c.items() if k1==brand and k2==mdl}
            vals = make_row_vals(mrc, brc, prev_year, curr_year, curr_months, "bev_child")
            write_data_row(ws, data_row, [mdl], vals)
            data_row += 1

    print("      7.BEV by Model done")


# ── Sheet: 8.Rank by BEV Model ───────────────────────────────────────────────
def build_rank_bev_model(wb, bev, prev_year, curr_year, curr_months, fmt_h):
    ws = wb.add_worksheet("8.Rank by BEV Model")
    extras = ["diff","growth_mom","growth_yoy"]
    data_row = write_header(ws, 0, "Ranking of Registration of Major BEV Models",
                            "รย.1,2,3,6,9,10,11", "BEV Major", "Model",
                            prev_year, curr_year, curr_months, extras, fmt_h,
                            second_label_col="Brand")

    grand_c  = bev.groupby(["ปี","เดือน"])["จำนวนรถ"].sum().to_dict()
    model_c  = bev.groupby(["รุ่นรถ2","ยี่ห้อรถ2","ปี","เดือน"])["จำนวนรถ"].sum().to_dict()

    model_brand_pairs = bev.groupby(["รุ่นรถ2","ยี่ห้อรถ2"]).size().reset_index()[["รุ่นรถ2","ยี่ห้อรถ2"]]
    pairs = [(r.รุ่นรถ2, r.ยี่ห้อรถ2) for _, r in model_brand_pairs.iterrows()]

    pairs_sorted = sorted(
        pairs,
        key=lambda p: -_ytd({(yr,m):v for (k1,k2,yr,m),v in model_c.items() if k1==p[0] and k2==p[1]},
                             curr_year, curr_months)
    )

    vals = make_row_vals(grand_c, grand_c, prev_year, curr_year, curr_months, "bev_rank")
    write_data_row(ws, data_row, ["Grand Total", None], vals, fmt_h)
    data_row += 1

    for (mdl, brand) in pairs_sorted:
        mrc = {(yr,m):v for (k1,k2,yr,m),v in model_c.items() if k1==mdl and k2==brand}
        vals = make_row_vals(mrc, grand_c, prev_year, curr_year, curr_months, "bev_rank")
        write_data_row(ws, data_row, [mdl, brand], vals)
        data_row += 1

    print("      8.Rank by BEV Model done")


# ── Copy sheets from source workbook ─────────────────────────────────────────
def copy_sheet_raw(src_path, sheet_name, wb_out, out_name=None):
    """Copy a sheet cell-by-cell (values only) into wb_out."""
    import openpyxl
    out_name = out_name or sheet_name
    wb_src = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    if sheet_name not in wb_src.sheetnames:
        print(f"      Warning: sheet '{sheet_name}' not found in {Path(src_path).name}")
        wb_src.close(); return
    ws_src = wb_src[sheet_name]
    ws_dst = wb_out.add_worksheet(out_name)
    for r_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val is not None:
                ws_dst.write(r_idx, c_idx, val)
    wb_src.close()
    print(f"      {out_name} copied")


# ── Copy Cleaned Data rows ────────────────────────────────────────────────────
def copy_cleaned_data(cleaned_path, wb_out, df_raw):
    """Write Data sheet from the cleaned DataFrame."""
    import openpyxl
    wb_src = openpyxl.load_workbook(cleaned_path, read_only=True, data_only=True)
    ws_src = wb_src["Cleaned Data"]
    ws_dst = wb_out.add_worksheet("Data")
    for r_idx, row in enumerate(ws_src.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val is not None:
                ws_dst.write(r_idx, c_idx, val)
    wb_src.close()
    print("      Data copied")


# ── Output filename ───────────────────────────────────────────────────────────
def make_output_name(prev_year, curr_year, curr_months):
    thai_month = curr_months[-1]
    month_num  = THAI_TO_NUM[thai_month]
    greg_year  = curr_year - 543
    prefix     = f"{greg_year}{month_num:02d}"
    end_month_en = thai_month
    return BASE / f"{prefix}_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - {end_month_en} {curr_year}(analyst).xlsx"


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    sys.stdout.reconfigure(encoding="utf-8")

    if not CLEANED_FILE.exists():
        print(f"ERROR: {CLEANED_FILE.name} not found — run /create-model-report first"); sys.exit(1)
    model_file = find_file(MODEL_PATTERN, "Model template")
    print(f"Cleaned : {CLEANED_FILE.name}")
    print(f"Model   : {model_file.name}")

    # ── 1. Load Cleaned Data ──────────────────────────────────────────────────
    print("\n[1/4] Loading Cleaned Data...", flush=True)
    df_raw = pd.read_excel(str(CLEANED_FILE), sheet_name="Cleaned Data", header=5)
    df_raw["จำนวนรถ"] = pd.to_numeric(df_raw["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    df_raw["ปี"]      = pd.to_numeric(df_raw["ปี"],      errors="coerce").dropna().astype(int)
    df_raw = df_raw.dropna(subset=["ปี"]).copy()
    df_raw["ปี"] = df_raw["ปี"].astype(int)
    # Use only File 1 rows (ชนิดเชื้อเพลิง present) and mapped Powertrain types
    KNOWN_PT = {"ICE","BEV","HEV","PHEV"}
    df_file1 = df_raw[df_raw["ชนิดเชื้อเพลิง"].notna() & df_raw["Powertrain"].isin(KNOWN_PT)].copy() \
               if "ชนิดเชื้อเพลิง" in df_raw.columns else df_raw.copy()
    df = filter_ry(df_file1)
    print(f"      {len(df_raw):,} total rows → {len(df_file1):,} File1/mapped → {len(df):,} after รย filter")

    prev_year, curr_year, prev_prev_year, curr_months = _get_year_info(df)
    curr_month = curr_months[-1]
    print(f"      Years: prev={prev_year}, curr={curr_year}, curr_months={curr_months}")

    # ── 2. Load BEV Major from model Data sheet ───────────────────────────────
    print("\n[2/4] Loading BEV Major data...", flush=True)
    df_model_data = pd.read_excel(str(model_file), sheet_name="Data", header=6,
                                   usecols=range(10))
    df_model_data.columns = ["ปี","เดือน","ประเภทรถ","จังหวัด","ยี่ห้อรถ",
                              "ยี่ห้อรถ2","รุ่นรถ","รุ่นรถ2","Powertrain","จำนวนรถ"]
    df_model_data = df_model_data.dropna(subset=["ปี"]).copy()
    df_model_data["ปี"]      = df_model_data["ปี"].astype(int)
    df_model_data["จำนวนรถ"] = pd.to_numeric(df_model_data["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    bev = filter_ry(df_model_data[df_model_data["Powertrain"] == "BEV Major"])
    print(f"      BEV Major rows: {len(bev):,}")

    # ── 3. Build output file ──────────────────────────────────────────────────
    out_file = make_output_name(prev_year, curr_year, curr_months)
    print(f"\n[3/4] Writing {out_file.name}...", flush=True)

    wb  = xlsxwriter.Workbook(str(out_file))
    fmh = wb.add_format({"bold": True, "bg_color": "#4472C4", "font_color": "#FFFFFF",
                          "border": 1, "align": "center", "valign": "vcenter"})

    build_pivot(wb, df, prev_year, curr_year, curr_months, fmh)
    build_reg_powertrain(wb, df, prev_year, curr_year, prev_prev_year, curr_months, fmh)
    build_brand_sheet(wb, df, "2.Rank by Brand",   "Registration by Brand",     "ALL",  prev_year, curr_year, curr_months, fmh, "rank")
    build_brand_sheet(wb, df, "3.ICE by Brand",    "ICE Registration by Brand", "ICE",  prev_year, curr_year, curr_months, fmh, "brand_pt")
    build_brand_sheet(wb, df, "4.BEV by Brand",    "BEV Registration by Brand", "BEV",  prev_year, curr_year, curr_months, fmh, "brand_pt")
    build_brand_sheet(wb, df, "5.HEV by Brand",    "HEV Registration by Brand", "HEV",  prev_year, curr_year, curr_months, fmh, "brand_pt")
    build_brand_sheet(wb, df, "6.PHEV by Brand",   "PHEV Registration by Brand","PHEV", prev_year, curr_year, curr_months, fmh, "brand_pt")
    build_bev_by_model(wb, bev, prev_year, curr_year, curr_months, fmh)
    build_rank_bev_model(wb, bev, prev_year, curr_year, curr_months, fmh)

    # ── 4. Copy reference sheets ──────────────────────────────────────────────
    print("\n[4/4] Copying reference sheets...", flush=True)
    copy_cleaned_data(str(CLEANED_FILE), wb, df_raw)
    copy_sheet_raw(str(model_file), "master powertrain",     wb)
    copy_sheet_raw(str(model_file), "BEV Series Name Table", wb)

    wb.close()
    print(f"\nOutput: {out_file.name}")


def _get_year_info(df):
    years = sorted(df["ปี"].unique())
    curr_year      = int(years[-1])
    prev_year      = int(years[-2]) if len(years) >= 2 else curr_year
    prev_prev_year = int(years[-3]) if len(years) >= 3 else prev_year - 1
    curr_months    = [m for m in MONTH_ORDER if m in df[df["ปี"] == curr_year]["เดือน"].unique()]
    return prev_year, curr_year, prev_prev_year, curr_months


if __name__ == "__main__":
    main()
