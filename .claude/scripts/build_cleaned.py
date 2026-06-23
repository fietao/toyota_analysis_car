#!/usr/bin/env python3
"""
build_cleaned.py — Data Cleaner Agent script.

Steps:
  1. Read reference tables from Model.xlsx (powertrain map, BEV series name table)
  2. Read both raw DLT files (fuel + model)
  3. Add derived columns: ยี่ห้อรถ2, รุ่นรถ2, Powertrain
  4. Write Cleaned Data + master powertrain sheets
  5. Save df_cleaned as pickle for build_pivots.py

Output:
  test_model_1.xlsx  (Cleaned Data, master powertrain)
  test_model_cleaned.pkl  (intermediate for pivot builder)
"""

import glob, sys, os, zipfile, re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")


def read_dlt_file(path: Path) -> pd.DataFrame:
    df = pd.read_excel(str(path), header=5)
    if "จำนวนรถ" in df.columns:
        df["จำนวนรถ"] = pd.to_numeric(df["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    return df
    
BASE = Path(__file__).resolve().parents[2]
RAW1_PATTERN  = str(BASE / "raw data" / "รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564*.xlsx")
RAW2_PATTERN  = str(BASE / "raw data" / "รถใหม่_ยี่ห้อรถ-รุ่นรถ-จังหวัด ปี 2564*.xlsx")
MODEL_PATTERN = str(BASE / "refer" / "*- Model.xlsx")
CALC_PATTERN  = str(BASE / "refer" / "*(calculation).xlsx")

BRAND2_MAP = {
    "GWM TANK":             "GWM",
    "HAVAL":                "GWM",
    "ORA":                  "GWM",
    "GAC":                  "AION",
    "DEEPAL":               "Deepal+ChangAn",
    "CHANGAN":              "ChangAn+Deepal",
    "CHANG AN":             "ChangAn+Deepal",
    "MERCEDES":             "Mercedes-Benz",
    "MERCEDES BENZ":        "Mercedes-Benz",
    "MERCEDES-AMG":         "Mercedes-Benz",
    "MERCEDESBENZ-MAYBACH": "Mercedes-Benz",
    "ZX AUTO":              "ZX Auto",
    "ZXAUTO":               "ZX Auto",
    "STAR8":                "Star8",
    "STAR 8":               "Star8",
    "STAR8-V":              "Star8",
    "ไม่ระบุ":              "ไม่ระบุ",
    "พ่วง/กึ่งพ่วง":        "ไม่ระบุ",
}

BRAND2_TABLE = [
    ("Brand1",  "Brand2"),
    ("GWM",     "GWM"),    ("GWM Tank", "GWM"),  ("Haval",   "GWM"),
    ("ORA",     "GWM"),    ("GAC",      "AION"),  ("Deepal",  "Deepal+ChangAn"),
    ("ChangAn", "ChangAn+Deepal"), ("Mercedes", "Mercedes-Benz"),
    ("ZX Auto", "ZX Auto"), ("Star8",   "Star8"), ("ไม่ระบุ", "ไม่ระบุ"),
]

THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]


FINAL_COLS = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด",
              "ยี่ห้อรถ", "ยี่ห้อรถ2", "รุ่นรถ", "รุ่นรถ2",
              "ชนิดเชื้อเพลิง", "Powertrain", "จำนวนรถ"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_file(pattern, label):
    matches = [m for m in glob.glob(pattern)
               if not any(kw in Path(m).name.lower()
                          for kw in ["~$", "(cleaned data)", "analyst", "test_model", "output"])]
    if not matches:
        print(f"ERROR: No {label} found: {pattern}"); sys.exit(1)
    return Path(max(matches, key=os.path.getmtime))


# what is kwargs 
def read_sheet_raw(path, sheet_name, **kwargs):
    try:
        return pd.read_excel(str(path), sheet_name=sheet_name, header=None, **kwargs)
    except Exception as e:
        print(f"  Warning: could not read '{sheet_name}': {e}")
        return pd.DataFrame()


def read_brand2_rows(wb):
    try:
        ws = wb["Data"]
        rows = []
        for raw_brand, brand2 in ws.iter_rows(
            min_row=1, max_row=8, min_col=14, max_col=15, values_only=True
        ):
            e = str(raw_brand).strip() if raw_brand is not None else ""
            f = str(brand2).strip() if brand2 is not None else ""
            if (
                e
                and e.upper() not in ("NAN", "BRAND1", "ยี่ห้อรถ", "E", "BRAND")
                and f
                and f.upper() not in ("NAN", "BRAND2", "ยี่ห้อรถ2", "F")
            ):
                rows.append((e, f))
        return rows
    except Exception as e:
        print(f"  Warning: could not read brand2 rows: {e}")
        return []


def add_brand2(df, brand_map):
    upper = df["ยี่ห้อรถ"].astype(str).str.strip().str.upper()
    df.insert(df.columns.get_loc("ยี่ห้อรถ") + 1, "ยี่ห้อรถ2",
              upper.map(brand_map).fillna(df["ยี่ห้อรถ"]))


def ordered_cols(df):
    cols  = [c for c in FINAL_COLS if c in df.columns]
    extra = [c for c in df.columns if c not in FINAL_COLS]
    return df[cols + extra]


def write_rows(ws, df, start_row=0):
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row):
        for c_idx, val in enumerate(row):
            if not pd.isna(val):
                ws.write(r_idx, c_idx, val)


def clean_powertrain_value(value):
    if pd.isna(value):
        return "OTH"
    value = str(value).strip()
    return value if value and value.lower() != "nan" and value != "(blank)" else "OTH"


def sort_cleaned_data(df, brand2_order=None):
    out = df.copy()
    month_rank = {month: idx for idx, month in enumerate(MONTH_ORDER, start=1)}
    out["_month_sort"] = out["เดือน"].map(month_rank).fillna(99)
    out["_source_sort"] = out["รุ่นรถ"].notna().astype(int)

    if brand2_order:
        brand_rank = {b: i for i, b in enumerate(brand2_order)}
        out["_brand2_sort"] = out["ยี่ห้อรถ2"].map(brand_rank).fillna(len(brand2_order))
    else:
        out["_brand2_sort"] = 0

    sort_cols = [
        "ปี", "_month_sort", "ประเภทรถ", "จังหวัด",
        "_brand2_sort", "ยี่ห้อรถ2", "ยี่ห้อรถ", "_source_sort",
        "ชนิดเชื้อเพลิง", "รุ่นรถ2", "รุ่นรถ",
    ]
    sort_cols = [col for col in sort_cols if col in out.columns]
    out = out.sort_values(sort_cols, kind="mergesort", na_position="last")
    return out.drop(columns=["_month_sort", "_source_sort", "_brand2_sort"], errors="ignore")



def enable_pivot_refresh(wb):
    """Set refreshOnLoad to True for all pivot tables in the workbook."""
    try:
        for ws in wb.worksheets:
            for pivot in ws._pivots:
                if pivot.cache:
                    pivot.cache.refreshOnLoad = True
    except Exception as e:
        print(f"  Warning: Could not set refreshOnLoad on pivots: {e}")


def _col(n):
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))


def rewrite_data_rows(workbook_path, dataframe, data_start_row):
    """Splice new data rows into Data sheet starting at data_start_row.
    Rows 1..(data_start_row-1) are preserved byte-for-byte (shared strings, styles, brand helper).
    """
    path = Path(workbook_path)

    with zipfile.ZipFile(str(path), "r") as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m = re.search(r'<sheet\b[^>]*\bname="Data"[^>]*\br:id="([^"]+)"', wb_xml)
    if not m:
        m = re.search(r'<sheet\b[^>]*\br:id="([^"]+)"[^>]*\bname="Data"', wb_xml)
    if not m:
        raise ValueError("'Data' sheet not found in xl/workbook.xml")
    rid = m.group(1)
    m2 = re.search(
        rf'<Relationship\b[^>]*\bId="{re.escape(rid)}"[^>]*\bTarget="([^"]+)"', rels_xml
    )
    if not m2:
        raise ValueError(f"No relationship for Id='{rid}'")
    target = m2.group(1)
    sheet_xml_name = f"xl/{target}" if not target.startswith("xl/") else target

    with zipfile.ZipFile(str(path), "r") as z:
        old_xml = z.read(sheet_xml_name).decode("utf-8")

    cut_m = re.search(rf'<row\b[^>]*\br="{data_start_row}"', old_xml)
    prefix = old_xml[:cut_m.start()] if cut_m else old_xml[:old_xml.rfind("</sheetData>")]
    suffix = old_xml[old_xml.rfind("</sheetData>"):]

    rows_xml = []
    for r_idx, row in enumerate(dataframe.itertuples(index=False), data_start_row):
        cells = []
        for c_idx, v in enumerate(row, 1):
            if pd.notna(v):
                col = _col(c_idx)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cells.append(f'<c r="{col}{r_idx}"><v>{v}</v></c>')
                else:
                    cells.append(
                        f'<c r="{col}{r_idx}" t="inlineStr"><is><t>{_esc(v)}</t></is></c>'
                    )
        if cells:
            rows_xml.append(f'<row r="{r_idx}">{"".join(cells)}</row>')

    new_xml = (prefix + "".join(rows_xml) + suffix).encode("utf-8")

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(str(path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == sheet_xml_name:
                zout.writestr(item.filename, new_xml)
            elif "pivotCacheDefinition" in item.filename and item.filename.endswith(".xml"):
                raw = zin.read(item.filename).decode("utf-8")
                patched = re.sub(
                    r'(<pivotCacheDefinition\b)([^>]*?)(/?>)',
                    lambda m: (
                        m.group(1) + m.group(2)
                        + (' refreshOnLoad="1"' if "refreshOnLoad" not in m.group(2) else "")
                        + m.group(3)
                    ),
                    raw, count=1,
                )
                zout.writestr(item.filename, patched.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    tmp.replace(path)
    print(f"      Saved: {path.name} ({len(dataframe):,} rows from row {data_start_row})")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    raw1_file  = find_file(RAW1_PATTERN,  "fuel raw data")
    raw2_file  = find_file(RAW2_PATTERN,  "model raw data")

    # Look for master Model in root first, fallback to refer folder
    model_matches = [p for p in glob.glob(str(BASE / "*master Model.xlsx")) if "~$" not in p]
    if model_matches:
        model_file = Path(max(model_matches, key=os.path.getmtime))
        print(f"Using master Model from root: {model_file.name}")
    else:
        #what is this part for 
        model_file = find_file(MODEL_PATTERN, "Model template")

    # Look for master Cal in root first, fallback to refer folder
    calc_matches = [p for p in glob.glob(str(BASE / "*(master cal).xlsx")) if "~$" not in p]
    if calc_matches:
        calc_file = Path(max(calc_matches, key=os.path.getmtime))
        print(f"Using master Cal from root: {calc_file.name}")
    else:
        calc_file = find_file(CALC_PATTERN, "Calculation template")

    print(f"Raw 1 (Fuel) : {raw1_file.name}")
    print(f"Raw 2 (Model): {raw2_file.name}")
    print(f"Model Temp   : {model_file.name}")
    print(f"Calc Temp    : {calc_file.name}")

    # ── Parquet path (used for incremental check and final save) ──────────────
    pq_path = BASE / "test_model_cleaned.parquet"
    df_existing = None
    existing_key_set = set()

    if pq_path.exists():
        print(f"\n[Incremental] Loading existing: {pq_path.name}")
        df_existing = pd.read_parquet(str(pq_path))
        ex_key_col = df_existing["ปี"].astype(str) + "_" + df_existing["เดือน"].astype(str)
        existing_key_set = set(ex_key_col.unique())
        print(f"  {len(existing_key_set)} year-month pairs | {len(df_existing):,} rows")
    else:
        print("\n[Full build] No existing parquet — rebuilding from scratch.")

    # 1. Reference tables
    print("\n[1/4] Reading reference tables...", flush=True)
    wb_ref = load_workbook(str(model_file), read_only=True, data_only=True)

    # Read master powertrain
    pt_rows = []
    ws_pt = wb_ref["master powertrain"]
    for row in ws_pt.iter_rows(values_only=True):
        pt_rows.append(row)
    df_pt = pd.DataFrame(pt_rows)
    _pt    = df_pt.iloc[7:, [4, 5]].dropna(subset=[4])
    powertrain_map = {str(k).strip(): clean_powertrain_value(v)
                      for k, v in zip(_pt.iloc[:, 0], _pt.iloc[:, 1]) if not pd.isna(k)}
    print(f"      {len(powertrain_map)} powertrain mappings")

    # Read brand → ยี่ห้อรถ2 table from refer1 Data!N:O rows 1-8.
    brand_rows = read_brand2_rows(wb_ref)

    if brand_rows:
        ref_brand2_map = {e.upper(): f for e, f in brand_rows}
        brand2_order   = list(dict.fromkeys(f for _, f in brand_rows))
        brand2_table   = [("Brand1", "Brand2")] + brand_rows
        merged_brand2_map = {**BRAND2_MAP, **ref_brand2_map}
        print(f"      {len(brand_rows)} ยี่ห้อรถ2 entries from refer1 (Data!N:O rows 1-8)")
    else:
        merged_brand2_map = BRAND2_MAP
        brand2_order = list(dict.fromkeys(v for v in BRAND2_MAP.values()))
        brand2_table = list(BRAND2_TABLE)
        print("      ยี่ห้อรถ2: using hardcoded table (Data!N:O rows 1-8 empty)")

    # Read BEV Series Name Table
    bev_rows = []
    ws_bev = wb_ref["BEV Series Name Table"]
    for row in ws_bev.iter_rows(values_only=True):
        bev_rows.append(row)
    df_bev_tbl = pd.DataFrame(bev_rows)
    _bev       = df_bev_tbl.iloc[1:, [1, 2, 3]].dropna(subset=[1])
    model2_map        = {str(k).strip().upper(): str(v).strip() for k, v in zip(_bev.iloc[:, 0], _bev.iloc[:, 1])}
    pt_from_model_map = {str(k).strip().upper(): str(v).strip() for k, v in zip(_bev.iloc[:, 0], _bev.iloc[:, 2])}

    wb_ref.close()

    csv_map_path = BASE / "refer" / "model2_map.csv"
    if csv_map_path.exists():
        df_csv = pd.read_csv(str(csv_map_path), encoding="utf-8-sig")
        csv_model2_map = dict(zip(
            df_csv["รุ่นรถ_raw"].astype(str).str.strip().str.upper(),
            df_csv["รุ่นรถ2"].astype(str).str.strip()
        ))
        model2_map.update(csv_model2_map)
        print(f"      Loaded {len(csv_model2_map)} mappings from model2_map.csv")

    print(f"      {len(model2_map)} รุ่นรถ2 mappings")

    # 2. Raw files
    print("\n[2/4] Reading raw data...", flush=True)
    df_fuel  = read_dlt_file(raw1_file)
    df_model = read_dlt_file(raw2_file)
    print(f"      {len(df_fuel):,} fuel rows | {len(df_model):,} model rows")

    # 3. Derived columns
    print("\n[3/4] Adding derived columns...", flush=True)

    # df_fuel used only for master powertrain summary — not included in cleaned data.
    add_brand2(df_fuel, merged_brand2_map)
    df_fuel.insert(df_fuel.columns.get_loc("ยี่ห้อรถ2") + 1, "รุ่นรถ",  None)
    df_fuel.insert(df_fuel.columns.get_loc("รุ่นรถ")    + 1, "รุ่นรถ2", None)
    df_fuel["Powertrain"] = df_fuel["ชนิดเชื้อเพลิง"].apply(
        lambda f: powertrain_map.get(str(f).strip(), "OTH") if pd.notna(f) else None)

    # Model rows: brand2, model2, powertrain (fuel first, model fallback)
    add_brand2(df_model, merged_brand2_map)
    if "รุ่นรถ" in df_model.columns:
        upper = df_model["รุ่นรถ"].astype(str).str.strip().str.upper()
        df_model.insert(df_model.columns.get_loc("รุ่นรถ") + 1, "รุ่นรถ2",
                        upper.map(model2_map).fillna(df_model["รุ่นรถ"]))

    pt_fuel  = (df_model["ชนิดเชื้อเพลิง"].astype(str).str.strip().map(powertrain_map)
                if "ชนิดเชื้อเพลิง" in df_model.columns
                else pd.Series(dtype=object, index=df_model.index))
    pt_model = (df_model["รุ่นรถ"].astype(str).str.strip().str.upper().map(pt_from_model_map)
                if "รุ่นรถ" in df_model.columns
                else pd.Series(dtype=object, index=df_model.index))
    powertrain_col = pt_fuel.combine_first(pt_model).fillna("OTH")
    if "Powertrain" in df_model.columns:
        df_model["Powertrain"] = powertrain_col
    else:
        df_model.insert(df_model.columns.get_loc("รุ่นรถ2") + 1, "Powertrain", powertrain_col)

    # df_fuel is used above for the master powertrain summary sheet only.
    # Final cleaned data uses df_model only — matches refer file row count (636,333).
    df_cleaned = ordered_cols(df_model)
    df_cleaned = sort_cleaned_data(df_cleaned, brand2_order)
    print(f"      {len(df_cleaned):,} raw rows processed | cols: {list(df_cleaned.columns)}")

    # ── Incremental merge: new months + corrections ────────────────────────────
    if df_existing is not None:
        raw_key = df_cleaned["ปี"].astype(str) + "_" + df_cleaned["เดือน"].astype(str)

        # Only check last year (full Jan-Dec) and current year (Jan-current month)
        current_be = date.today().year + 543
        check_years = {str(current_be - 1), str(current_be)}
        scoped_keys = {k for k in raw_key.unique() if k.split("_", 1)[0] in check_years}
        print(f"  Scope: BE years {sorted(check_years)} | {len(scoped_keys)} months to check")

        new_keys = scoped_keys - existing_key_set
        overlap_keys = scoped_keys & existing_key_set

        corrected_keys = set()
        correction_log = []
        for key in sorted(overlap_keys):
            yr, mo = key.split("_", 1)
            ex_mask  = ex_key_col == key
            raw_mask = raw_key == key
            old_rows  = int(ex_mask.sum())
            new_rows  = int(raw_mask.sum())
            old_total = int(df_existing.loc[ex_mask,  "จำนวนรถ"].sum())
            new_total = int(df_cleaned.loc[raw_mask, "จำนวนรถ"].sum())
            if old_rows != new_rows or old_total != new_total:
                corrected_keys.add(key)
                correction_log.append({"year": yr, "month": mo,
                                       "old_rows": old_rows, "new_rows": new_rows,
                                       "old_total": old_total, "new_total": new_total,
                                       "diff": new_total - old_total})

        if correction_log:
            print(f"\n  *** CORRECTIONS DETECTED ({len(correction_log)} months) ***")
            for c in correction_log:
                row_diff = c["new_rows"] - c["old_rows"]
                row_part = (f"rows {c['old_rows']:,}→{c['new_rows']:,} ({row_diff:+,}),  "
                            if row_diff != 0 else "")
                print(f"      {c['year']} {c['month']}: {row_part}"
                      f"total {c['old_total']:,}→{c['new_total']:,} (diff {c['diff']:+,})")
            import json, datetime
            log_path = BASE / "corrections_log.json"
            existing_log = json.loads(log_path.read_text("utf-8")) if log_path.exists() else []
            existing_log.append({
                "run": str(datetime.date.today()),
                "corrections": [
                    {**c, "row_diff": c["new_rows"] - c["old_rows"]}
                    for c in correction_log
                ]
            })
            log_path.write_text(json.dumps(existing_log, ensure_ascii=False, indent=2), "utf-8")
            print(f"      Logged → {log_path.name}")

        update_keys = new_keys | corrected_keys

        if not update_keys:
            print("  All data up to date. Using existing cleaned data.")
            df_cleaned = df_existing
        else:
            df_keep = df_existing[~ex_key_col.isin(corrected_keys)]
            df_add  = df_cleaned[raw_key.isin(update_keys)]
            df_combined = pd.concat([df_keep, df_add], ignore_index=True)
            df_cleaned  = sort_cleaned_data(df_combined, brand2_order)
            if new_keys:
                print(f"  Added {len(new_keys)} new month(s): {', '.join(sorted(new_keys))}")
            if corrected_keys:
                print(f"  Refreshed {len(corrected_keys)} corrected month(s): "
                      f"{', '.join(sorted(corrected_keys))}")
            print(f"  Final row count: {len(df_cleaned):,}")

    # ── Report new BEV models not in BEV Series Name Table ───────────────────
    if "รุ่นรถ" in df_cleaned.columns and "Powertrain" in df_cleaned.columns:
        bev_rows = df_cleaned[
            df_cleaned["Powertrain"].astype(str).str.startswith("BEV", na=False) &
            df_cleaned["รุ่นรถ"].notna() &
            (df_cleaned["รุ่นรถ"].astype(str).str.strip() != "")
        ][["ยี่ห้อรถ", "รุ่นรถ"]].drop_duplicates()
        known_models = set(model2_map.keys())
        new_bev = bev_rows[
            ~bev_rows["รุ่นรถ"].astype(str).str.strip().str.upper().isin(known_models)
        ]
        if not new_bev.empty:
            print(f"\n*** NEW BEV MODELS not in BEV Series Name Table ({len(new_bev)}) ***")
            for _, row in new_bev.sort_values(["ยี่ห้อรถ", "รุ่นรถ"]).iterrows():
                print(f"  {str(row['ยี่ห้อรถ']):<20}  {row['รุ่นรถ']}")
            print("  → Add these to BEV Series Name Table before next analyst run")

    # Save intermediate for pivot builder (parquet: safe, fast, preserves dtypes)
    # Cast object columns to str to avoid mixed-type ArrowTypeError
    for col in df_cleaned.select_dtypes(include="object").columns:
        df_cleaned[col] = df_cleaned[col].astype(str).replace("nan", pd.NA)
    df_cleaned.to_parquet(str(pq_path), index=False)
    print(f"      Saved intermediate: {pq_path.name}")

    # 4. Write output xlsx (preserve formatting from template)
    out_file = model_file if model_file.parent == BASE else BASE / "test_model_1.xlsx"
    calc_out = calc_file if calc_file.parent == BASE else BASE / "test_calculation.xlsx"
    import shutil
    print(f"\n[4/4] Preparing output files: Model={out_file.name}, Cal={calc_out.name}...", flush=True)
    
    # Copy the templates if they are different paths
    if model_file.resolve() != out_file.resolve():
        shutil.copy2(model_file, out_file)
        print("      Copied Model template.")

    if calc_file.resolve() != calc_out.resolve():
        shutil.copy2(calc_file, calc_out)
        print(f"      Calculation template copied → {calc_out.name}")

    print("      Writing new Data sheet...")

    rewrite_data_rows(out_file, df_cleaned, data_start_row=8)

    df_calc = df_cleaned.drop(columns=["Powertrain"], errors="ignore")
    rewrite_data_rows(calc_out, df_calc, data_start_row=7)

    print(f"\nOutput : {out_file}")
    print(f"         {calc_out.name}")
    print(f"  Rows : {len(df_cleaned):,}")
    print("  → Data sheet replaced. All original template sheets, pivots, and charts are preserved.")


if __name__ == "__main__":
    main()
