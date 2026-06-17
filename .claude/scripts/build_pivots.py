#!/usr/bin/env python3
"""
build_pivots.py — Analyst Agent script.

Reads the cleaned data pickle from build_cleaned.py and adds BEV/BMW
pivot sheets to test_model_1.xlsx using openpyxl (append mode).

Steps:
  1. Load df_cleaned from test_model_cleaned.pkl
  2. Filter BEV Major rows and BMW rows
  3. Build pivot sheets: BEV by Model, BEV by Model (2), BMW
  4. Append sheets to test_model_1.xlsx
  5. Delete the pickle

Run AFTER build_cleaned.py.
"""

import sys
from pathlib import Path

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE     = Path(__file__).resolve().parents[2]
PKL_PATH = BASE / "test_model_cleaned.parquet"
OUT_PATH = BASE / "test_model_1.xlsx"

THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]


# ── Pivot helpers ─────────────────────────────────────────────────────────────

def _ym_structure(df, last_n_years=None):
    all_years = sorted(df["ปี"].unique())
    if last_n_years:
        all_years = all_years[-last_n_years:]
    months_by_yr = df.groupby("ปี")["เดือน"].apply(set)
    return {yr: [m for m in MONTH_ORDER if m in months_by_yr.get(yr, set())]
            for yr in all_years}


def _col_specs(ym, bmw_style=False):
    specs, years = [], sorted(ym.keys())
    max_yr = years[-1] if years else None
    for yr in years:
        if bmw_style and yr != max_yr:
            specs.append((yr, "ytotal"))
        else:
            for m in ym[yr]:
                specs.append((yr, m))
            specs.append((yr, "ytotal"))
    specs.append(("grand", "grand"))
    return specs


def _pivot_wide(df, index_cols):
    return df.pivot_table(index=index_cols, columns=["ปี", "เดือน"],
                          values="จำนวนรถ", aggfunc="sum", fill_value=0)


def _row_values(row_series, specs):
    counts = row_series.to_dict()
    grand  = int(row_series.sum())
    result = []
    for yr, ctype in specs:
        if yr == "grand":
            result.append(grand or None)
        elif ctype == "ytotal":
            s = int(sum(counts.get((yr, m), 0) for m in MONTH_ORDER))
            result.append(s or None)
        else:
            v = int(counts.get((yr, ctype), 0))
            result.append(v or None)
    return result


def _write_vals(ws, row, vals, start):
    for c, v in enumerate(vals, start=start):
        if v is not None:
            ws.write(row, c, v)


def _write_col_headers(ws, ym, specs, start_col, fmt_h):
    col, last_yr = start_col, None
    for yr, ctype in specs:
        if yr == "grand":
            ws.write(5, col, "Grand Total", fmt_h)
        elif ctype == "ytotal":
            label = yr if yr != last_yr else f"{yr} Total"
            ws.write(5, col, label, fmt_h)
        else:
            if yr != last_yr:
                ws.write(5, col, yr, fmt_h)
            ws.write(6, col, ctype, fmt_h)
        last_yr = yr
        col += 1


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_bev_by_model(workbook, bev, fmt_h):
    ym    = _ym_structure(bev, last_n_years=2)
    bev   = bev[bev["ปี"].isin(ym.keys())]
    specs = _col_specs(ym)

    ws = workbook.add_worksheet("BEV by Model")
    ws.write(0, 0, "ประเภทรถ");      ws.write(0, 1, "(Multiple Items)")
    ws.write(1, 0, "จังหวัด");       ws.write(1, 1, "(All)")
    ws.write(2, 0, "Powertrain");    ws.write(2, 1, "BEV Major")
    ws.write(4, 0, "Sum of จำนวนรถ"); ws.write(4, 1, "Column Labels")
    ws.write(6, 0, "Row Labels", fmt_h)
    _write_col_headers(ws, ym, specs, 1, fmt_h)

    model_piv   = _pivot_wide(bev, ["ยี่ห้อรถ2", "รุ่นรถ2"])
    brand_piv   = model_piv.groupby(level="ยี่ห้อรถ2").sum()
    brand_order = brand_piv.sum(axis=1).sort_values(ascending=False).index.tolist()

    data_row = 7
    for brand in brand_order:
        ws.write(data_row, 0, brand, fmt_h)
        _write_vals(ws, data_row, _row_values(brand_piv.loc[brand], specs), 1)
        data_row += 1
        models_df   = model_piv.loc[[brand]]
        model_order = models_df.sum(axis=1).sort_values(ascending=False).index.tolist()
        for (_, model) in model_order:
            ws.write(data_row, 0, model)
            _write_vals(ws, data_row, _row_values(models_df.loc[(brand, model)], specs), 1)
            data_row += 1
    print("      BEV by Model done")


def _flat_pivot(workbook, df, sheet_name, index_cols, label_names, pt_label, bmw_style, fmt_h):
    ym    = _ym_structure(df, last_n_years=2)
    df    = df[df["ปี"].isin(ym.keys())]
    specs = _col_specs(ym, bmw_style=bmw_style)
    start = len(label_names)

    ws = workbook.add_worksheet(sheet_name)
    ws.write(0, 0, "ประเภทรถ");   ws.write(0, 1, "(Multiple Items)")
    ws.write(1, 0, "จังหวัด");    ws.write(1, 1, "(All)")
    ws.write(2, 0, "Powertrain"); ws.write(2, 1, pt_label)
    ws.write(4, 0, "Sum of จำนวนรถ")
    for i, name in enumerate(label_names):
        ws.write(6, i, name, fmt_h)
    _write_col_headers(ws, ym, specs, start, fmt_h)

    piv       = _pivot_wide(df, index_cols)
    row_order = piv.sum(axis=1).sort_values(ascending=False).index.tolist()

    data_row = 7
    for key in row_order:
        key_vals = key if isinstance(key, tuple) else (key,)
        for i, v in enumerate(key_vals):
            ws.write(data_row, i, v)
        _write_vals(ws, data_row, _row_values(piv.loc[key], specs), start)
        data_row += 1
    print(f"      {sheet_name} done")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not PKL_PATH.exists():
        print(f"ERROR: {PKL_PATH.name} not found — run build_cleaned.py first")
        sys.exit(1)
    if not OUT_PATH.exists():
        print(f"ERROR: {OUT_PATH.name} not found — run build_cleaned.py first")
        sys.exit(1)

    print(f"Loading {PKL_PATH.name}...", flush=True)
    df_cleaned = pd.read_parquet(str(PKL_PATH))
    print(f"  {len(df_cleaned):,} rows loaded")

    # Model rows only (fuel rows have รุ่นรถ = None)
    df_model = df_cleaned[df_cleaned["รุ่นรถ"].notna()].copy()
    bev_data = df_model[df_model["Powertrain"] == "BEV Major"]
    bmw_data = df_model[df_model["ยี่ห้อรถ2"] == "BMW"]
    print(f"  BEV Major rows: {len(bev_data):,}")
    print(f"  BMW rows      : {len(bmw_data):,}")

    # Write pivot sheets to a temp file, then copy into the main xlsx
    tmp_path = BASE / "_pivots_tmp.xlsx"
    print(f"\nBuilding pivot sheets...", flush=True)

    workbook = xlsxwriter.Workbook(str(tmp_path))
    fmt_h = workbook.add_format({"bold": True, "bg_color": "#4472C4",
                                  "font_color": "#FFFFFF", "border": 1,
                                  "align": "center", "valign": "vcenter"})

    build_bev_by_model(workbook, bev_data, fmt_h)
    _flat_pivot(workbook, bev_data, "BEV by Model (2)",
                ["รุ่นรถ2", "ยี่ห้อรถ2"], ["รุ่นรถ2", "ยี่ห้อรถ2"], "BEV Major", False, fmt_h)
    _flat_pivot(workbook, bmw_data, "BMW",
                ["ยี่ห้อรถ2", "รุ่นรถ2"], ["ยี่ห้อรถ2", "รุ่นรถ2"], "(All)", True, fmt_h)
    workbook.close()

    # Append pivot sheets into the main xlsx using openpyxl
    print(f"\nAppending pivot sheets to {OUT_PATH.name}...", flush=True)
    wb_main = load_workbook(str(OUT_PATH))
    wb_tmp  = load_workbook(str(tmp_path))

    for sheet_name in wb_tmp.sheetnames:
        ws_src = wb_tmp[sheet_name]
        ws_dst = wb_main.create_sheet(sheet_name)
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row))

    wb_main.save(str(OUT_PATH))
    wb_main.close()
    wb_tmp.close()

    # Cleanup
    tmp_path.unlink(missing_ok=True)
    PKL_PATH.unlink(missing_ok=True)
    print(f"  Cleaned up temp files")

    print(f"\nOutput: {OUT_PATH}")
    print(f"  Sheets: Cleaned Data | master powertrain | BEV Series Name Table | BEV by Model | BEV by Model (2) | BMW")
    print("Done.")


if __name__ == "__main__":
    main()
