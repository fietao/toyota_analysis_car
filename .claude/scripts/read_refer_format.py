#!/usr/bin/env python3
"""
read_refer_format.py
Extract the format blueprint from the reference Model.xlsx.
Outputs a format spec: header colors, column widths, sort order,
table structure, frozen panes, autofilter — sheet by sheet.
Used by the analyze-therefer skill so agents know exactly what to match.
"""

import sys
import glob
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(__file__).resolve().parents[2]


def find_ref():
    files = [f for f in glob.glob(str(BASE / "refer/*Model.xlsx")) if "~$" not in f]
    if not files:
        print("ERROR: reference Model.xlsx not found in refer/", file=sys.stderr)
        sys.exit(1)
    return Path(sorted(files)[-1])


def rgb(color):
    if color and color.type == "rgb" and color.rgb not in ("00000000", "FFFFFFFF"):
        return f"#{color.rgb[2:]}"   # strip alpha → #RRGGBB
    return None


def cell_format(cell):
    info = {}
    if cell.font:
        if cell.font.bold:              info["bold"] = True
        c = rgb(cell.font.color)
        if c:                           info["font_color"] = c
        if cell.font.size:              info["font_size"] = cell.font.size
    if cell.fill and cell.fill.fgColor:
        c = rgb(cell.fill.fgColor)
        if c:                           info["bg_color"] = c
    if cell.alignment:
        if cell.alignment.horizontal:   info["align"] = cell.alignment.horizontal
    if cell.number_format and cell.number_format != "General":
        info["number_format"] = cell.number_format
    return info


def main():
    ref_file = find_ref()
    print(f"REFERENCE FILE: {ref_file.name}")
    print(f"{'='*60}")

    wb = load_workbook(str(ref_file), read_only=False, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSHEET: {sheet_name}")
        print(f"  Dimensions  : {ws.max_row} rows × {ws.max_column} cols")

        # Frozen panes
        if ws.freeze_panes:
            print(f"  Frozen panes: {ws.freeze_panes}")

        # Autofilter
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"  Autofilter  : {ws.auto_filter.ref}")

        # Column widths
        col_widths = {}
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                col_widths[col_letter] = round(dim.width, 1)
        if col_widths:
            print(f"  Col widths  : {col_widths}")

        # Header row (row 1 or find first non-empty row)
        header_row_idx = None
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
            if any(cell.value is not None for cell in row):
                header_row_idx = row[0].row
                break

        if header_row_idx:
            print(f"  Header row  : {header_row_idx}")
            header_cells = list(ws.iter_rows(
                min_row=header_row_idx, max_row=header_row_idx))[0]
            headers = []
            for cell in header_cells:
                fmt = cell_format(cell)
                entry = {"col": get_column_letter(cell.column),
                         "value": cell.value, "format": fmt}
                headers.append(entry)
            for h in headers:
                if h["value"] is not None:
                    print(f"    {h['col']:>3}  {str(h['value']):<30}  {h['format']}")

        # Sample data rows (first 3 data rows after header)
        data_start = (header_row_idx or 1) + 1
        data_rows = list(ws.iter_rows(
            min_row=data_start,
            max_row=min(data_start + 2, ws.max_row),
            values_only=True))
        if data_rows:
            print(f"  Sample data :")
            for row in data_rows:
                print(f"    {list(row)}")

        # Data row format (check first data row for cell styles)
        if header_row_idx and ws.max_row > header_row_idx:
            first_data = list(ws.iter_rows(
                min_row=data_start, max_row=data_start))[0]
            row_fmt = {}
            for cell in first_data:
                fmt = cell_format(cell)
                if fmt:
                    row_fmt[get_column_letter(cell.column)] = fmt
            if row_fmt:
                print(f"  Data row fmt: {row_fmt}")

    wb.close()
    print(f"\n{'='*60}")
    print("[FORMAT BLUEPRINT DONE]")


if __name__ == "__main__":
    main()
