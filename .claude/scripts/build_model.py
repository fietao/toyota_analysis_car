"""
Builds the monthly model output file from raw data.

Steps:
  1. Read reference tables from model file (master powertrain, BEV Series Name Table)
  2. Pull both raw files -> concat into one table
  3. Add derived columns: ยี่ห้อรถ2, รุ่นรถ2, Powertrain -> organize column order
  4. Write Cleaned Data + static sheets
  5. Build BEV pivot sheets (BEV by Model, BEV by Model (2), BMW)

Output: test_model_1.xlsx in project root
"""

import glob, sys, os
from datetime import date
from pathlib import Path
import pandas as pd
import xlsxwriter

# ── Project root ──────────────────────────────────────────────────────────────
def _find_project_root():
    p = Path(__file__).resolve()
    for parent in p.parents:
        if (parent / "CLAUDE.md").exists():
            return parent
    raise RuntimeError(f"Could not find project root (no CLAUDE.md above {p})")

BASE = _find_project_root()
RAW1_PATTERN  = str(BASE / "raw data" / "รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564*.xlsx")
RAW2_PATTERN  = str(BASE / "raw data" / "รถใหม่_ยี่ห้อรถ-รุ่นรถ-จังหวัด ปี 2564*.xlsx")
MODEL_PATTERN = str(BASE / "refer" / "*- Model.xlsx")   # hyphen required → avoids lowercase model.xlsx

# ── Brand2 mapping ────────────────────────────────────────────────────────────
BRAND2_MAP = {
    "GWM TANK":             "GWM",
    "HAVAL":                "GWM",
    "ORA":                  "GWM",
    "GAC":                  "AION",
    "DEEPAL":               "Deepal+ChangAn",
    "CHANGAN":              "ChangAn+Deepal",
    "CHANG AN":             "ChangAn+Deepal",
    "MERCEDES":             "Mercedes-Benz",
    "MERCEDES BENZ":        "Mercedes-Benz",
    "MERCEDES-AMG":         "Mercedes-Benz",
    "MERCEDESBENZ-MAYBACH": "Mercedes-Benz",
    "ZX AUTO":              "ZX Auto",
    "ZXAUTO":               "ZX Auto",
    "STAR8":                "Star8",
    "STAR 8":               "Star8",
    "STAR8-V":              "Star8",
    "ไม่ระบุ":              "ไม่ระบุ",
    "พ่วง/กึ่งพ่วง":        "ไม่ระบุ",
}


BRAND2_TABLE = [
    ("Brand1",    "Brand2"),
    ("GWM",       "GWM"),
    ("GWM Tank",  "GWM"),
    ("Haval",     "GWM"),
    ("ORA",       "GWM"),
    ("GAC",       "AION"),
    ("Deepal",    "Deepal+ChangAn"),
    ("ChangAn",   "ChangAn+Deepal"),
    ("Mercedes",  "Mercedes-Benz"),
    ("ZX Auto",   "ZX Auto"),
    ("Star8",     "Star8"),
    ("ไม่ระบุ",   "ไม่ระบุ"),
]

THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]

MASTER_POWERTRAIN_ORDER = [
    ("CNG", "ICE"),
    ("CNG-LPG", "ICE"),
    ("CNG-ดีเซล", "ICE"),
    ("CNG-เบนซิน", "ICE"),
    ("LPG-ดีเซล-ไฟฟ้าแบบเสียบปลั๊ก", "PHEV"),
    ("LPG-เบนซิน-ไฟฟ้า", "HEV"),
    ("LPGและดีเซล", "ICE"),
    ("LPGและเบนซิน", "ICE"),
    ("ก๊าซ LPG", "ICE"),
    ("ดีเซล", "ICE"),
    ("ดีเซล-ไฟฟ้า", "HEV"),
    ("ดีเซล-ไฟฟ้าแบบเสียบปลั๊ก", "PHEV"),
    ("เบนซิน", "ICE"),
    ("เบนซิน-ไฟฟ้า", "HEV"),
    ("เบนซิน-ไฟฟ้าแบบเสียบปลั๊ก", "PHEV"),
    ("ไฟฟ้า", "BEV"),
    ("ไม่ใช้เชื้อเพลิง", None),
    (None, None),
    ("ไฮโดรเจน", "ICE"),
]

FINAL_COLS = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด",
              "ยี่ห้อรถ", "ยี่ห้อรถ2",
              "รุ่นรถ", "รุ่นรถ2",
              "ชนิดเชื้อเพลิง", "Powertrain",
              "จำนวนรถ"]


# ── File helpers ──────────────────────────────────────────────────────────────

def find_file(pattern, label):
    matches = glob.glob(pattern)
    # Ignore lock files and output files
    exclude_keywords = ["~$", "(cleaned data)", "(analyst)", "test_model", "output"]
    matches = [m for m in matches if not any(kw in Path(m).name.lower() for kw in exclude_keywords)]
    if not matches:
        print(f"ERROR: No {label} file found: {pattern}"); sys.exit(1)
    return Path(max(matches, key=os.path.getmtime))


def find_all_files(pattern, label):
    matches = glob.glob(pattern)
    matches = [Path(m) for m in matches if not Path(m).name.startswith("~$")]
    if not matches:
        print(f"ERROR: No {label} file found: {pattern}"); sys.exit(1)
    return sorted(matches, key=os.path.getmtime)


def read_sheet_raw(model_file, sheet_name):
    try:
        return pd.read_excel(model_file, sheet_name=sheet_name, header=None)
    except Exception as e:
        print(f"  Warning: could not read sheet '{sheet_name}': {e}")
        return pd.DataFrame()


def write_rows(ws, df, start_row=0):
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row):
        for c_idx, val in enumerate(row):
            if pd.isna(val):
                continue
            ws.write(r_idx, c_idx, val)


def add_brand2(df):
    brand_upper = df["ยี่ห้อรถ"].astype(str).str.strip().str.upper()
    df.insert(df.columns.get_loc("ยี่ห้อรถ") + 1, "ยี่ห้อรถ2",
              brand_upper.map(BRAND2_MAP).fillna(df["ยี่ห้อรถ"]))


def ordered_cols(df):
    cols = [c for c in FINAL_COLS if c in df.columns]
    extra = [c for c in df.columns if c not in FINAL_COLS]
    return df[cols + extra]


def compare_reference_structure(out_file, model_file):
    import openpyxl

    wb_out = openpyxl.load_workbook(out_file, read_only=True, data_only=True)
    wb_ref = openpyxl.load_workbook(model_file, read_only=True, data_only=True)
    print("\n── Reference compare ───────────────────────────────────────")
    print(f"  Generated sheets: {wb_out.sheetnames}")
    print(f"  Reference sheets: {wb_ref.sheetnames}")

    for out_name, ref_name in [
        ("Cleaned Data", "Data"),
        ("master powertrain", "master powertrain"),
        ("BEV Series Name Table", "BEV Series Name Table"),
        ("BEV by Model", "BEV by Model"),
        ("BEV by Model (2)", "BEV by Model (2)"),
        ("BMW", "BMW"),
    ]:
        if out_name not in wb_out.sheetnames or ref_name not in wb_ref.sheetnames:
            print(f"  Missing compare sheet: generated={out_name}, reference={ref_name}")
            continue
        ws_out = wb_out[out_name]
        ws_ref = wb_ref[ref_name]
        print(f"  {out_name} vs {ref_name}: {ws_out.max_row}x{ws_out.max_column} vs {ws_ref.max_row}x{ws_ref.max_column}")

    wb_out.close()
    wb_ref.close()


# ── Data sheet reader ─────────────────────────────────────────────────────────
def read_data_sheet(model_file):
    """Read the Data sheet (brand+model DLT source). Header is at row index 6."""
    df = pd.read_excel(str(model_file), sheet_name="Data", header=6,
                       usecols=range(10))
    df.columns = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด", "ยี่ห้อรถ",
                  "ยี่ห้อรถ2", "รุ่นรถ", "รุ่นรถ2", "Powertrain", "จำนวนรถ"]
    df = df.dropna(subset=["ปี"]).copy()
    df["ปี"] = df["ปี"].astype(int)
    df["จำนวนรถ"] = pd.to_numeric(df["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    return df


# ── Pivot helpers ─────────────────────────────────────────────────────────────
def _ym_structure(df, last_n_years=None):
    """Return {year: [ordered months present in df]}, optionally capped to last N years."""
    all_years = sorted(df["ปี"].unique())
    if last_n_years:
        all_years = all_years[-last_n_years:]
    months_by_yr = df.groupby("ปี")["เดือน"].apply(set)
    return {yr: [m for m in MONTH_ORDER if m in months_by_yr.get(yr, set())]
            for yr in all_years}


def _col_specs(ym, bmw_style=False):
    """Return list of (year, col_type) where col_type is a month name, 'ytotal', or 'grand'."""
    specs = []
    years = sorted(ym.keys())
    max_yr = years[-1] if years else None
    for yr in years:
        if bmw_style and yr != max_yr:
            specs.append((yr, "ytotal"))
        else:
            for m in ym[yr]:
                specs.append((yr, m))
            specs.append((yr, "ytotal"))
    specs.append(("grand", "grand"))
    return specs


def _pivot_wide(df, index_cols):
    """Pivot df to wide format: index=index_cols, columns=(ปี, เดือน)."""
    return df.pivot_table(
        index=index_cols,
        columns=["ปี", "เดือน"],
        values="จำนวนรถ",
        aggfunc="sum",
        fill_value=0,
    )


def _row_values(row_series, specs):
    """Convert a pivot row Series to a list of cell values matching specs."""
    counts = row_series.to_dict()
    grand_total = int(row_series.sum())
    result = []
    for yr, ctype in specs:
        if yr == "grand":
            result.append(grand_total if grand_total else None)
        elif ctype == "ytotal":
            yr_sum = int(sum(counts.get((yr, m), 0) for m in MONTH_ORDER))
            result.append(yr_sum if yr_sum else None)
        else:
            val = int(counts.get((yr, ctype), 0))
            result.append(val if val else None)
    return result


def _write_vals(ws, row, vals, start):
    """Write a list of cell values to a worksheet row, skipping None."""
    for c, v in enumerate(vals, start=start):
        if v is not None:
            ws.write(row, c, v)


def _write_col_headers(ws, ym, specs, data_start_col, fmt_h):
    """Write year header (row 5) and month header (row 6) for a pivot sheet."""
    col = data_start_col
    last_yr_written = None
    for yr, ctype in specs:
        if yr == "grand":
            ws.write(5, col, "Grand Total", fmt_h)
        elif ctype == "ytotal":
            if yr != last_yr_written:
                ws.write(5, col, yr, fmt_h)
                last_yr_written = yr
            else:
                ws.write(5, col, f"{yr} Total", fmt_h)
        else:
            if yr != last_yr_written:
                ws.write(5, col, yr, fmt_h)
                last_yr_written = yr
            ws.write(6, col, ctype, fmt_h)
        col += 1


# ── Master powertrain sheet builder ──────────────────────────────────────────
def build_master_powertrain_sheet(workbook, df, df_template, fmt_h, fmt_bold_blue, fmt_warn, powertrain_map):
    ws = workbook.add_worksheet("master powertrain")

    fuel_totals = (df.groupby("ชนิดเชื้อเพลิง", dropna=False)["จำนวนรถ"]
                   .sum().to_dict())
    active_fuels = {
        str(fuel).strip()
        for fuel, total in fuel_totals.items()
        if pd.notna(fuel) and total > 0
    }

    # Cols E-F: copy reference mapping first; active rows are marked after A:C is built from it.
    for r_idx, tup in enumerate(df_template.itertuples(index=False, name=None)):
        val_e = tup[4] if len(tup) > 4 else None
        val_f = tup[5] if len(tup) > 5 else None
        is_active = not pd.isna(val_e) and str(val_e).strip() in active_fuels

        if r_idx == 6:
            fmt = fmt_h
        elif is_active:
            fmt = fmt_bold_blue
        else:
            fmt = None

        if not pd.isna(val_e):
            ws.write(r_idx, 4, val_e, fmt)
        if not pd.isna(val_f):
            ws.write(r_idx, 5, val_f, fmt)

    ws.write(5, 0, "Sum of จำนวนรถ", fmt_h)
    ws.write(6, 0, "ชนิดเชื้อเพลิง", fmt_h)
    ws.write(6, 1, "Powertrain", fmt_h)
    ws.write(6, 2, "Total", fmt_h)

    r = 7
    written_fuels = set()
    for fuel, pt in MASTER_POWERTRAIN_ORDER:
        total = int(fuel_totals.get(fuel, 0))
        if fuel is not None:
            ws.write(r, 0, fuel)
            written_fuels.add(fuel)
        if pt is not None:
            ws.write(r, 1, pt)
        ws.write(r, 2, total)
        r += 1

    extras = []
    for fuel, total in fuel_totals.items():
        if pd.isna(fuel):
            continue
        fuel_key = str(fuel).strip()
        if fuel_key not in written_fuels:
            extras.append((fuel_key, powertrain_map.get(fuel_key, "Other"), int(total)))
    for fuel, pt, total in sorted(extras, key=lambda item: item[0]):
        ws.write(r, 0, fuel, fmt_warn)
        if pt is not None:
            ws.write(r, 1, pt, fmt_warn)
        ws.write(r, 2, total, fmt_warn)
        r += 1

    ws.autofilter(6, 0, r - 1, 2)
    ws.write(r, 0, "Grand Total")
    ws.write(r, 2, int(df["จำนวนรถ"].sum()))

    if extras:
        print(f"      master powertrain extra fuel types: {[fuel for fuel, _, _ in extras]}")
    print("      master powertrain done")


# ── BEV sheet builders ────────────────────────────────────────────────────────
def build_bev_by_model(workbook, bev, fmt_h):
    """BEV by Model: grouped by Brand2 (parent) then รุ่นรถ2 (child). bev pre-filtered."""
    ym = _ym_structure(bev, last_n_years=2)
    bev = bev[bev["ปี"].isin(ym.keys())]
    specs = _col_specs(ym)
    data_start = 1

    ws = workbook.add_worksheet("BEV by Model")
    ws.write(0, 0, "ประเภทรถ");   ws.write(0, 1, "(Multiple Items)")
    ws.write(1, 0, "จังหวัด");    ws.write(1, 1, "(All)")
    ws.write(2, 0, "Powertrain"); ws.write(2, 1, "BEV Major")
    ws.write(4, 0, "Sum of จำนวนรถ"); ws.write(4, 1, "Column Labels")
    ws.write(6, 0, "Row Labels", fmt_h)
    _write_col_headers(ws, ym, specs, data_start, fmt_h)

    model_piv = _pivot_wide(bev, ["ยี่ห้อรถ2", "รุ่นรถ2"])
    brand_piv = model_piv.groupby(level="ยี่ห้อรถ2").sum()
    brand_order = brand_piv.sum(axis=1).sort_values(ascending=False).index.tolist()

    data_row = 7
    for brand in brand_order:
        ws.write(data_row, 0, brand, fmt_h)
        _write_vals(ws, data_row, _row_values(brand_piv.loc[brand], specs), data_start)
        data_row += 1

        models_df = model_piv.loc[[brand]]
        model_order = models_df.sum(axis=1).sort_values(ascending=False).index.tolist()
        for (_, model) in model_order:
            ws.write(data_row, 0, model)
            _write_vals(ws, data_row, _row_values(models_df.loc[(brand, model)], specs), data_start)
            data_row += 1

    print("      BEV by Model done")


def _build_flat_pivot_sheet(workbook, df, sheet_name, index_cols, label_names,
                             powertrain_label, bmw_style, fmt_h):
    """Build a flat (non-grouped) pivot sheet. df must be pre-filtered."""
    ym = _ym_structure(df, last_n_years=2)
    df = df[df["ปี"].isin(ym.keys())]
    specs = _col_specs(ym, bmw_style=bmw_style)
    data_start = len(label_names)

    ws = workbook.add_worksheet(sheet_name)
    ws.write(0, 0, "ประเภทรถ");   ws.write(0, 1, "(Multiple Items)")
    ws.write(1, 0, "จังหวัด");    ws.write(1, 1, "(All)")
    ws.write(2, 0, "Powertrain"); ws.write(2, 1, powertrain_label)
    ws.write(4, 0, "Sum of จำนวนรถ")
    ws.write(4, data_start, "ปี"); ws.write(4, data_start + 1, "เดือน")
    for i, name in enumerate(label_names):
        ws.write(6, i, name, fmt_h)
    _write_col_headers(ws, ym, specs, data_start, fmt_h)

    piv = _pivot_wide(df, index_cols)
    row_order = piv.sum(axis=1).sort_values(ascending=False).index.tolist()

    data_row = 7
    for key in row_order:
        key_vals = key if isinstance(key, tuple) else (key,)
        for i, v in enumerate(key_vals):
            ws.write(data_row, i, v)
        _write_vals(ws, data_row, _row_values(piv.loc[key], specs), data_start)
        data_row += 1

    print(f"      {sheet_name} done")


def build_bev_by_model_2(workbook, bev, fmt_h):
    _build_flat_pivot_sheet(workbook, bev, "BEV by Model (2)",
                            ["รุ่นรถ2", "ยี่ห้อรถ2"], ["รุ่นรถ2", "ยี่ห้อรถ2"],
                            "BEV Major", bmw_style=False, fmt_h=fmt_h)


def build_bmw(workbook, bmw, fmt_h):
    _build_flat_pivot_sheet(workbook, bmw, "BMW",
                            ["ยี่ห้อรถ2", "รุ่นรถ2"], ["ยี่ห้อรถ2", "รุ่นรถ2"],
                            "(All)", bmw_style=True, fmt_h=fmt_h)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    sys.stdout.reconfigure(encoding="utf-8")

    raw_file   = find_file(RAW2_PATTERN,  "model raw data")  # Model file uses RAW2 (model-level raw data)
    raw1_file  = find_file(RAW1_PATTERN,  "fuel raw data")
    model_file = find_file(MODEL_PATTERN, "Model template")

    print(f"Raw 1 (Fuel) : {raw1_file.name}")
    print(f"Raw 2 (Model): {raw_file.name}")
    print(f"Model Temp   : {model_file.name}")

    # ── 1. Read reference tables ──────────────────────────────────────────────
    print("\n[1/5] Reading reference tables...", flush=True)

    df_powertrain = read_sheet_raw(model_file, "master powertrain")
    _pt = df_powertrain.iloc[7:, [4, 5]].dropna(subset=[4])
    powertrain_map = {}
    for k, v in zip(_pt.iloc[:, 0], _pt.iloc[:, 1]):
        if pd.isna(k):
            continue
        key = str(k).strip()
        val = str(v).strip() if pd.notna(v) else None
        powertrain_map[key] = val
    print(f"      {len(powertrain_map)} powertrain mappings (master powertrain)")

    df_bev_table = read_sheet_raw(model_file, "BEV Series Name Table")
    _bev = df_bev_table.iloc[1:, [1, 2, 3]].dropna(subset=[1])
    model2_map        = {str(k).strip().upper(): str(v).strip() for k, v in zip(_bev.iloc[:, 0], _bev.iloc[:, 1])}
    pt_from_model_map = {str(k).strip().upper(): str(v).strip() for k, v in zip(_bev.iloc[:, 0], _bev.iloc[:, 2])}
    print(f"      {len(model2_map)} รุ่นรถ2 mappings (BEV Series Name Table)")

    # ── 2. Pull raw files ──────────────────────────────────────────────────────
    print("\n[2/5] Reading raw data...", flush=True)
    df = pd.read_excel(str(raw_file), header=5)
    print(f"      {len(df):,} model rows loaded")

    df_fuel_raw = pd.read_excel(str(raw1_file), header=5)
    df_fuel_raw["จำนวนรถ"] = pd.to_numeric(df_fuel_raw["จำนวนรถ"], errors="coerce").fillna(0).astype(int)

    def get_fuel_pt(fuel_val):
        if pd.isna(fuel_val):
            return None
        fuel_str = str(fuel_val).strip()
        if fuel_str in powertrain_map:
            return powertrain_map[fuel_str]
        return "Other"

    add_brand2(df_fuel_raw)
    df_fuel_raw.insert(df_fuel_raw.columns.get_loc("ยี่ห้อรถ2") + 1, "รุ่นรถ", None)
    df_fuel_raw.insert(df_fuel_raw.columns.get_loc("รุ่นรถ") + 1, "รุ่นรถ2", None)
    df_fuel_raw["Powertrain"] = df_fuel_raw["ชนิดเชื้อเพลิง"].apply(get_fuel_pt)
    print(f"      {len(df_fuel_raw):,} fuel rows loaded")

    # ── 3. Add derived columns → organize ────────────────────────────────────
    print("\n[3/5] Adding derived columns...", flush=True)

    # ยี่ห้อรถ2
    add_brand2(df)

    # รุ่นรถ2 (only when รุ่นรถ column exists)
    if "รุ่นรถ" in df.columns:
        model_upper = df["รุ่นรถ"].astype(str).str.strip().str.upper()
        df.insert(df.columns.get_loc("รุ่นรถ") + 1, "รุ่นรถ2",
                  model_upper.map(model2_map).fillna(df["รุ่นรถ"]))

    # Powertrain: ชนิดเชื้อเพลิง-based (File 1) takes priority; รุ่นรถ-based (File 2) fills the rest
    pt_fuel  = (df["ชนิดเชื้อเพลิง"].astype(str).str.strip().map(powertrain_map)
                if "ชนิดเชื้อเพลิง" in df.columns
                else pd.Series(dtype=object, index=df.index))
    pt_model = (df["รุ่นรถ"].astype(str).str.strip().str.upper().map(pt_from_model_map)
                if "รุ่นรถ" in df.columns
                else pd.Series(dtype=object, index=df.index))
    powertrain = pt_fuel.combine_first(pt_model).fillna("Other")

    if "ชนิดเชื้อเพลิง" in df.columns:
        df.insert(df.columns.get_loc("ชนิดเชื้อเพลิง") + 1, "Powertrain", powertrain)
    else:
        df["Powertrain"] = powertrain

    df.insert(df.columns.get_loc("รุ่นรถ2") + 1, "ชนิดเชื้อเพลิง", None)
    df = ordered_cols(df)
    df_fuel_raw = ordered_cols(df_fuel_raw)
    df_cleaned = pd.concat([df_fuel_raw, df], ignore_index=True)
    print(f"      Columns: {list(df_cleaned.columns)}")

    # ── Filter subsets for BEV pivot builders ────────────────────────────────
    bev_data = df[df["Powertrain"] == "BEV Major"]
    bmw_data = df[df["ยี่ห้อรถ2"] == "BMW"]
    print(f"      {len(df_cleaned):,} combined rows | years: {sorted(df_cleaned['ปี'].unique())}")
    print(f"      BEV Major rows: {len(bev_data):,}")

    # ── 4. Build output file ──────────────────────────────────────────────────
    out_file = BASE / "test_model_1.xlsx"
    print(f"\n[4/5] Writing {out_file.name}...", flush=True)

    workbook = xlsxwriter.Workbook(str(out_file))

    fmt_header = workbook.add_format({
        "bold": True, "bg_color": "#4472C4", "font_color": "#FFFFFF",
        "border": 1, "align": "center", "valign": "vcenter",
    })
    fmt_title = workbook.add_format({"bold": True, "font_size": 12})
    fmt_bold_blue = workbook.add_format({
        "bold": True,
        "font_color": "#4472C4"
    })
    fmt_warn = workbook.add_format({
        "bold": True,
        "bg_color": "#FFF2CC",
        "font_color": "#9C6500"
    })

    # ── Cleaned Data sheet ────────────────────────────────────────────────────
    ws = workbook.add_worksheet("Cleaned Data")
    DATA_ROW = 5

    max_yr_raw = int(df_cleaned["ปี"].max())
    last_yr_rows = df_cleaned[df_cleaned["ปี"] == max_yr_raw]
    end_month = [m for m in MONTH_ORDER if m in last_yr_rows["เดือน"].unique()][-1]
    end_year  = max_yr_raw

    today = date.today()
    proc_month = THAI_MONTHS.get(today.month, "")
    proc_year  = today.year + 543

    ws.write(0, 0, "สถิติการจดทะเบียนรถใหม่ ตามกฎหมายว่าด้วยรถยนต์ จำแนกตามยี่ห้อรถ ชนิดเชื้อเพลิง และจังหวัด", fmt_title)
    ws.write(1, 0, f"เดือนมกราคม ปี พ.ศ. 2564 - เดือน{end_month} ปี พ.ศ. {end_year}")
    ws.write(2, 0, "หน่วย: คัน")
    ws.write(3, 0, f"ประมวลผลข้อมูล วันที่ {today.day} เดือน{proc_month} ปี พ.ศ. {proc_year}")
    ws.write(4, 0, "หมายเหตุ: นับเฉพาะรถใหม่ ไม่รวมรถที่ใช้แล้วนำกลับมาจดทะเบียนใหม่")

    for r, (b1, b2) in enumerate(BRAND2_TABLE):
        fmt = fmt_header if r == 0 else None
        ws.write(r, 13, b1, fmt)
        ws.write(r, 14, b2, fmt)

    total = len(df_cleaned)
    CHUNK = 10000
    for i in range(0, total, CHUNK):
        write_rows(ws, df_cleaned.iloc[i:i+CHUNK], start_row=DATA_ROW + 1 + i)
        print(f"      Written {min(i+CHUNK, total):,} / {total:,} rows...", flush=True)

    ws.add_table(DATA_ROW, 0, DATA_ROW + total, len(df_cleaned.columns) - 1, {
        'style': 'Table Style Medium 2',
        'columns': [{'header': col, 'header_format': fmt_header} for col in df_cleaned.columns],
    })
    col_widths = {
        "ปี": 8, "เดือน": 12, "ประเภทรถ": 35, "จังหวัด": 20,
        "ยี่ห้อรถ": 20, "ยี่ห้อรถ2": 20,
        "รุ่นรถ": 28, "รุ่นรถ2": 25,
        "ชนิดเชื้อเพลิง": 25, "Powertrain": 15, "จำนวนรถ": 12,
    }
    for i, col in enumerate(df_cleaned.columns):
        if col in col_widths:
            ws.set_column(i, i, col_widths[col])
    ws.set_column(13, 14, 18)  # BRAND2_TABLE reference
    print("      Cleaned Data done")

    # ── Static / generated sheets ─────────────────────────────────────────────
    build_master_powertrain_sheet(workbook, df_fuel_raw, df_powertrain, fmt_header, fmt_bold_blue, fmt_warn, powertrain_map)

    ws_bev = workbook.add_worksheet("BEV Series Name Table")
    if not df_bev_table.empty:
        write_rows(ws_bev, df_bev_table)
    print("      BEV Series Name Table done")

    # ── 5. BEV pivot sheets ───────────────────────────────────────────────────
    print("\n[5/5] Building BEV pivot sheets...", flush=True)
    build_bev_by_model(workbook, bev_data, fmt_header)
    build_bev_by_model_2(workbook, bev_data, fmt_header)
    build_bmw(workbook, bmw_data, fmt_header)

    workbook.close()
    print(f"\nOutput: {out_file}")
    compare_reference_structure(out_file, model_file)

    # ── End-of-run report ─────────────────────────────────────────────────────
    print("\n── Report ───────────────────────────────────────────────────")
    print(f"  Raw 1 file used: {raw1_file.name}")
    print(f"  Raw 2 file used: {raw_file.name}")
    print(f"  Total rows     : {len(df_cleaned):,}")

    if "รุ่นรถ" in df.columns:
        unmapped_models = df["รุ่นรถ"].dropna().str.strip()
        unmapped_models = unmapped_models[~unmapped_models.isin(model2_map)].unique()
        if len(unmapped_models):
            print(f"  รุ่นรถ not in BEV Series Name Table: {len(unmapped_models)} unique")
            print(f"    First 5: {list(unmapped_models[:5])}")
        else:
            print("  รุ่นรถ: all mapped ✓")

    if "ชนิดเชื้อเพลิง" in df.columns:
        other_fuels = df[df["Powertrain"] == "Other"]["ชนิดเชื้อเพลิง"].dropna().unique()
        if len(other_fuels):
            print(f"  ชนิดเชื้อเพลิง → 'Other': {list(other_fuels)}")
        else:
            print("  ชนิดเชื้อเพลิง: all mapped ✓")

    print("Done.")


if __name__ == "__main__":
    main()
