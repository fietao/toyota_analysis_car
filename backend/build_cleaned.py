#!/usr/bin/env python3
"""
build_cleaned.py — Data Cleaner Agent script.

Steps:
  1. Read both raw DLT files (fuel + model)
  2. Append raw model names missing from the existing master mapping sheet
  3. Read reference tables from Model.xlsx plus repo mapping CSVs
  4. Add derived columns: ยี่ห้อรถ2, รุ่นรถ2, Powertrain, include_in_bev_model_report
  5. Update Cleaned Data in the existing master workbook in place
  6. Save df_cleaned as parquet for downstream scripts

Output:
  existing masterModel.xlsx updated in place
  test_model_cleaned.parquet  (intermediate for downstream scripts)
"""

import csv, glob, sys, os, zipfile, re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from schema import validate_model, validate_fuel
from series_registry import canonical_series_map, normalize_key, verified_powertrain_map

sys.stdout.reconfigure(encoding="utf-8")


def read_dlt_file(path: Path) -> pd.DataFrame:
    df = pd.read_excel(str(path), header=5, engine="calamine")
    if "จำนวนรถ" in df.columns:
        df["จำนวนรถ"] = pd.to_numeric(df["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    return df

BASE = Path(__file__).resolve().parent
OUTPUT_BASE = Path(os.environ["BUILD_CLEANED_OUTPUT_DIR"]) if os.environ.get("BUILD_CLEANED_OUTPUT_DIR") else BASE
RAW1_PATTERN  = str(BASE / "raw data" / "**" / "รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564*.xlsx")
RAW2_PATTERN  = str(BASE / "raw data" / "**" / "รถใหม่_ยี่ห้อรถ-รุ่นรถ-จังหวัด ปี 2564*.xlsx")





THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]


FINAL_COLS = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด",
              "ยี่ห้อรถ", "ยี่ห้อรถ2", "รุ่นรถ", "รุ่นรถ2",
              "ชนิดเชื้อเพลิง", "Powertrain", "จำนวนรถ"]


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_file(pattern, label):
    override = os.environ.get(f"{label.upper().replace(' ', '_')}_PATH")
    if override:
        path = Path(override)
        if not path.exists():
            print(f"ERROR: {label} override not found: {path}"); sys.exit(1)
        return path

    raw_root = (BASE / "raw data").resolve()
    matches = [m for m in glob.glob(pattern, recursive=True)
               if Path(m).parent.resolve() == raw_root
               and not any(kw in Path(m).name.lower()
                           for kw in ["~$", "(cleaned data)", "analyst", "test_model", "output", "fake test"])]
    if not matches:
        print(f"ERROR: No {label} found: {pattern}"); sys.exit(1)
    return Path(max(matches, key=os.path.getmtime))


# what is kwargs
def read_sheet_raw(path, sheet_name, **kwargs):
    try:
        return pd.read_excel(str(path), sheet_name=sheet_name, header=None, engine="calamine", **kwargs)
    except Exception as e:
        print(f"  Warning: could not read '{sheet_name}': {e}")
        return pd.DataFrame()


def read_brand2_rows(wb):
    try:
        ws = wb["Data"]
        rows = []
        for raw_brand, brand2 in ws.iter_rows(
            min_row=1, max_row=8, min_col=14, max_col=15, values_only=True
        ):
            e = str(raw_brand).strip() if raw_brand is not None else ""
            f = str(brand2).strip() if brand2 is not None else ""
            if (
                e
                and e.upper() not in ("NAN", "BRAND1", "ยี่ห้อรถ", "E", "BRAND")
                and f
                and f.upper() not in ("NAN", "BRAND2", "ยี่ห้อรถ2", "F")
            ):
                rows.append((e, f))
        return rows
    except Exception as e:
        print(f"  Warning: could not read brand2 rows: {e}")
        return []


def add_brand2(df, brand_map):
    upper = df["ยี่ห้อรถ"].astype(str).str.strip().str.upper()
    df.insert(df.columns.get_loc("ยี่ห้อรถ") + 1, "ยี่ห้อรถ2",
              upper.map(brand_map).fillna(df["ยี่ห้อรถ"]))


def ordered_cols(df):
    cols  = [c for c in FINAL_COLS if c in df.columns]
    extra = [c for c in df.columns if c not in FINAL_COLS]
    return df[cols + extra]


def write_rows(ws, df, start_row=0):
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row):
        for c_idx, val in enumerate(row):
            if not pd.isna(val):
                ws.write(r_idx, c_idx, val)


def clean_powertrain_value(value):
    if pd.isna(value):
        return "OTH"
    value = str(value).strip()
    return value if value and value.lower() != "nan" and value != "(blank)" else "OTH"


def dump_map_to_csv(mapping: dict, path: str, key_field="fuel", value_field="powertrain"):
    with open(path, mode="w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([key_field, value_field])
        for key in sorted(mapping.keys()):
            writer.writerow([key, mapping[key]])


def load_powertrain_map(path: str, key_field="fuel", value_field="powertrain") -> dict:
    powertrain_map = {}
    with open(path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            key = row[key_field].strip()
            value = row[value_field].strip()
            powertrain_map[key] = value
    return powertrain_map


def sort_cleaned_data(df, brand2_order=None):
    out = df.copy()
    month_rank = {month: idx for idx, month in enumerate(MONTH_ORDER, start=1)}
    out["_month_sort"] = out["เดือน"].map(month_rank).fillna(99)
    out["_source_sort"] = out["รุ่นรถ"].notna().astype(int)

    if brand2_order:
        brand_rank = {b: i for i, b in enumerate(brand2_order)}
        out["_brand2_sort"] = out["ยี่ห้อรถ2"].map(brand_rank).fillna(len(brand2_order))
    else:
        out["_brand2_sort"] = 0

    sort_cols = [
        "ปี", "_month_sort", "ประเภทรถ", "จังหวัด",
        "_brand2_sort", "ยี่ห้อรถ2", "ยี่ห้อรถ", "_source_sort",
        "Powertrain", "รุ่นรถ2", "รุ่นรถ",
    ]
    sort_cols = [col for col in sort_cols if col in out.columns]
    out = out.sort_values(sort_cols, kind="mergesort", na_position="last")
    return out.drop(columns=["_month_sort", "_source_sort", "_brand2_sort"], errors="ignore")



def enable_pivot_refresh(wb):
    """Set refreshOnLoad to True for all pivot tables in the workbook."""
    try:
        for ws in wb.worksheets:
            for pivot in ws._pivots:
                if pivot.cache:
                    pivot.cache.refreshOnLoad = True
    except Exception as e:
        print(f"  Warning: Could not set refreshOnLoad on pivots: {e}")


def _col(n):
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))


def _header_key(value):
    return re.sub(r"[\s_\-().]+", "", str(value or "").strip().lower())


def _find_header(headers, aliases):
    wanted = {_header_key(alias) for alias in aliases}
    for idx, header in enumerate(headers, start=1):
        if _header_key(header) in wanted:
            return idx
    return None


def _sheet_xml_path(workbook_path, sheet_name):
    with zipfile.ZipFile(str(workbook_path), "r") as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    rid_m = (
        re.search(rf'<sheet\b[^>]*\bname="{re.escape(sheet_name)}"[^>]*\br:id="([^"]+)"', wb_xml)
        or re.search(rf'<sheet\b[^>]*\br:id="([^"]+)"[^>]*\bname="{re.escape(sheet_name)}"', wb_xml)
    )
    if not rid_m:
        return None

    rel_m = re.search(rf'<Relationship\b(?=[^>]*\bId="{re.escape(rid_m.group(1))}")[^>]*/?>', rels_xml)
    target_m = re.search(r'\bTarget="([^"]+)"', rel_m.group(0)) if rel_m else None
    if not target_m:
        return None
    target = target_m.group(1).lstrip("/")
    return f"xl/{target}" if not target.startswith("xl/") else target


def _normalise_model_key(value):
    return str(value).strip().upper() if pd.notna(value) else ""


def _suggest_missing_model_rows(df_model):
    brand_map = load_powertrain_map(str(BASE / "config" / "brand_map.csv"), "brand", "brand2")
    brand_map = {str(k).strip().upper(): str(v).strip() for k, v in brand_map.items()}
    powertrain_map = load_powertrain_map(str(BASE / "config" / "powertrain_map.csv"))
    series_name_map = canonical_series_map()

    required = {"ยี่ห้อรถ", "รุ่นรถ"}
    if not required.issubset(df_model.columns):
        return []

    work = df_model.copy()
    work["_raw_model"] = work["รุ่นรถ"].astype(str).str.strip()
    work["_model_key"] = work["_raw_model"].str.upper()
    work = work[work["_raw_model"] != ""]
    if work.empty:
        return []

    work["_brand"] = (
        work["ยี่ห้อรถ"].astype(str).str.strip().str.upper()
        .map(brand_map)
        .fillna(work["ยี่ห้อรถ"].astype(str).str.strip())
    )
    if "ชนิดเชื้อเพลิง" in work.columns:
        work["_powertrain"] = work["ชนิดเชื้อเพลิง"].map(powertrain_map).map(clean_powertrain_value)
    else:
        work["_powertrain"] = "OTH"
    work["_units"] = pd.to_numeric(work.get("จำนวนรถ", 1), errors="coerce").fillna(1)
    work["_order"] = range(len(work))

    # Keep the first-seen model order for append order, but choose suggestions from
    # the highest-volume brand/powertrain tuple for that raw model.
    ranked = (
        work.groupby(["_model_key", "_raw_model", "_brand", "_powertrain"], dropna=False)
        .agg(_units=("_units", "sum"), _order=("_order", "min"))
        .reset_index()
        .sort_values(["_model_key", "_units", "_order"], ascending=[True, False, True])
        .drop_duplicates(subset=["_model_key"], keep="first")
    )
    first_order = work.groupby("_model_key")["_order"].min().to_dict()
    ranked["_first_order"] = ranked["_model_key"].map(first_order)
    ranked = ranked.sort_values("_first_order", kind="mergesort")

    rows = []
    for _, row in ranked.iterrows():
        raw_model = str(row["_raw_model"]).strip()
        brand = str(row["_brand"]).strip() or "N/A"
        registry_key = normalize_key(brand, raw_model)
        rows.append({
            "Brand": brand,
            "รุ่นรถ": raw_model,
            "รุ่นรถ2": series_name_map.get(registry_key, raw_model),
            "Powertrain": clean_powertrain_value(row["_powertrain"]),
        })
    return rows


def append_missing_master_model_rows(workbook_path: Path, df_model: pd.DataFrame) -> int:
    """Append raw models missing from the workbook mapping sheet, without rewriting it.

    The current workbook uses "BEV Series Name Table" as the raw model mapping sheet.
    Future templates may add optional Index/Status columns; when present, they are
    populated, but existing rows/cells are never edited.
    """
    sheet_name = "BEV Series Name Table"
    path = Path(workbook_path)
    wb = load_workbook(str(path), read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            print(f"      {sheet_name}: not found — skipping missing-model append")
            return 0
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        cols = {
            "index": _find_header(headers, ["Index", "Idx", "#", "No", "No.", "ลำดับ"]),
            "brand": _find_header(headers, ["Brand", "ยี่ห้อรถ2", "ยี่ห้อรถ"]),
            "raw_model": _find_header(headers, ["รุ่นรถ", "raw model", "model", "รุ่นรถ_raw"]),
            "canonical": _find_header(headers, ["รุ่นรถ2", "canonical model", "model2"]),
            "powertrain": _find_header(headers, ["Powertrain", "powertrain suggestion"]),
            "status": _find_header(headers, ["Status", "สถานะ", "Review Status"]),
        }
        required_cols = ["brand", "raw_model", "canonical", "powertrain"]
        missing_cols = [name for name in required_cols if not cols[name]]
        if missing_cols:
            print(f"      {sheet_name}: missing column(s) {', '.join(missing_cols)} — skipping")
            return 0

        existing_models = set()
        last_data_row = 1
        max_index = 0
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=False):
            raw_value = row[cols["raw_model"] - 1].value
            raw_key = _normalise_model_key(raw_value)
            if raw_key:
                existing_models.add(raw_key)
                last_data_row = max(last_data_row, row[0].row)
            if cols["index"]:
                idx_value = row[cols["index"] - 1].value
                try:
                    max_index = max(max_index, int(idx_value))
                except (TypeError, ValueError):
                    pass
    finally:
        wb.close()

    suggestions = [
        row for row in _suggest_missing_model_rows(df_model)
        if _normalise_model_key(row["รุ่นรถ"]) not in existing_models
    ]
    if not suggestions:
        print(f"      {sheet_name}: no new raw model names")
        return 0

    sheet_path = _sheet_xml_path(path, sheet_name)
    if not sheet_path:
        print(f"      {sheet_name}: could not resolve sheet XML — skipping")
        return 0

    with zipfile.ZipFile(str(path), "r") as z:
        old_xml = z.read(sheet_path).decode("utf-8")

    style_by_col = {}
    last_row_m = re.search(rf'<row\b[^>]*\br="{last_data_row}"[^>]*>(.*?)</row>', old_xml, flags=re.S)
    if last_row_m:
        for cell_m in re.finditer(r'<c\b([^>]*)\br="([A-Z]+)\d+"([^>]*)>', last_row_m.group(1)):
            attrs = cell_m.group(1) + cell_m.group(3)
            style_m = re.search(r'\bs="([^"]+)"', attrs)
            if style_m:
                style_by_col[cell_m.group(2)] = style_m.group(1)

    def cell_xml(col_idx, row_idx, value):
        col = _col(col_idx)
        style = f' s="{style_by_col[col]}"' if col in style_by_col else ""
        value = "" if value is None else str(value).strip()
        return f'<c r="{col}{row_idx}"{style} t="inlineStr"><is><t>{_esc(value)}</t></is></c>'

    rows_xml = []
    for offset, row in enumerate(suggestions, start=1):
        row_idx = last_data_row + offset
        values = {}
        if cols["index"]:
            values[cols["index"]] = max_index + offset
        values[cols["brand"]] = row["Brand"]
        values[cols["raw_model"]] = row["รุ่นรถ"]
        values[cols["canonical"]] = row["รุ่นรถ2"]
        values[cols["powertrain"]] = row["Powertrain"]
        if cols["status"]:
            values[cols["status"]] = "NEW"

        cells = "".join(cell_xml(col_idx, row_idx, values[col_idx]) for col_idx in sorted(values))
        rows_xml.append(f'<row r="{row_idx}">{cells}</row>')

    new_last_row = last_data_row + len(suggestions)
    new_xml = old_xml.replace("</sheetData>", "".join(rows_xml) + "</sheetData>", 1)
    last_col = _col(len(headers))
    if re.search(r'<dimension\b[^>]*\bref="[^"]+"', new_xml):
        new_xml = re.sub(r'(<dimension\b[^>]*\bref=")[^"]+"', rf'\g<1>A1:{last_col}{new_last_row}"', new_xml, count=1)
    if re.search(r'<autoFilter\b[^>]*\bref="[^"]+"', new_xml):
        new_xml = re.sub(r'(<autoFilter\b[^>]*\bref=")[^"]+"', rf'\g<1>A1:{last_col}{new_last_row}"', new_xml, count=1)

    tmp = path.with_suffix(".tmp-model-map.xlsx")
    with zipfile.ZipFile(str(path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = new_xml.encode("utf-8") if item.filename == sheet_path else zin.read(item.filename)
            zout.writestr(item, data)
    tmp.replace(path)
    print(f"      {sheet_name}: appended {len(suggestions)} new raw model row(s)")
    return len(suggestions)


def _field_indices(xml, tag, index_attr):
    m = re.search(rf'<{tag}\s+count="\d+">(.*?)</{tag}>', xml, re.S)
    if not m:
        return set()
    return {int(v) for v in re.findall(rf'{index_attr}="(\d+)"', m.group(1))}


def _remap_pivot_field_list(xml, tag, index_attr, remap):
    """Remap a rowFields|colFields|pageFields|dataFields block's index attribute
    (x= or fld=) through `remap` (old cache-field index -> new index, or None
    to drop). Any element referencing a dropped index — a stale ชนิดเชื้อเพลิง
    field, possibly duplicated many times by a prior non-idempotent writer run
    — is pruned and the block's count attribute is adjusted to match."""
    m = re.search(rf'<{tag}\s+count="\d+">(.*?)</{tag}>', xml, re.S)
    if not m:
        return xml
    elements = re.findall(r'<[^>]+/>', m.group(1))
    kept = []
    for el in elements:
        idx_m = re.search(rf'\b{index_attr}="(\d+)"', el)
        if not idx_m:
            kept.append(el)
            continue
        new_idx = remap.get(int(idx_m.group(1)))
        if new_idx is None:
            continue
        kept.append(re.sub(rf'\b{index_attr}="\d+"', f'{index_attr}="{new_idx}"', el, count=1))
    new_block = f'<{tag} count="{len(kept)}">{"".join(kept)}</{tag}>' if kept else ""
    return xml[:m.start()] + new_block + xml[m.end():]


def _patch_pivot_table_def1(xml, remap, new_count):
    """Rebuild a model-grain pivot table (pivotTable1/2/3, cache def1) so its
    pivotFields match the current (ชนิดเชื้อเพลิง-free) model schema.

    Row/col/page/data field lists are remapped and pruned first via `remap`;
    pivotFields is then rebuilt from scratch sized to `new_count`, assigning
    each field's axis from where it landed in the remapped field lists. No
    cached <items> are carried over — refreshOnLoad (set by the caller)
    repopulates them from live data.
    """
    xml = _remap_pivot_field_list(xml, "rowFields", "x", remap)
    xml = _remap_pivot_field_list(xml, "colFields", "x", remap)
    xml = _remap_pivot_field_list(xml, "pageFields", "fld", remap)
    xml = _remap_pivot_field_list(xml, "dataFields", "fld", remap)

    row_idx = _field_indices(xml, "rowFields", "x")
    col_idx = _field_indices(xml, "colFields", "x")
    page_idx = _field_indices(xml, "pageFields", "fld")
    data_idx = _field_indices(xml, "dataFields", "fld")

    fields = []
    for i in range(new_count):
        if i in data_idx:
            fields.append('<pivotField dataField="1" compact="0" outline="0" numFmtId="3" showAll="0"/>')
        elif i in row_idx:
            fields.append('<pivotField axis="axisRow" compact="0" outline="0" showAll="0"/>')
        elif i in col_idx:
            fields.append('<pivotField axis="axisCol" compact="0" outline="0" showAll="0"/>')
        elif i in page_idx:
            fields.append('<pivotField axis="axisPage" compact="0" outline="0" showAll="0"/>')
        else:
            fields.append('<pivotField compact="0" outline="0" showAll="0"/>')
    new_pf = f'<pivotFields count="{new_count}">' + "".join(fields) + '</pivotFields>'

    start = xml.find("<pivotFields")
    end   = xml.find("</pivotFields>") + len("</pivotFields>")
    if start >= 0 and end > start:
        xml = xml[:start] + new_pf + xml[end:]
    return xml


def _rebuild_table_columns(raw, new_cols, header_row, n_data_rows):
    """Regenerate <tableColumns> to exactly match new_cols.

    Drops any <calculatedColumnFormula> (ยี่ห้อรถ2/รุ่นรถ2/Powertrain previously
    recomputed live via XLOOKUP against the untrusted legacy BEV Series Name
    Table / brand-map helper range). The pipeline now writes each row's
    literal registry-derived value directly, so the Table must carry those
    literals rather than recompute over them on every open.
    """
    cols_xml = "".join(
        f'<tableColumn id="{i}" name="{_esc(name)}"/>'
        for i, name in enumerate(new_cols, 1)
    )
    new_block = f'<tableColumns count="{len(new_cols)}">{cols_xml}</tableColumns>'
    raw = re.sub(r'<tableColumns\s+count="\d+">.*?</tableColumns>', new_block, raw, count=1, flags=re.S)

    last_col = _col(len(new_cols))
    last_row = header_row + n_data_rows
    raw = re.sub(
        r'(ref=")([^"]+)(")',
        rf'\g<1>A{header_row}:{last_col}{last_row}\g<3>',
        raw,
    )
    return raw


def _rebuild_cache_fields(raw, new_cols):
    """Regenerate <cacheFields> (pivotCacheDefinition1, model-grain) to exactly
    match new_cols — self-heals any number of stale duplicate fields left by
    a prior non-idempotent insertion."""
    fields_xml = "".join(
        f'<cacheField name="{_esc(name)}" numFmtId="0"><sharedItems/></cacheField>'
        for name in new_cols
    )
    new_block = f'<cacheFields count="{len(new_cols)}">{fields_xml}</cacheFields>'
    return re.sub(r'<cacheFields\s+count="\d+">.*?</cacheFields>', new_block, raw, count=1, flags=re.S)


def rewrite_data_rows(workbook_path, dataframe, data_start_row):
    """Splice new data rows into Data sheet starting at data_start_row.
    Rows 1..(data_start_row-2) are preserved byte-for-byte (metadata, brand helper).
    The header row immediately above data_start_row is refreshed from dataframe columns.
    """
    path = Path(workbook_path)

    with zipfile.ZipFile(str(path), "r") as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m = re.search(r'<sheet\b[^>]*\bname="Data"[^>]*\br:id="([^"]+)"', wb_xml)
    if not m:
        m = re.search(r'<sheet\b[^>]*\br:id="([^"]+)"[^>]*\bname="Data"', wb_xml)
    if not m:
        raise ValueError("'Data' sheet not found in xl/workbook.xml")
    rid = m.group(1)
    rel_m = re.search(rf'<Relationship\b(?=[^>]*\bId="{re.escape(rid)}")[^>]*/?>', rels_xml)
    m2 = re.search(r'\bTarget="([^"]+)"', rel_m.group(0)) if rel_m else None
    if not m2:
        raise ValueError(f"No relationship for Id='{rid}'")
    target = m2.group(1).lstrip("/")
    sheet_xml_name = f"xl/{target}" if not target.startswith("xl/") else target

    with zipfile.ZipFile(str(path), "r") as z:
        old_xml = z.read(sheet_xml_name).decode("utf-8")

    header_row = data_start_row - 1
    cut_m = re.search(rf'<row\b[^>]*\br="{header_row}"', old_xml)
    if not cut_m:
        cut_m = re.search(rf'<row\b[^>]*\br="{data_start_row}"', old_xml)
    prefix = old_xml[:cut_m.start()] if cut_m else old_xml[:old_xml.rfind("</sheetData>")]
    suffix = old_xml[old_xml.rfind("</sheetData>"):]

    rows_xml = []
    header_cells = []
    for c_idx, col_name in enumerate(dataframe.columns, 1):
        col = _col(c_idx)
        header_cells.append(
            f'<c r="{col}{header_row}" t="inlineStr"><is><t>{_esc(col_name)}</t></is></c>'
        )
    rows_xml.append(f'<row r="{header_row}">{"".join(header_cells)}</row>')

    for r_idx, row in enumerate(dataframe.itertuples(index=False), data_start_row):
        cells = []
        for c_idx, v in enumerate(row, 1):
            if pd.notna(v):
                col = _col(c_idx)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cells.append(f'<c r="{col}{r_idx}"><v>{v}</v></c>')
                else:
                    cells.append(
                        f'<c r="{col}{r_idx}" t="inlineStr"><is><t>{_esc(v)}</t></is></c>'
                    )
        if cells:
            rows_xml.append(f'<row r="{r_idx}">{"".join(cells)}</row>')

    new_xml = (prefix + "".join(rows_xml) + suffix).encode("utf-8")

    # Build pivot-table → cache-definition mapping. Only cache def1 and its
    # pivot tables are model-grain (bound to the Data sheet we just rewrote);
    # cache def2 / pivotTable4 ("master powertrain") is a fuel-grain view —
    # ชนิดเชื้อเพลิง is legitimately still part of its premise, and rebuilding it
    # from fuel-source data is a separate, later slice. It is left untouched.
    pt_to_cache: dict[str, int] = {}
    with zipfile.ZipFile(str(path), "r") as _zr:
        for _name in _zr.namelist():
            if (_name.startswith("xl/pivotTables/") and _name.endswith(".xml")
                    and "_rels" not in _name):
                _rels = _name.replace("xl/pivotTables/", "xl/pivotTables/_rels/") + ".rels"
                if _rels in _zr.namelist():
                    _rd = _zr.read(_rels).decode("utf-8")
                    _cm = re.search(r"pivotCacheDefinition(\d+)\.xml", _rd)
                    pt_to_cache[_name] = int(_cm.group(1)) if _cm else 1

    # Model-schema cache field remap: old cacheField NAME order -> new column
    # position. Any old field no longer present in the new schema (ชนิดเชื้อเพลิง,
    # dropped from model facts — and possibly duplicated many times over by a
    # prior non-idempotent writer run) maps to None and is pruned everywhere
    # it's referenced (pivotFields, row/col/page/dataFields).
    new_cols = list(dataframe.columns)
    with zipfile.ZipFile(str(path), "r") as _zr:
        old_cache1_xml = _zr.read("xl/pivotCache/pivotCacheDefinition1.xml").decode("utf-8")
    old_names = re.findall(r'<cacheField\s+name="([^"]*)"', old_cache1_xml)
    remap: dict[int, int | None] = {
        old_idx: (new_cols.index(name) if name in new_cols else None)
        for old_idx, name in enumerate(old_names)
    }

    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(str(path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == sheet_xml_name:
                zout.writestr(item.filename, new_xml)
            elif item.filename == "xl/workbook.xml":
                raw = zin.read(item.filename).decode("utf-8")
                raw = re.sub(
                    r'<calcPr\b[^>]*/>',
                    '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>',
                    raw,
                    count=1,
                )
                zout.writestr(item.filename, raw.encode("utf-8"))
            elif item.filename == "xl/calcChain.xml":
                pass  # stale after row rewrite; Excel regenerates on open
            elif item.filename.startswith("xl/tables/") and item.filename.endswith(".xml"):
                raw = zin.read(item.filename).decode("utf-8")
                if re.search(r'\bref="A' + str(header_row) + r':', raw):
                    raw = _rebuild_table_columns(raw, new_cols, header_row, len(dataframe))
                zout.writestr(item.filename, raw.encode("utf-8"))
            elif item.filename == "xl/pivotCache/pivotCacheDefinition1.xml":
                raw = _rebuild_cache_fields(old_cache1_xml, new_cols)
                raw = re.sub(
                    r'(<pivotCacheDefinition\b)([^>]*?)(/?>)',
                    lambda m: (
                        m.group(1) + m.group(2)
                        + (' refreshOnLoad="1"' if "refreshOnLoad" not in m.group(2) else "")
                        + m.group(3)
                    ),
                    raw, count=1,
                )
                zout.writestr(item.filename, raw.encode("utf-8"))
            elif (item.filename.startswith("xl/pivotTables/") and item.filename.endswith(".xml")
                  and "_rels" not in item.filename
                  and pt_to_cache.get(item.filename, 1) == 1):
                raw = zin.read(item.filename).decode("utf-8")
                raw = _patch_pivot_table_def1(raw, remap, len(new_cols))
                zout.writestr(item.filename, raw.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    tmp.replace(path)
    print(f"      Saved: {path.name} ({len(dataframe):,} rows from row {data_start_row})")


def _add_summary_sheet(workbook_path, sheet_name, df_cleaned):
    """Add a new sheet with two side-by-side tables from df_cleaned:
    Cols A-C: ชนิดเชื้อเพลิง | Powertrain | Total (all fuel types)
    Cols E-I: Brand | รุ่นรถ | รุ่นรถ2 | Powertrain | Total (BEV series only)
    """
    path = Path(workbook_path)

    tbl_a = (
        df_cleaned
        .groupby(["ชนิดเชื้อเพลิง", "Powertrain"], dropna=False)["จำนวนรถ"]
        .sum()
        .reset_index()
        .sort_values(["Powertrain", "ชนิดเชื้อเพลิง"])
        .reset_index(drop=True)
    )

    bev_mask = df_cleaned["ชนิดเชื้อเพลิง"].astype(str).str.strip() == "ไฟฟ้า"
    tbl_b = (
        df_cleaned[bev_mask]
        .groupby(["ยี่ห้อรถ2", "รุ่นรถ", "รุ่นรถ2", "Powertrain"], dropna=False)["จำนวนรถ"]
        .sum()
        .reset_index()
        .sort_values(["ยี่ห้อรถ2", "รุ่นรถ2", "รุ่นรถ"])
        .reset_index(drop=True)
    )

    A_HEADERS = ["ชนิดเชื้อเพลิง", "Powertrain", "Total"]
    B_HEADERS = ["Brand", "รุ่นรถ", "รุ่นรถ2", "Powertrain", "Total"]
    A_OFF, B_OFF = 1, 5

    def _fmt(v):
        try:
            return f"{int(v):,}"
        except (TypeError, ValueError):
            return ""

    def _cell(col_idx, row_idx, value):
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            pass
        col = _col(col_idx)
        ref = f"{col}{row_idx}"
        s = str(value).strip()
        if not s or s.lower() in ("nan", "none", "<na>"):
            return ""
        return f'<c r="{ref}" t="inlineStr"><is><t>{_esc(s)}</t></is></c>'

    # Collect all cells keyed by row number to avoid duplicate <row> elements
    row_cells: dict[int, list[str]] = {}

    def _add(r, *cells):
        row_cells.setdefault(r, []).extend(c for c in cells if c)

    # Header row
    _add(1,
        *[_cell(A_OFF + i, 1, h) for i, h in enumerate(A_HEADERS)],
        *[_cell(B_OFF + i, 1, h) for i, h in enumerate(B_HEADERS)],
    )

    # Table A data rows
    for idx, row in tbl_a.reset_index(drop=True).iterrows():
        r = idx + 2
        _add(r,
            _cell(A_OFF,     r, row["ชนิดเชื้อเพลิง"]),
            _cell(A_OFF + 1, r, row["Powertrain"]),
            _cell(A_OFF + 2, r, _fmt(row["จำนวนรถ"])),
        )

    # Grand Total — placed in the row right after Table A data
    grand_total = int(tbl_a["จำนวนรถ"].sum())
    gt_r = len(tbl_a) + 2
    _add(gt_r,
        _cell(A_OFF,     gt_r, "Grand Total"),
        _cell(A_OFF + 2, gt_r, _fmt(grand_total)),
    )

    # Table B data rows
    for idx, row in tbl_b.reset_index(drop=True).iterrows():
        r = idx + 2
        _add(r,
            _cell(B_OFF,     r, row["ยี่ห้อรถ2"]),
            _cell(B_OFF + 1, r, row["รุ่นรถ"]),
            _cell(B_OFF + 2, r, row["รุ่นรถ2"]),
            _cell(B_OFF + 3, r, row["Powertrain"]),
            _cell(B_OFF + 4, r, _fmt(row["จำนวนรถ"])),
        )

    rows_xml = [
        f'<row r="{r}">{"".join(cells)}</row>'
        for r, cells in sorted(row_cells.items())
        if cells
    ]

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheetData>' + "".join(rows_xml) + "</sheetData>"
        "</worksheet>"
    ).encode("utf-8")

    with zipfile.ZipFile(str(path), "r") as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        ct_xml   = z.read("[Content_Types].xml").decode("utf-8")
        all_parts = set(z.namelist())

    used_ns  = {int(m) for m in re.findall(r'xl/worksheets/sheet(\d+)\.xml', " ".join(all_parts))}
    next_n   = max(used_ns, default=0) + 1
    new_part = f"xl/worksheets/sheet{next_n}.xml"

    used_rids = {int(r) for r in re.findall(r'Id="rId(\d+)"', rels_xml)}
    new_rid   = f"rId{max(used_rids, default=0) + 1}"

    used_sids = {int(s) for s in re.findall(r'sheetId="(\d+)"', wb_xml)}
    next_sid  = max(used_sids, default=0) + 1

    esc_name = _esc(sheet_name)
    wb_xml = wb_xml.replace(
        "</sheets>",
        f'<sheet name="{esc_name}" sheetId="{next_sid}" r:id="{new_rid}"/></sheets>',
    )
    rels_xml = rels_xml.replace(
        "</Relationships>",
        f'<Relationship Id="{new_rid}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="worksheets/sheet{next_n}.xml"/>'
        "</Relationships>",
    )
    ct_xml = ct_xml.replace(
        "</Types>",
        f'<Override PartName="/xl/worksheets/sheet{next_n}.xml" '
        f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>",
    )

    tmp = path.with_suffix(".tmp2.xlsx")
    with zipfile.ZipFile(str(path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            fname = item.filename
            if fname == "xl/workbook.xml":
                zout.writestr(fname, wb_xml.encode("utf-8"))
            elif fname == "xl/_rels/workbook.xml.rels":
                zout.writestr(fname, rels_xml.encode("utf-8"))
            elif fname == "[Content_Types].xml":
                zout.writestr(fname, ct_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(fname))
        zout.writestr(new_part, sheet_xml)

    tmp.replace(path)
    print(f"      Added sheet '{sheet_name}': {len(tbl_a)} fuel rows | {len(tbl_b)} BEV series rows")


# ── Main ──────────────────────────────────────────────────────────────────────

def find_master_model_file() -> Path:
    override = os.environ.get("MASTER_MODEL_PATH")
    if override:
        path = Path(override)
        if not path.exists():
            print(f"ERROR: master model override not found: {path}"); sys.exit(1)
        return path

    # Find master Model template. Prefer refer/ when the operator keeps masters there.
    model_matches = [
        p for pattern in [
            str(BASE / "refer" / "*master*Model.xlsx"),
            str(BASE / "*master*Model.xlsx"),
            str(BASE / "refer" / "*masterModel.xlsx"),
            str(BASE / "*masterModel.xlsx"),
            str(BASE / "refer" / "*master*Model*.xlsx"),
            str(BASE / "*master*Model*.xlsx"),
            str(BASE / "output" / "*masterModel*.xlsx"),
        ]
        for p in glob.glob(pattern)
        if "~$" not in Path(p).name
    ]
    if not model_matches:
        print("ERROR: No masterModel.xlsx found in project root."); sys.exit(1)
    return Path(max(model_matches, key=os.path.getmtime))

def load_existing_parquet(pq_path: Path, is_fuel: bool = False) -> pd.DataFrame | None:
    if pq_path.exists():
        print(f"\n[Rolling rebuild] Loading existing: {pq_path.name}")
        df_existing = pd.read_parquet(str(pq_path))
        required_existing_cols = {"ปี", "เดือน", "ชนิดเชื้อเพลิง"} if is_fuel else {"ปี", "เดือน"}
        missing_existing_cols = required_existing_cols - set(df_existing.columns)
        if missing_existing_cols:
            print(
                "  Existing parquet is stale; missing column(s): "
                + ", ".join(sorted(missing_existing_cols))
            )
            print("  Rebuilding from raw data.")
            return None
        else:
            ex_key_col = df_existing["ปี"].astype(str) + "_" + df_existing["เดือน"].astype(str)
            print(f"  {len(set(ex_key_col.unique()))} year-month pairs | {len(df_existing):,} rows")
            return df_existing
    else:
        print("\n[Full build] No existing parquet — rebuilding from scratch.")
        return None

def load_reference_maps() -> dict:
    print("\n[1/4] Reading reference tables...", flush=True)

    powertrain_map = load_powertrain_map(str(BASE / "config" / "powertrain_map.csv"))
    print(f"      {len(powertrain_map)} powertrain mappings")

    brand_csv_map = load_powertrain_map(str(BASE / "config" / "brand_map.csv"), "brand", "brand2")
    ref_brand2_map = {k.upper(): v for k, v in brand_csv_map.items()}
    brand_rows = sorted(ref_brand2_map.items())
    brand2_order = list(dict.fromkeys(v for _, v in brand_rows))
    merged_brand2_map = ref_brand2_map
    print(f"      {len(brand_rows)} ยี่ห้อรถ2 entries from brand_map.csv")

    series_pt_map = verified_powertrain_map()
    print(f"      {len(series_pt_map)} verified series_registry powertrain mappings")

    series_name_map = canonical_series_map()
    print(f"      {len(series_name_map)} series_registry canonical-name mappings")

    return {
        "powertrain_map": powertrain_map,
        "brand2_order": brand2_order,
        "merged_brand2_map": merged_brand2_map,
        "series_powertrain_map": series_pt_map,
        "series_name_map": series_name_map,
        "unknown_fuels": set()
    }

def add_derived_columns(df_model: pd.DataFrame, df_fuel: pd.DataFrame, maps: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    print("\n[3/4] Adding derived columns...", flush=True)

    powertrain_map = maps["powertrain_map"]
    merged_brand2_map = maps["merged_brand2_map"]
    series_powertrain_map = maps["series_powertrain_map"]
    series_name_map = maps["series_name_map"]
    unknown_fuels = maps["unknown_fuels"]

    # df_fuel used only for master powertrain summary — not included in cleaned data.
    add_brand2(df_fuel, merged_brand2_map)
    df_fuel.insert(df_fuel.columns.get_loc("ยี่ห้อรถ2") + 1, "รุ่นรถ",  None)
    df_fuel.insert(df_fuel.columns.get_loc("รุ่นรถ")    + 1, "รุ่นรถ2", None)
    fuel_values = df_fuel["ชนิดเชื้อเพลิง"].dropna().astype(str).str.strip()
    unknown_fuels.update(f for f in fuel_values.unique() if f not in powertrain_map)
    df_fuel["Powertrain"] = df_fuel["ชนิดเชื้อเพลิง"].apply(
        lambda f: powertrain_map.get(str(f).strip(), "OTH") if pd.notna(f) else None)

    # Model rows: brand2, model2, and powertrain from the verified series_registry only.
    # The model source carries no fuel type of its own, so nothing here may fall back
    # to fuel-derived or dominant-fuel guesses — missing/unreviewed/conflicting series
    # default to N/A until a human verifies them in series_registry.csv.
    add_brand2(df_model, merged_brand2_map)
    model_key = None
    if {"ยี่ห้อรถ2", "รุ่นรถ"}.issubset(df_model.columns):
        model_key = df_model.apply(lambda r: normalize_key(r["ยี่ห้อรถ2"], r["รุ่นรถ"]), axis=1)

    if "รุ่นรถ" in df_model.columns:
        # series_registry canonical-name (any review_status) is the sole naming
        # authority; unmapped raw series display as-is (name only, never powertrain
        # — registry powertrain resolution below is untouched).
        registry_name = model_key.map(series_name_map) if model_key is not None else pd.Series(index=df_model.index, dtype=object)
        df_model.insert(df_model.columns.get_loc("รุ่นรถ") + 1, "รุ่นรถ2",
                        registry_name.fillna(df_model["รุ่นรถ"]))

    if model_key is not None:
        powertrain_col = model_key.map(series_powertrain_map).fillna("N/A")
    else:
        powertrain_col = pd.Series("N/A", index=df_model.index)
    if "Powertrain" in df_model.columns:
        df_model["Powertrain"] = powertrain_col
    else:
        df_model.insert(df_model.columns.get_loc("รุ่นรถ2") + 1, "Powertrain", powertrain_col)
    # include_in_bev_model_report is derived solely from the verified registry —
    # a legacy CSV/Excel inclusion flag can never override it.
    df_model["include_in_bev_model_report"] = powertrain_col == "BEV"

    return df_model, df_fuel

def rolling_merge(df_cleaned: pd.DataFrame, df_existing: pd.DataFrame | None, rebuild_years: set[int], maps: dict, is_fuel: bool = False) -> tuple[pd.DataFrame, dict]:
    raw_year = pd.to_numeric(df_cleaned["ปี"], errors="coerce")
    raw_rebuild_mask = raw_year.isin(rebuild_years)

    if df_existing is not None:
        existing_year = pd.to_numeric(df_existing["ปี"], errors="coerce")
        existing_rebuild_mask = existing_year.isin(rebuild_years)

        df_keep = df_existing[~existing_rebuild_mask]
        df_add = df_cleaned[raw_rebuild_mask]
        merged = pd.concat([df_keep, df_add], ignore_index=True)
        if not is_fuel:
            merged = sort_cleaned_data(merged, maps["brand2_order"])
        df_cleaned = ordered_cols(merged)

        if not is_fuel:
            raw_key = df_cleaned["ปี"].astype(str) + "_" + df_cleaned["เดือน"].astype(str)
            rebuilt_keys = set(raw_key[raw_rebuild_mask].unique())
            ex_key_col = df_existing["ปี"].astype(str) + "_" + df_existing["เดือน"].astype(str)
            existing_rebuild_key = ex_key_col[existing_rebuild_mask]
            existing_rebuild_keys = set(existing_rebuild_key.unique())
            new_keys = rebuilt_keys - existing_rebuild_keys
            corrected_keys = rebuilt_keys & existing_rebuild_keys

            rebuild_year_labels = sorted(str(y) for y in rebuild_years)
            print(f"  Rolling rebuild: BE years {rebuild_year_labels} | {len(rebuilt_keys)} month(s) from raw data")
            print(f"  Kept rows before {min(rebuild_years)}: {len(df_keep):,}")
            if new_keys:
                print(f"  Added {len(new_keys)} new month(s): {', '.join(sorted(new_keys))}")
            if corrected_keys:
                print(f"  Rebuilt {len(corrected_keys)} existing month(s): {', '.join(sorted(corrected_keys))}")
            print(f"  Final row count: {len(df_cleaned):,}")

            return df_cleaned, {"new_keys": new_keys, "corrected_keys": corrected_keys}
        else:
            return df_cleaned, {}
    else:
        if not is_fuel:
            raw_key = df_cleaned["ปี"].astype(str) + "_" + df_cleaned["เดือน"].astype(str)
            rebuilt_keys = set(raw_key[raw_rebuild_mask].unique())
            rebuild_year_labels = sorted(str(y) for y in rebuild_years)
            print(f"  Rolling rebuild: BE years {rebuild_year_labels} | {len(rebuilt_keys)} month(s) from raw data")
            return ordered_cols(df_cleaned), {"new_keys": set(), "corrected_keys": rebuilt_keys}
        return ordered_cols(df_cleaned), {}

def reapply_canonical_maps(df: pd.DataFrame, maps: dict, is_fuel: bool = False) -> pd.DataFrame:
    """Narrowly reapplies canonical brand and model identity fields to final dataset."""
    brand_map = maps["merged_brand2_map"]
    upper_brand = df["ยี่ห้อรถ"].astype(str).str.strip().str.upper()
    df["ยี่ห้อรถ2"] = upper_brand.map(brand_map).fillna(df["ยี่ห้อรถ"])

    if not is_fuel:
        series_powertrain_map = maps["series_powertrain_map"]
        series_name_map = maps.get("series_name_map", {})
        # Reclassify every retained row (not just the freshly rebuilt months) so a
        # powertrain guessed by the old dominant-fuel enrichment cannot survive.
        model_key = df.apply(lambda r: normalize_key(r["ยี่ห้อรถ2"], r["รุ่นรถ"]), axis=1)
        registry_name = model_key.map(series_name_map)
        df["รุ่นรถ2"] = registry_name.fillna(df["รุ่นรถ"])
        df["Powertrain"] = model_key.map(series_powertrain_map).fillna("N/A")
        df["include_in_bev_model_report"] = df["Powertrain"] == "BEV"
        df.drop(columns=["ชนิดเชื้อเพลิง"], errors="ignore", inplace=True)

    return df

def write_pipeline_state(base_dir: Path, out_file: Path, new_keys: set, corrected_keys: set, new_bev_records: list[dict], df_existing: pd.DataFrame | None) -> None:
    import json, datetime
    state = {
        "master_model":    str(out_file.relative_to(base_dir)),
        "run_date":        str(datetime.date.today()),
        "new_months":      sorted(new_keys) if df_existing is not None else [],
        "corrected_months": sorted(corrected_keys) if df_existing is not None else [],
        "new_bev_models":  new_bev_records,
    }
    (base_dir / "pipeline_state.json").write_text(
        json.dumps(state, ensure_ascii=False, indent=2), "utf-8"
    )

def update_known_models(base_dir: Path, df_cleaned: pd.DataFrame) -> None:
    known_models_path = base_dir / "config" / "known_models.txt"
    known_models_path.parent.mkdir(exist_ok=True, parents=True)

    if "รุ่นรถ" in df_cleaned.columns and "ยี่ห้อรถ2" in df_cleaned.columns:
        curr_models_df = df_cleaned[["รุ่นรถ", "ยี่ห้อรถ2"]].dropna().copy()
        curr_models_df["รุ่นรถ_clean"] = curr_models_df["รุ่นรถ"].astype(str).str.strip()
        curr_models_df["ยี่ห้อรถ2_clean"] = curr_models_df["ยี่ห้อรถ2"].astype(str).str.strip()
        curr_models_df = curr_models_df[curr_models_df["รุ่นรถ_clean"] != ""]
        curr_models_df = curr_models_df.drop_duplicates(subset=["รุ่นรถ_clean"])

        if not known_models_path.exists():
            unique_models = sorted(curr_models_df["รุ่นรถ_clean"].unique())
            with open(known_models_path, "w", encoding="utf-8") as f:
                for m in unique_models:
                    f.write(f"{m}\n")
            print(f"known_models.txt created with {len(unique_models)} models")
        else:
            with open(known_models_path, "r", encoding="utf-8") as f:
                known = {line.strip() for line in f if line.strip()}

            new_models = []
            for _, row in curr_models_df.iterrows():
                m = row["รุ่นรถ_clean"]
                if m not in known:
                    b = row["ยี่ห้อรถ2_clean"]
                    print(f"NEW MODEL DETECTED: {m} (brand: {b})")
                    new_models.append(m)

            if new_models:
                with open(known_models_path, "a", encoding="utf-8") as f:
                    for m in new_models:
                        f.write(f"{m}\n")

def main():
    raw1_file = find_file(RAW1_PATTERN, "fuel raw data")
    raw2_file = find_file(RAW2_PATTERN, "model raw data")
    model_file = find_master_model_file()

    print(f"Raw 1 (Fuel) : {raw1_file.name}")
    print(f"Raw 2 (Model): {raw2_file.name}")
    print(f"Master Model : {model_file.name}")

    pq_path = OUTPUT_BASE / "test_model_cleaned.parquet"
    df_existing = load_existing_parquet(pq_path, is_fuel=False)

    print("\n[2/4] Reading raw data...", flush=True)
    df_fuel = read_dlt_file(raw1_file)
    df_model = read_dlt_file(raw2_file)
    print(f"      {len(df_fuel):,} fuel rows | {len(df_model):,} model rows")

    print("\n[Mapping update] Checking master model mapping...", flush=True)
    append_missing_master_model_rows(model_file, df_model)

    maps = load_reference_maps()

    df_model, df_fuel = add_derived_columns(df_model, df_fuel, maps)

    df_cleaned = ordered_cols(df_model)
    df_cleaned = sort_cleaned_data(df_cleaned, maps["brand2_order"])
    validate_model(df_cleaned)
    print(f"      {len(df_cleaned):,} raw rows processed | cols: {list(df_cleaned.columns)}")

    raw_year = pd.to_numeric(df_cleaned["ปี"], errors="coerce")
    latest_raw_year = int(raw_year.max())
    rebuild_years = {latest_raw_year - 1, latest_raw_year}

    df_cleaned, merge_info = rolling_merge(df_cleaned, df_existing, rebuild_years, maps, is_fuel=False)
    df_cleaned = reapply_canonical_maps(df_cleaned, maps, is_fuel=False)
    new_keys = merge_info.get("new_keys", set())
    corrected_keys = merge_info.get("corrected_keys", set())

    # Save intermediate for pivot builder (parquet: safe, fast, preserves dtypes)
    for col in df_cleaned.select_dtypes(include=["object", "str"]).columns:
        df_cleaned[col] = df_cleaned[col].astype(str).replace("nan", pd.NA)
    df_cleaned.to_parquet(str(pq_path), index=False)
    print(f"      Saved intermediate: {pq_path.name}")

    # Save fuel parquet
    fuel_pq_path = OUTPUT_BASE / "test_fuel_cleaned.parquet"
    df_fuel_save = ordered_cols(df_fuel).copy()
    for col in df_fuel_save.select_dtypes(include=["object", "str"]).columns:
        df_fuel_save[col] = df_fuel_save[col].astype(str).replace("nan", pd.NA)

    df_fuel_existing = load_existing_parquet(fuel_pq_path, is_fuel=True)
    df_fuel_save, _ = rolling_merge(df_fuel_save, df_fuel_existing, rebuild_years, maps, is_fuel=True)
    df_fuel_save = reapply_canonical_maps(df_fuel_save, maps, is_fuel=True)
    validate_fuel(df_fuel_save)
    df_fuel_save.to_parquet(str(fuel_pq_path), index=False)
    print(f"      Saved fuel intermediate: {fuel_pq_path.name}")

    out_file = model_file
    print(f"\n[4/4] Updating in place: {out_file.name}", flush=True)

    rewrite_data_rows(out_file, df_cleaned, data_start_row=8)

    write_pipeline_state(OUTPUT_BASE, out_file, new_keys, corrected_keys, [], df_existing)
    update_known_models(OUTPUT_BASE, df_cleaned)

    print(f"\n  Rows : {len(df_cleaned):,}")
    print(f"  → {out_file.name}")
    print("  → pipeline_state.json updated")
    print(f"Unknown fuel types (->OTH): {sorted(maps['unknown_fuels'])}")

if __name__ == '__main__':
    main()
