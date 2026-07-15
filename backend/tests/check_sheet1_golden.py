"""Deterministic regression checks for Sheet 1.

Goal 1: recompute _build_sheet1_data using a synthetic fixture, compare every entry against manually specified expected values.
Goal 2: write a fresh standalone workbook to a temp path, assert text labels and values using openpyxl.
Goal 3: run a real-data smoke check verifying database invariants.

Run from any directory. Exits 0 on PASS, 1 on FAIL.
"""
import sys
import math
import tempfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

TESTS = Path(__file__).resolve().parent
BACKEND = TESTS.parent
sys.path.insert(0, str(BACKEND))

from build_analyst import _build_sheet1_data, write_sheet1_standalone, filter_ry

failures = []

def fail(msg):
    failures.append(msg)

# ── Goal 1: Deterministic Synthetic Test Fixture ─────────────────────────────

# Create a small synthetic dataset covering ICE, BEV, HEV, and PHEV
# for a previous full year and a current partial year (3 months in).
# Fixture years are arbitrary BE years; only the fixture data itself is a
# source of truth — nothing downstream re-hardcodes these numbers.
FIXTURE_PREV_YEAR = 2568
FIXTURE_CURR_YEAR = 2569
FIXTURE_CURR_MONTH = 3

synthetic_records = []

# Prev Year - 12 months
# Jan (1), Feb (2), Mar (3), Apr-Dec (4-12)
for month_num in range(1, 13):
    month_name = {
        1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน",
        5: "พฤษภาคม", 6: "มิถุนายน", 7: "กรกฎาคม", 8: "สิงหาคม",
        9: "กันยายน", 10: "ตุลาคม", 11: "พฤศจิกายน", 12: "ธันวาคม"
    }[month_num]

    # Let's add varying units to exercise YTD, growth, shares, zero/missing handling
    if month_num == 1:
        units = {"ICE": 10, "BEV": 5, "HEV": 2, "PHEV": 1}
    elif month_num == 2:
        units = {"ICE": 12, "BEV": 6, "HEV": 3, "PHEV": 2}
    elif month_num == 3:
        units = {"ICE": 15, "BEV": 8, "HEV": 4, "PHEV": 1}
    else:
        # April to December
        units = {"ICE": 10, "BEV": 5, "HEV": 2, "PHEV": 1}

    for pt, val in units.items():
        synthetic_records.append({
            "ปี": FIXTURE_PREV_YEAR,
            "เดือน": month_name,
            "ประเภทรถ": "รย.1", # passes filter_ry
            "จังหวัด": "กรุงเทพมหานคร",
            "ยี่ห้อรถ": "TOYOTA",
            "ยี่ห้อรถ2": "TOYOTA",
            "รุ่นรถ": "YARIS",
            "รุ่นรถ2": "YARIS",
            "ชนิดเชื้อเพลิง": "เบนซิน",
            "Powertrain": pt,
            "จำนวนรถ": val
        })

# Curr Year - 3 months (Partial year)
# Jan (1), Feb (2), Mar (3)
for month_num in range(1, FIXTURE_CURR_MONTH + 1):
    month_name = {
        1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม"
    }[month_num]

    if month_num == 1:
        units = {"ICE": 14, "BEV": 7, "HEV": 4, "PHEV": 2}
    elif month_num == 2:
        units = {"ICE": 18, "BEV": 9, "HEV": 5, "PHEV": 3}
    elif month_num == 3:
        # March: PHEV is intentionally missing/0 to test missing prior-period and zero-denom handling
        units = {"ICE": 20, "BEV": 12, "HEV": 6}

    for pt, val in units.items():
        synthetic_records.append({
            "ปี": FIXTURE_CURR_YEAR,
            "เดือน": month_name,
            "ประเภทรถ": "รย.1",
            "จังหวัด": "กรุงเทพมหานคร",
            "ยี่ห้อรถ": "TOYOTA",
            "ยี่ห้อรถ2": "TOYOTA",
            "รุ่นรถ": "YARIS",
            "รุ่นรถ2": "YARIS",
            "ชนิดเชื้อเพลิง": "เบนซิน",
            "Powertrain": pt,
            "จำนวนรถ": val
        })

df_synthetic = pd.DataFrame(synthetic_records)
df_ry_syn = filter_ry(df_synthetic)

# Years/month derived from the fixture's own metadata, not re-hardcoded.
fixture_years = sorted(df_synthetic["ปี"].unique())
if fixture_years != [FIXTURE_PREV_YEAR, FIXTURE_CURR_YEAR]:
    fail(f"Synthetic fixture years {fixture_years} do not match [{FIXTURE_PREV_YEAR}, {FIXTURE_CURR_YEAR}]")
prev_year_syn, curr_year_syn = fixture_years[-2], fixture_years[-1]

# Run logic under test
computed_syn = _build_sheet1_data(df_ry_syn, curr_year_syn, prev_year_syn, FIXTURE_CURR_MONTH)

# Independent manual specification of EXPECTED values:
expected_syn = {
    # ── Grand ──
    ("grand", "B"): 18,
    ("grand", "C"): 23,
    ("grand", "F"): 28,
    ("grand", "G"): 1.0,
    ("grand", "H"): 69,
    ("grand", "I"): 1.0,
    ("grand", "J"): 18,
    ("grand", "K"): 18,
    ("grand", "L"): 18,
    ("grand", "M"): 18,
    ("grand", "N"): 18,
    ("grand", "O"): 18,
    ("grand", "P"): 18,
    ("grand", "Q"): 231,
    ("grand", "R"): 1.0,
    ("grand", "S"): 27,
    ("grand", "T"): 35,
    ("grand", "W"): 38,
    ("grand", "X"): 1.0,
    ("grand", "Y"): None,
    ("grand", "Z"): 10 / 28,
    ("grand", "AA"): 100,
    ("grand", "AB"): 1.0,
    ("grand", "AC"): 31 / 69,
    ("grand", "SEC_I"): 69 / 231,

    # ── ICE ──
    ("ICE", "B"): 10,
    ("ICE", "C"): 12,
    ("ICE", "F"): 15,
    ("ICE", "G"): 15 / 28,
    ("ICE", "H"): 37,
    ("ICE", "I"): 37 / 69,
    ("ICE", "J"): 10,
    ("ICE", "K"): 10,
    ("ICE", "L"): 10,
    ("ICE", "M"): 10,
    ("ICE", "N"): 10,
    ("ICE", "O"): 10,
    ("ICE", "P"): 10,
    ("ICE", "Q"): 127,
    ("ICE", "R"): 127 / 231,
    ("ICE", "S"): 14,
    ("ICE", "T"): 18,
    ("ICE", "W"): 20,
    ("ICE", "X"): 20 / 38,
    ("ICE", "Y"): None,
    ("ICE", "Z"): 5 / 15,
    ("ICE", "AA"): 52,
    ("ICE", "AB"): 52 / 100,
    ("ICE", "AC"): 15 / 37,
    ("ICE", "SEC_I"): 37 / 127,

    # ── BEV ──
    ("BEV", "B"): 5,
    ("BEV", "C"): 6,
    ("BEV", "F"): 8,
    ("BEV", "G"): 8 / 28,
    ("BEV", "H"): 19,
    ("BEV", "I"): 19 / 69,
    ("BEV", "J"): 5,
    ("BEV", "K"): 5,
    ("BEV", "L"): 5,
    ("BEV", "M"): 5,
    ("BEV", "N"): 5,
    ("BEV", "O"): 5,
    ("BEV", "P"): 5,
    ("BEV", "Q"): 64,
    ("BEV", "R"): 64 / 231,
    ("BEV", "S"): 7,
    ("BEV", "T"): 9,
    ("BEV", "W"): 12,
    ("BEV", "X"): 12 / 38,
    ("BEV", "Y"): None,
    ("BEV", "Z"): 4 / 8,
    ("BEV", "AA"): 28,
    ("BEV", "AB"): 28 / 100,
    ("BEV", "AC"): 9 / 19,
    ("BEV", "SEC_I"): 19 / 64,

    # ── HEV ──
    ("HEV", "B"): 2,
    ("HEV", "C"): 3,
    ("HEV", "F"): 4,
    ("HEV", "G"): 4 / 28,
    ("HEV", "H"): 9,
    ("HEV", "I"): 9 / 69,
    ("HEV", "J"): 2,
    ("HEV", "K"): 2,
    ("HEV", "L"): 2,
    ("HEV", "M"): 2,
    ("HEV", "N"): 2,
    ("HEV", "O"): 2,
    ("HEV", "P"): 2,
    ("HEV", "Q"): 27,
    ("HEV", "R"): 27 / 231,
    ("HEV", "S"): 4,
    ("HEV", "T"): 5,
    ("HEV", "W"): 6,
    ("HEV", "X"): 6 / 38,
    ("HEV", "Y"): None,
    ("HEV", "Z"): 2 / 4,
    ("HEV", "AA"): 15,
    ("HEV", "AB"): 15 / 100,
    ("HEV", "AC"): 6 / 9,
    ("HEV", "SEC_I"): 9 / 27,

    # ── PHEV ──
    ("PHEV", "B"): 1,
    ("PHEV", "C"): 2,
    ("PHEV", "F"): 1,
    ("PHEV", "G"): 1 / 28,
    ("PHEV", "H"): 4,
    ("PHEV", "I"): 4 / 69,
    ("PHEV", "J"): 1,
    ("PHEV", "K"): 1,
    ("PHEV", "L"): 1,
    ("PHEV", "M"): 1,
    ("PHEV", "N"): 1,
    ("PHEV", "O"): 1,
    ("PHEV", "P"): 1,
    ("PHEV", "Q"): 13,
    ("PHEV", "R"): 13 / 231,
    ("PHEV", "S"): 2,
    ("PHEV", "T"): 3,
    ("PHEV", "W"): None, # March PHEV is missing/None
    ("PHEV", "X"): None,
    ("PHEV", "Y"): None,
    ("PHEV", "Z"): None,
    ("PHEV", "AA"): 5,
    ("PHEV", "AB"): 5 / 100,
    ("PHEV", "AC"): 1 / 4,
    ("PHEV", "SEC_I"): 4 / 13,
}

# Assert all synthetic results
for key_tuple, expected in expected_syn.items():
    actual = computed_syn.get(key_tuple)
    if expected is None and actual is None:
        continue
    if expected is None or actual is None:
        fail(f"Synthetic {key_tuple}: expected {expected!r}, got {actual!r}")
        continue
    if isinstance(expected, int):
        if actual != expected:
            fail(f"Synthetic {key_tuple}: expected {expected}, got {actual}")
    else:
        if not math.isclose(actual, expected, abs_tol=1e-12, rel_tol=0):
            fail(f"Synthetic {key_tuple}: expected {expected}, got {actual} (diff {abs(actual - expected):.3e})")

# Check for extra/missing keys
syn_keys = set(computed_syn.keys())
expected_keys = set(expected_syn.keys())
for k in sorted(syn_keys - expected_keys):
    fail(f"EXTRA KEY in synthetic computed: {k}")
for k in sorted(expected_keys - syn_keys):
    fail(f"MISSING KEY in synthetic expected: {k}")


# ── Goal 1b: Edge cases — missing prior month, missing prior-year month ─────
# Two powertrains (ICE, BEV) only, curr_month_num=5 so V (month curr-1) is
# actually populated by _build_sheet1_data (it needs >= 4 prior months).
EDGE_PREV_YEAR = 3001
EDGE_CURR_YEAR = 3002
EDGE_CURR_MONTH = 5  # May

def _row(year, month_num, pt, units):
    month_name = {
        1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน", 5: "พฤษภาคม",
    }[month_num]
    return {
        "ปี": year, "เดือน": month_name, "ประเภทรถ": "รย.1",
        "จังหวัด": "กรุงเทพมหานคร", "ยี่ห้อรถ": "TOYOTA", "ยี่ห้อรถ2": "TOYOTA",
        "รุ่นรถ": "YARIS", "รุ่นรถ2": "YARIS", "ชนิดเชื้อเพลิง": "เบนซิน",
        "Powertrain": pt, "จำนวนรถ": units,
    }

edge_records = []
# Prev year: Jan-Apr present for both PTs. May (= curr month) is entirely
# absent for both PTs -> F must be None (missing prior-year month), not 0.
for m, (ice, bev) in {1: (10, 5), 2: (11, 6), 3: (12, 7), 4: (13, 8)}.items():
    edge_records.append(_row(EDGE_PREV_YEAR, m, "ICE", ice))
    edge_records.append(_row(EDGE_PREV_YEAR, m, "BEV", bev))
# Curr year: Jan-Mar present for both. April (V) present for ICE only -> BEV's
# V is missing (not zero), so BEV's growth-vs-prior-month (Y) must be None.
for m, (ice, bev) in {1: (20, 9), 2: (21, 10), 3: (22, 11)}.items():
    edge_records.append(_row(EDGE_CURR_YEAR, m, "ICE", ice))
    edge_records.append(_row(EDGE_CURR_YEAR, m, "BEV", bev))
edge_records.append(_row(EDGE_CURR_YEAR, 4, "ICE", 23))  # BEV April intentionally omitted
# May (W, the curr month) present for both.
edge_records.append(_row(EDGE_CURR_YEAR, 5, "ICE", 30))
edge_records.append(_row(EDGE_CURR_YEAR, 5, "BEV", 15))

df_edge = pd.DataFrame(edge_records)
computed_edge = _build_sheet1_data(filter_ry(df_edge), EDGE_CURR_YEAR, EDGE_PREV_YEAR, EDGE_CURR_MONTH)

# Missing prior-year month (F): independently known to be absent from the fixture -> None.
for pt in ("ICE", "BEV", "grand"):
    if computed_edge.get((pt, "F")) is not None:
        fail(f"Edge case missing-prior-year-month: expected ({pt}, F) is None, got {computed_edge.get((pt, 'F'))}")
    if computed_edge.get((pt, "G")) is not None:
        fail(f"Edge case missing-prior-year-month: expected ({pt}, G) is None, got {computed_edge.get((pt, 'G'))}")
    if computed_edge.get((pt, "Z")) is not None:
        fail(f"Edge case missing-prior-year-month: expected ({pt}, Z) is None, got {computed_edge.get((pt, 'Z'))}")

# Missing prior month (V): ICE has April data, BEV does not -> only BEV's Y is None.
if computed_edge.get(("ICE", "V")) != 23:
    fail(f"Edge case missing-prior-month: expected (ICE, V) == 23, got {computed_edge.get(('ICE', 'V'))}")
if computed_edge.get(("BEV", "V")) is not None:
    fail(f"Edge case missing-prior-month: expected (BEV, V) is None, got {computed_edge.get(('BEV', 'V'))}")
if computed_edge.get(("ICE", "Y")) is None:
    fail("Edge case missing-prior-month: expected (ICE, Y) to be a real number since ICE has April data")
if computed_edge.get(("BEV", "Y")) is not None:
    fail(f"Edge case missing-prior-month: expected (BEV, Y) is None (V missing), got {computed_edge.get(('BEV', 'Y'))}")


# ── Goal 1c: Edge case — zero denominator (an explicit recorded 0 row) ──────
# `aggregate._monthly` drops any group whose summed units are 0 (`> 0` filter,
# see aggregate.py) before _build_sheet1_data ever sees it — read directly
# from that source, not from calling the function under test. So a real
# recorded zero row is, by design, indistinguishable from an absent one by
# the time it reaches get(): both surface as None. The behavior worth
# guaranteeing here is that the downstream ratios stay None — no
# ZeroDivisionError, no NaN, no crash — when every powertrain in a month is
# an explicit zero.
ZERO_PREV_YEAR = 3003
ZERO_CURR_YEAR = 3004
ZERO_CURR_MONTH = 2  # Feb

zero_records = [
    _row(ZERO_PREV_YEAR, 1, "ICE", 5), _row(ZERO_PREV_YEAR, 1, "BEV", 3),
    # Prev-year curr-month (Feb) is an explicit recorded zero for both PTs.
    _row(ZERO_PREV_YEAR, 2, "ICE", 0), _row(ZERO_PREV_YEAR, 2, "BEV", 0),
    _row(ZERO_CURR_YEAR, 1, "ICE", 6), _row(ZERO_CURR_YEAR, 1, "BEV", 4),
    _row(ZERO_CURR_YEAR, 2, "ICE", 7), _row(ZERO_CURR_YEAR, 2, "BEV", 2),
]
df_zero = pd.DataFrame(zero_records)
computed_zero = _build_sheet1_data(filter_ry(df_zero), ZERO_CURR_YEAR, ZERO_PREV_YEAR, ZERO_CURR_MONTH)

for pt in ("ICE", "BEV", "grand"):
    if computed_zero.get((pt, "F")) is not None:
        fail(f"Edge case zero-denominator: expected ({pt}, F) is None, got {computed_zero.get((pt, 'F'))}")
    if computed_zero.get((pt, "G")) is not None:
        fail(f"Edge case zero-denominator: expected ({pt}, G) is None, got {computed_zero.get((pt, 'G'))}")
    if computed_zero.get((pt, "Z")) is not None:
        fail(f"Edge case zero-denominator: expected ({pt}, Z) is None, got {computed_zero.get((pt, 'Z'))}")


# ── Goal 1d: Edge case — rank changes between two periods ───────────────────
# Hand-computed from the raw literals below (not from the ranking behavior of
# the function under test): ICE leads BEV in the prior period, BEV overtakes
# ICE in the current period.
RANK_PREV_YEAR = 3005
RANK_CURR_YEAR = 3006
RANK_CURR_MONTH = 1  # Jan
PREV_UNITS = {"ICE": 50, "BEV": 10}
CURR_UNITS = {"ICE": 8, "BEV": 40}

rank_records = [
    _row(RANK_PREV_YEAR, 1, pt, u) for pt, u in PREV_UNITS.items()
] + [
    _row(RANK_CURR_YEAR, 1, pt, u) for pt, u in CURR_UNITS.items()
]
df_rank = pd.DataFrame(rank_records)
computed_rank = _build_sheet1_data(filter_ry(df_rank), RANK_CURR_YEAR, RANK_PREV_YEAR, RANK_CURR_MONTH)

prev_rank_order = sorted(PREV_UNITS, key=lambda pt: PREV_UNITS[pt], reverse=True)
curr_rank_order = sorted(CURR_UNITS, key=lambda pt: CURR_UNITS[pt], reverse=True)
if prev_rank_order == curr_rank_order:
    fail("Rank-changes fixture did not produce a rank flip between periods — fixture design invalid")
if prev_rank_order != ["ICE", "BEV"] or curr_rank_order != ["BEV", "ICE"]:
    fail(f"Rank-changes fixture: unexpected ordering prev={prev_rank_order} curr={curr_rank_order}")

# Cross-check the function's own F/W outputs agree with the same hand-specified literals.
for pt, u in PREV_UNITS.items():
    if computed_rank.get((pt, "F")) != u:
        fail(f"Rank-changes: expected ({pt}, F) == {u}, got {computed_rank.get((pt, 'F'))}")
for pt, u in CURR_UNITS.items():
    if computed_rank.get((pt, "W")) != u:
        fail(f"Rank-changes: expected ({pt}, W) == {u}, got {computed_rank.get((pt, 'W'))}")


# ── Goal 2: Standalone Label and Value Assertions ────────────────────────────
tmp_path = Path(tempfile.gettempdir()) / "sheet1_syn_standalone.xlsx"
write_sheet1_standalone(tmp_path, df_ry_syn, curr_year_syn, prev_year_syn, FIXTURE_CURR_MONTH)

wb = load_workbook(str(tmp_path))
ws = wb["1.Reg by Powertrain"]
wb.close()
tmp_path.unlink(missing_ok=True)

def chk_cell(cell_ref, expected):
    actual = ws[cell_ref].value
    if actual != expected:
        fail(f"Standalone cell {cell_ref}: expected {expected!r}, got {actual!r}")

# Label assertions
chk_cell("A1",  "Registration by Powertrain")
chk_cell("A3",  "ประเภทรถ : รย.1,2,3,6,9,10,11")
chk_cell("A5",  "Powertrain")
chk_cell("A8",  "Grand Total")
chk_cell("A9",  "ICE")
chk_cell("A10", "BEV")
chk_cell("A11", "HEV")
chk_cell("A12", "PHEV")

# Years derived from fixture metadata
chk_cell("F5",  str(prev_year_syn))
chk_cell("S5",  str(curr_year_syn))

# Header labels
chk_cell("F6",  "Mar")
chk_cell("H6",  "Jan-Mar")
chk_cell("Q6",  f"{prev_year_syn} Total")
chk_cell("S6",  "Jan")
chk_cell("T6",  "Feb")
chk_cell("W6",  "Mar")
chk_cell("AA6", f"{curr_year_syn} Total")

chk_cell("B7",  "มกราคม")
chk_cell("C7",  "กุมภาพันธ์")
chk_cell("F7",  "Units")
chk_cell("G7",  "Share")
chk_cell("H7",  "Units")
chk_cell("I7",  "Share")
chk_cell("Q7",  "Units")
chk_cell("R7",  "Share")
chk_cell("S7",  "Units")
chk_cell("T7",  "Units")
chk_cell("W7",  "Units")
chk_cell("X7",  "Share")
chk_cell("Y7",  "Growth vs Feb 2569")
chk_cell("Z7",  "Growth vs Mar 2568")
chk_cell("AA7", "Units")
chk_cell("AB7", "Share")
chk_cell("AC7", "Growth")

# Check arbitrary data value from the sheet (e.g. Grand Total row 8, column H is prev YTD units = 69)
chk_cell("H8", 69)
# ICE Row 9, column Q is prev full year units = 127
chk_cell("Q9", 127)
# BEV Row 10, column AA is YTD units = 28
chk_cell("AA10", 28)
# HEV Row 11, column W is curr month units = 6
chk_cell("W11", 6)
# PHEV Row 12, column W is None
chk_cell("W12", None)


# ── Goal 3: Live Real-Data Smoke Check ───────────────────────────────────────
def run_real_data_smoke_check():
    pq_path = BACKEND / "test_model_cleaned.parquet"
    manifest_path = BACKEND / ".." / "frontend" / "public" / "data" / "cleaned_data_manifest.json"

    if not pq_path.exists():
        fail(f"Real data parquet {pq_path} not found.")
        return

    df_real = pd.read_parquet(str(pq_path))

    # 1. Required schema exists
    expected_cols = {"ปี", "เดือน", "ประเภทรถ", "จังหวัด", "ยี่ห้อรถ", "ยี่ห้อรถ2", "รุ่นรถ", "รุ่นรถ2", "ชนิดเชื้อเพลิง", "Powertrain", "จำนวนรถ"}
    missing_cols = expected_cols - set(df_real.columns)
    if missing_cols:
        fail(f"Real data missing columns: {missing_cols}")

    # 2. No invalid negative units
    negative_units = (df_real["จำนวนรถ"] < 0).sum()
    if negative_units > 0:
        fail(f"Real data contains {negative_units} rows with negative registration counts.")

    # 3. Model and fuel manifest totals agree (if manifest exists)
    if manifest_path.exists():
        import json
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        model_units = manifest.get("model_total_units")
        fuel_units = manifest.get("fuel_total_units")
        if model_units != fuel_units:
            fail(f"Manifest totals disagree: model_total_units={model_units}, fuel_total_units={fuel_units}")

    # 4. Current reporting period can be detected (max year and max month)
    try:
        max_year = df_real["ปี"].max()
        max_month = df_real[df_real["ปี"] == max_year]["เดือน"].unique()
        if pd.isna(max_year) or len(max_month) == 0:
            fail("Could not detect current reporting period from real data.")
    except Exception as e:
        fail(f"Error checking reporting period: {e}")

run_real_data_smoke_check()


# ── Result ───────────────────────────────────────────────────────────────────
if failures:
    print(f"FAIL — {len(failures)} issue(s):")
    for f in failures:
        print(f)
    sys.exit(1)

print("PASS — Deterministic synthetic fixture check and real data smoke check both passed successfully.")
sys.exit(0)
