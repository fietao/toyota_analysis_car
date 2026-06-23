"""
Builds the monthly analyst report from the cleaned parquet.

Input:  test_model_cleaned.parquet
        *(master cal).xlsx  (root) or *(master*cal)*.xlsx / *(calculation)*.xlsx (refer/)
Output: Refreshes the Data sheet in a copy of the master cal template.
"""

import glob, sys, os, shutil, zipfile, re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

BASE         = Path(__file__).resolve().parent
CLEANED_FILE = BASE / "test_model_cleaned.parquet"

_calc_matches = [
    p for p in glob.glob(str(BASE / "*cal*.xlsx"))
    if "~$" not in p and "Model" not in Path(p).name
]
_refer_calc_matches = [
    p for p in glob.glob(str(BASE / "refer" / "*cal*.xlsx"))
    if "~$" not in p and "Model" not in Path(p).name
]
_nested_refer_calc_matches = [
    p for p in glob.glob(str(BASE / "refer" / "refer" / "*cal*.xlsx"))
    if "~$" not in p and "Model" not in Path(p).name
]
CALC_FILE = (
    Path(max(_calc_matches, key=os.path.getmtime))
    if _calc_matches
    else Path(max(_refer_calc_matches, key=os.path.getmtime)) if _refer_calc_matches
    else Path(max(_nested_refer_calc_matches, key=os.path.getmtime)) if _nested_refer_calc_matches
    else BASE / "test_calculation.xlsx"
)

THAI_MONTHS = {
    1:"มกราคม", 2:"กุมภาพันธ์", 3:"มีนาคม",    4:"เมษายน",
    5:"พฤษภาคม", 6:"มิถุนายน",  7:"กรกฎาคม",   8:"สิงหาคม",
    9:"กันยายน", 10:"ตุลาคม",   11:"พฤศจิกายน", 12:"ธันวาคม",
}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]
INCLUDE_RY  = {'1','2','3','6','9','10','11'}

def filter_ry(df):
    """Keep only ประเภทรถ รย.1,2,3,6,9,10,11."""
    codes = df["ประเภทรถ"].str.extract(r"รย\.(\d+)")[0]
    return df[codes.isin(INCLUDE_RY)].copy()


def _get_year_info(df):
    years = sorted(df["ปี"].unique())
    curr_year      = int(years[-1])
    prev_year      = int(years[-2]) if len(years) >= 2 else curr_year
    prev_prev_year = int(years[-3]) if len(years) >= 3 else prev_year - 1
    curr_months    = [m for m in MONTH_ORDER if m in df[df["ปี"] == curr_year]["เดือน"].unique()]
    return prev_year, curr_year, prev_prev_year, curr_months


def _col_letter(n):
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _xml_esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _find_data_sheet_xml_name(zip_path):
    """Return the xl/worksheets/sheetN.xml path for the sheet named 'Data'."""
    with zipfile.ZipFile(str(zip_path), "r") as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m = re.search(r'<sheet\b[^>]*\bname="Data"[^>]*\br:id="([^"]+)"', wb_xml)
    if not m:
        m = re.search(r'<sheet\b[^>]*\br:id="([^"]+)"[^>]*\bname="Data"', wb_xml)
    if not m:
        raise ValueError("'Data' sheet not found in xl/workbook.xml")
    rid = m.group(1)
    m2 = re.search(rf'<Relationship\b[^>]*\bId="{re.escape(rid)}"[^>]*\bTarget="([^"]+)"', rels_xml)
    if not m2:
        raise ValueError(f"No relationship for Id='{rid}' in workbook.xml.rels")
    target = m2.group(1)
    return f"xl/{target}" if not target.startswith("xl/") else target


def _build_sheet_xml(meta_rows, col_names, df):
    """Build minimal worksheet XML for the Data sheet (inline strings)."""
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
        "<sheetData>",
    ]
    r = 1
    for meta_row in meta_rows:
        cells = [
            f'<c r="{_col_letter(c)}{r}" t="inlineStr"><is><t>{_xml_esc(v)}</t></is></c>'
            for c, v in enumerate(meta_row, 1) if v is not None and str(v).strip()
        ]
        if cells:
            parts.append(f'<row r="{r}">{"".join(cells)}</row>')
        r += 1
    cells = [
        f'<c r="{_col_letter(c)}{r}" t="inlineStr"><is><t>{_xml_esc(n)}</t></is></c>'
        for c, n in enumerate(col_names, 1)
    ]
    parts.append(f'<row r="{r}">{"".join(cells)}</row>')
    r += 1
    for row in df.itertuples(index=False):
        cells = []
        for c, v in enumerate(row, 1):
            if pd.notna(v):
                col = _col_letter(c)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    cells.append(f'<c r="{col}{r}"><v>{v}</v></c>')
                else:
                    cells.append(f'<c r="{col}{r}" t="inlineStr"><is><t>{_xml_esc(v)}</t></is></c>')
        if cells:
            parts.append(f'<row r="{r}">{"".join(cells)}</row>')
        r += 1
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _clear_hard_formula_errors(xml):
    hard_errors = ("#DIV/0!", "#REF!", "#VALUE!", "#NAME?")
    if not any(err in xml for err in hard_errors):
        return xml

    def repl(match):
        cell_xml = match.group(0)
        if not any(err in cell_xml for err in hard_errors):
            return cell_xml
        attrs = re.sub(r'\s+t="e"', "", match.group(1))
        return f"<c{attrs}/>"

    return re.sub(r"<c\b([^>]*)>.*?</c>", repl, xml)


def write_calculation_data_sheet(path, df):
    """Replace the Data sheet in an xlsx without loading all rows into openpyxl."""
    cols = [
        "ปี", "เดือน", "ประเภทรถ", "จังหวัด", "ยี่ห้อรถ", "ยี่ห้อรถ2",
        "ชนิดเชื้อเพลิง", "Powertrain", "จำนวนรถ",
    ]
    df_out = df[[c for c in cols if c in df.columns]].copy()

    wb_ro = load_workbook(str(path), read_only=True, data_only=True)
    ws_ro = wb_ro["Data"]
    meta_rows = [list(row) for row in ws_ro.iter_rows(min_row=1, max_row=5, values_only=True)]
    wb_ro.close()

    data_xml_name = _find_data_sheet_xml_name(path)
    new_xml = _build_sheet_xml(meta_rows, list(df_out.columns), df_out).encode("utf-8")

    path = Path(path)
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(str(path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == data_xml_name:
                zout.writestr(item.filename, new_xml)
            elif item.filename == "xl/calcChain.xml":
                pass
            elif item.filename == "xl/workbook.xml":
                raw = zin.read(item.filename).decode("utf-8")
                raw = re.sub(
                    r'<calcPr\b[^>]*/>',
                    '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>',
                    raw,
                    count=1,
                )
                zout.writestr(item.filename, raw.encode("utf-8"))
            elif item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                raw = zin.read(item.filename).decode("utf-8")
                zout.writestr(item.filename, _clear_hard_formula_errors(raw).encode("utf-8"))
            elif "pivotCacheDefinition" in item.filename and item.filename.endswith(".xml"):
                raw = zin.read(item.filename).decode("utf-8")
                patched = re.sub(
                    r'(<pivotCacheDefinition\b)([^>]*?)(/?>)',
                    lambda m: (
                        m.group(1) + m.group(2)
                        + (' refreshOnLoad="1"' if 'refreshOnLoad' not in m.group(2) else '')
                        + m.group(3)
                    ),
                    raw, count=1,
                )
                zout.writestr(item.filename, patched.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    tmp.replace(path)
    print(f"      Saved: {path.name} ({len(df_out):,} data rows)")


def _col_to_num(col):
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _find_sheet1_xml_name(zip_path):
    """Return xl/worksheets/sheetN.xml for '1.Reg by Powertrain'."""
    with zipfile.ZipFile(str(zip_path), "r") as z:
        wb_xml  = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    m = re.search(r'<sheet\b[^>]*name="1\.Reg[^"]*"[^>]*r:id="([^"]+)"', wb_xml)
    if not m:
        m = re.search(r'<sheet\b[^>]*r:id="([^"]+)"[^>]*name="1\.Reg[^"]*"', wb_xml)
    if not m:
        raise ValueError("'1.Reg by Powertrain' sheet not found in workbook.xml")
    rid = m.group(1)
    m2 = re.search(rf'<Relationship\b[^>]*Id="{re.escape(rid)}"[^>]*Target="([^"]+)"', rels_xml)
    target = m2.group(1)
    return f"xl/{target}" if not target.startswith("xl/") else target


def _build_sheet1_data(df_ry, curr_year, prev_year, curr_month_num):
    """
    Compute all cell values for rows 8-12 and 14-18 of 1.Reg by Powertrain.

    df_ry : already ry-filtered DataFrame (BEV Major still present)
    Returns dict: {(pt_key, col_letter): value_or_None}
      pt_key: 'grand' | 'ICE' | 'BEV' | 'HEV' | 'PHEV'
    """
    MONTH_TO_NUM = {v: k for k, v in THAI_MONTHS.items()}
    PTS = ("ICE", "BEV", "HEV", "PHEV")

    df = df_ry.copy()
    df["Powertrain"] = df["Powertrain"].replace("BEV Major", "BEV")
    df = df[df["Powertrain"].isin(set(PTS))].copy()
    df["month_num"] = df["เดือน"].map(MONTH_TO_NUM).fillna(0).astype(int)
    # Keep all prev_year; limit curr_year to <= curr_month_num
    df = df[
        (df["ปี"] < curr_year) |
        ((df["ปี"] == curr_year) & (df["month_num"] <= curr_month_num))
    ].copy()

    agg = df.groupby(["ปี", "month_num", "Powertrain"])["จำนวนรถ"].sum()

    def get(year, month, pt):
        try:
            return int(agg[(year, month, pt)])
        except KeyError:
            return None

    def total(year, month):
        t = sum(get(year, month, pt) or 0 for pt in PTS)
        return t or None

    def ytd(year, pt, thru):
        t = sum(get(year, m, pt) or 0 for m in range(1, thru + 1))
        return t or None

    def ytd_total(year, thru):
        t = sum(ytd(year, pt, thru) or 0 for pt in PTS)
        return t or None

    def full_year(year, pt):  return ytd(year, pt, 12)
    def full_total(year):     return ytd_total(year, 12)

    def safe_div(a, b):
        return (a / b) if a is not None and b else None

    data = {}
    all_pts = [("grand", None)] + [(pt, pt) for pt in PTS]

    def v_units(pt_key, pt, year, month):
        return total(year, month) if pt_key == "grand" else get(year, month, pt)

    def v_ytd(pt_key, pt, year, thru):
        return ytd_total(year, thru) if pt_key == "grand" else ytd(year, pt, thru)

    def v_full(pt_key, pt, year):
        return full_total(year) if pt_key == "grand" else full_year(year, pt)

    for pt_key, pt in all_pts:
        # Prev year: months before curr month (B=Jan … E=Apr)
        for col, mn in zip("BCDE", range(1, curr_month_num)):
            data[(pt_key, col)] = v_units(pt_key, pt, prev_year, mn)

        # F = prev year curr month
        data[(pt_key, "F")] = v_units(pt_key, pt, prev_year, curr_month_num)

        # J–P = prev year months after curr month
        for col, mn in zip("JKLMNOP", range(curr_month_num + 1, 13)):
            data[(pt_key, col)] = v_units(pt_key, pt, prev_year, mn)

        # Q = prev year full year
        data[(pt_key, "Q")] = v_full(pt_key, pt, prev_year)

        # Curr year: S–V = months 1 … curr_month-1
        for col, mn in zip("STUV", range(1, curr_month_num)):
            data[(pt_key, col)] = v_units(pt_key, pt, curr_year, mn)

        # W = curr year curr month
        data[(pt_key, "W")] = v_units(pt_key, pt, curr_year, curr_month_num)

        # AA = curr year YTD
        data[(pt_key, "AA")] = v_ytd(pt_key, pt, curr_year, curr_month_num)

        # H = prev year YTD (written here so G/I/AC can reference it)
        data[(pt_key, "H")] = v_ytd(pt_key, pt, prev_year, curr_month_num)

    # Derived columns — require grand total refs to already be in data
    for pt_key, pt in all_pts:
        f  = data[(pt_key, "F")];  gt_f  = data[("grand", "F")]
        h  = data[(pt_key, "H")];  gt_h  = data[("grand", "H")]
        q  = data[(pt_key, "Q")];  gt_q  = data[("grand", "Q")]
        w  = data[(pt_key, "W")];  gt_w  = data[("grand", "W")]
        v  = data.get((pt_key, "V"))
        aa = data[(pt_key, "AA")]; gt_aa = data[("grand", "AA")]

        data[(pt_key, "G")]     = safe_div(f, gt_f)
        data[(pt_key, "I")]     = safe_div(h, gt_h)
        data[(pt_key, "R")]     = safe_div(q, gt_q)
        data[(pt_key, "X")]     = safe_div(w, gt_w)
        data[(pt_key, "Y")]     = safe_div((w or 0) - (v or 0), v) if w and v else None
        data[(pt_key, "Z")]     = safe_div((w or 0) - (f or 0), f) if w and f else None
        data[(pt_key, "AB")]    = safe_div(aa, gt_aa)
        data[(pt_key, "AC")]    = safe_div((aa or 0) - (h or 0), h) if aa and h else None
        data[(pt_key, "SEC_I")] = safe_div(h, q)  # prev YTD as % of full year

    return data


# Columns B–AC that Python writes (in order)
_SHEET1_WRITE_COLS = (
    "B C D E F G H I J K L M N O P Q R S T U V W X Y Z AA AB AC".split()
)


def write_sheet1(out_path, df_ry, curr_year, prev_year, curr_month_num):
    """Surgically replace rows 8-12 and 14-18 in 1.Reg by Powertrain with static values."""
    data     = _build_sheet1_data(df_ry, curr_year, prev_year, curr_month_num)
    xml_name = _find_sheet1_xml_name(out_path)

    with zipfile.ZipFile(str(out_path), "r") as z:
        sheet_xml = z.read(xml_name).decode("utf-8")

    def extract_row(xml, row_num):
        m = re.search(rf'<row r="{row_num}"[^>]*>.*?</row>', xml, re.DOTALL)
        return m.group(0) if m else None

    def parse_styles(row_xml):
        """Return {col_letter: style_idx} for every cell in the row."""
        styles = {}
        for m in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>)', row_xml):
            col   = m.group(1)
            attrs = m.group(2)
            s     = re.search(r's="(\d+)"', attrs)
            styles[col] = int(s.group(1)) if s else 0
        return styles

    def make_cell(col, row_num, style, value):
        r = f"{col}{row_num}"
        if value is None:
            return f'<c r="{r}" s="{style}"/>'
        if isinstance(value, float) and value == int(value):
            return f'<c r="{r}" s="{style}"><v>{int(value)}</v></c>'
        return f'<c r="{r}" s="{style}"><v>{value}</v></c>'

    def split_row(row_xml):
        """Return (row_open_tag, {col: cell_xml}) for every cell."""
        row_open = re.match(r'<row [^>]+>', row_xml).group(0)
        cells = {}
        for m in re.finditer(r'<c\b[^>]*>.*?</c>|<c\b[^>]*/>', row_xml, re.DOTALL):
            cell = m.group(0)
            cm = re.search(r'<c r="([A-Z]+)\d+"', cell)
            if cm:
                cells[cm.group(1)] = cell
        return row_open, cells

    WRITE_SET = set(_SHEET1_WRITE_COLS)
    PT_ROWS   = {8: "grand", 9: "ICE", 10: "BEV", 11: "HEV", 12: "PHEV"}

    new_xml = sheet_xml

    for row_num, pt_key in PT_ROWS.items():
        row_xml = extract_row(sheet_xml, row_num)
        if not row_xml:
            continue
        styles            = parse_styles(row_xml)
        row_open, old_cells = split_row(row_xml)

        # Rebuild: col A kept as-is; B-AC replaced; AD+ kept as-is
        new_cells = []
        for col, cell in sorted(old_cells.items(), key=lambda x: _col_to_num(x[0])):
            n = _col_to_num(col)
            if col == "A":
                new_cells.append(cell)
            elif col in WRITE_SET:
                new_cells.append(make_cell(col, row_num, styles.get(col, 0), data.get((pt_key, col))))
            else:  # AD+ formula cells — keep
                new_cells.append(cell)

        new_row = f"{row_open}{''.join(new_cells)}</row>"
        new_xml = new_xml.replace(row_xml, new_row, 1)

    # Secondary section: rows 14-18 — only col I (ratio); keep col H label
    SEC_ROWS = {14: "grand", 15: "ICE", 16: "BEV", 17: "HEV", 18: "PHEV"}
    for row_num, pt_key in SEC_ROWS.items():
        row_xml = extract_row(sheet_xml, row_num)
        if not row_xml:
            continue
        styles            = parse_styles(row_xml)
        row_open, old_cells = split_row(row_xml)

        i_val  = data.get((pt_key, "SEC_I"))

        h_cell = old_cells.get("H", "")
        i_cell = make_cell("I", row_num, styles.get("I", 0), i_val)
        new_row = f"{row_open}{h_cell}{i_cell}</row>"
        new_xml = new_xml.replace(row_xml, new_row, 1)

    tmp = Path(out_path).with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(str(out_path), "r") as zin, \
         zipfile.ZipFile(str(tmp), "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == xml_name:
                zout.writestr(item.filename, new_xml.encode("utf-8"))
            else:
                zout.writestr(item, zin.read(item.filename))
    tmp.replace(Path(out_path))
    print(f"      Sheet 1 written.")


_MONTH_ENG = {
    1: "Jan", 2: "Feb", 3: "Mar", 4:  "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8:  "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}


def write_sheet1_standalone(out_path, df_ry, curr_year, prev_year, curr_month_num):
    """Write a fresh xlsx with '1.Reg by Powertrain' — no template required."""
    from openpyxl import Workbook

    data = _build_sheet1_data(df_ry, curr_year, prev_year, curr_month_num)

    curr_m_eng = _MONTH_ENG[curr_month_num]
    prev_m_eng = _MONTH_ENG[curr_month_num - 1] if curr_month_num > 1 else _MONTH_ENG[12]
    ytd_label  = f"Jan-{curr_m_eng}" if curr_month_num > 1 else curr_m_eng

    wb = Workbook()
    ws = wb.active
    ws.title = "1.Reg by Powertrain"

    # Row 1 — title
    ws["A1"] = "Registration by Powertrain"
    # Row 3 — vehicle type filter note
    ws["A3"] = "ประเภทรถ : รย.1,2,3,6,9,10,11"

    # Row 5 — year group labels
    ws["A5"] = "Powertrain"
    ws["F5"] = str(prev_year)
    ws["S5"] = str(curr_year)

    # Row 6 — period labels
    ws["F6"] = curr_m_eng
    ws["H6"] = ytd_label
    ws["Q6"] = f"{prev_year} Total"
    for col, mn in zip("STUV", range(1, curr_month_num)):
        ws[f"{col}6"] = _MONTH_ENG[mn]
    ws["W6"] = curr_m_eng
    ws["AA6"] = f"{curr_year} Total"

    # Row 7 — column headers
    for col, mn in zip("BCDE", range(1, curr_month_num)):
        ws[f"{col}7"] = THAI_MONTHS[mn]
    ws["F7"] = "Units"
    ws["G7"] = "Share"
    ws["H7"] = "Units"
    ws["I7"] = "Share"
    for col, mn in zip("JKLMNOP", range(curr_month_num + 1, 13)):
        ws[f"{col}7"] = THAI_MONTHS[mn]
    ws["Q7"] = "Units"
    ws["R7"] = "Share"
    for col, mn in zip("STUV", range(1, curr_month_num)):
        ws[f"{col}7"] = "Units"
    ws["W7"] = "Units"
    ws["X7"] = "Share"
    ws["Y7"] = f"Growth vs {prev_m_eng} {curr_year}"
    ws["Z7"] = f"Growth vs {curr_m_eng} {prev_year}"
    ws["AA7"] = "Units"
    ws["AB7"] = "Share"
    ws["AC7"] = "Growth"

    # Rows 8–12 — primary data
    PT_ROWS   = {8: "grand", 9: "ICE", 10: "BEV", 11: "HEV", 12: "PHEV"}
    PT_LABELS = {"grand": "Grand Total", "ICE": "ICE", "BEV": "BEV", "HEV": "HEV", "PHEV": "PHEV"}
    for row_num, pt_key in PT_ROWS.items():
        ws[f"A{row_num}"] = PT_LABELS[pt_key]
        for col in _SHEET1_WRITE_COLS:
            val = data.get((pt_key, col))
            if val is not None:
                ws[f"{col}{row_num}"] = val

    # Rows 14–18 — secondary data (hidden in template: prev YTD % of full year)
    SEC_ROWS = {14: "grand", 15: "ICE", 16: "BEV", 17: "HEV", 18: "PHEV"}
    for row_num, pt_key in SEC_ROWS.items():
        ws[f"H{row_num}"] = PT_LABELS[pt_key]
        sec_i = data.get((pt_key, "SEC_I"))
        if sec_i is not None:
            ws[f"I{row_num}"] = sec_i

    wb.save(str(out_path))
    print(f"      Standalone Sheet 1: {Path(out_path).name}")


def _verify_sheet1(df_ry, curr_year, prev_year, curr_month_num):
    """Print independent verification totals for sheet 1."""
    MONTH_TO_NUM = {v: k for k, v in THAI_MONTHS.items()}
    df = df_ry.copy()
    df["Powertrain"] = df["Powertrain"].replace("BEV Major", "BEV")
    df = df[df["Powertrain"].isin({"ICE", "BEV", "HEV", "PHEV"})].copy()
    df["month_num"] = df["เดือน"].map(MONTH_TO_NUM).fillna(0).astype(int)

    def s1_filter(year, month):
        return df[(df["ปี"] == year) & (df["month_num"] == month)]

    def s1_ytd(year, thru):
        return df[(df["ปี"] == year) & (df["month_num"] <= thru)]

    curr_m_total  = s1_filter(curr_year, curr_month_num)["จำนวนรถ"].sum()
    prev_m_total  = s1_filter(prev_year, curr_month_num)["จำนวนรถ"].sum()
    curr_ytd      = s1_ytd(curr_year, curr_month_num)["จำนวนรถ"].sum()
    prev_ytd      = s1_ytd(prev_year, curr_month_num)["จำนวนรถ"].sum()
    prev_full     = df[df["ปี"] == prev_year]["จำนวนรถ"].sum()

    print("\n--- Sheet 1 verification (independent pandas) ---")
    print(f"  Curr year  curr month total : {curr_m_total:,}")
    print(f"  Prev year  curr month total : {prev_m_total:,}")
    print(f"  Curr year  YTD             : {curr_ytd:,}")
    print(f"  Prev year  YTD             : {prev_ytd:,}")
    print(f"  Prev year  full year       : {prev_full:,}")
    print(f"  MoM  (grand total)         : {(curr_m_total - s1_filter(curr_year, curr_month_num-1)['จำนวนรถ'].sum()) / s1_filter(curr_year, curr_month_num-1)['จำนวนรถ'].sum():.4f}")
    print(f"  YoY  (grand total)         : {(curr_m_total - prev_m_total) / prev_m_total:.4f}")
    print(f"  YTD YoY                    : {(curr_ytd - prev_ytd) / prev_ytd:.4f}")
    print(f"  Prev YTD/Full year         : {prev_ytd / prev_full:.4f}")
    # Per-powertrain curr month spot-check
    for pt in ("ICE", "BEV", "HEV", "PHEV"):
        n = s1_filter(curr_year, curr_month_num).query(f"Powertrain=='{pt}'")["จำนวนรถ"].sum()
        print(f"  {pt:4s} curr month            : {n:,}")
    print("---------------------------------------------------")


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    if not CLEANED_FILE.exists():
        print(f"ERROR: {CLEANED_FILE.name} not found — run build_cleaned.py first")
        sys.exit(1)
    if not CALC_FILE.exists():
        print(f"ERROR: {CALC_FILE.name} not found")
        sys.exit(1)
    print(f"Cleaned : {CLEANED_FILE.name}")
    print(f"Calc    : {CALC_FILE.name}")

    print("Loading cleaned data...", flush=True)
    df_raw = pd.read_parquet(str(CLEANED_FILE))
    df_raw["จำนวนรถ"] = pd.to_numeric(df_raw["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    df_raw["ปี"]      = pd.to_numeric(df_raw["ปี"],      errors="coerce").dropna().astype(int)
    df_raw = df_raw.dropna(subset=["ปี"]).copy()
    df_raw["ปี"] = df_raw["ปี"].astype(int)

    KNOWN_PT = {"ICE", "BEV", "HEV", "PHEV"}
    df_file1 = (
        df_raw[df_raw["ชนิดเชื้อเพลิง"].notna() & df_raw["Powertrain"].isin(KNOWN_PT)].copy()
        if "ชนิดเชื้อเพลิง" in df_raw.columns
        else df_raw.copy()
    )
    df = filter_ry(df_file1)
    print(f"      {len(df_raw):,} total rows → {len(df_file1):,} File1/mapped → {len(df):,} after รย filter")

    prev_year, curr_year, prev_prev_year, curr_months = _get_year_info(df)
    print(f"      Years: prev={prev_year}, curr={curr_year}, curr_months={curr_months}")

    import datetime
    thai_month = curr_months[-1]
    month_num  = {v: k for k, v in THAI_MONTHS.items()}[thai_month]
    greg_year  = curr_year - 543
    prefix     = f"{greg_year}{month_num:02d}"
    out_file   = BASE / f"{prefix}_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - {thai_month} {curr_year}(test analyst).xlsx"

    print(f"\nCopying template → {out_file.name}...", flush=True)
    shutil.copy2(str(CALC_FILE), str(out_file))

    print("Refreshing Data sheet...", flush=True)
    write_calculation_data_sheet(out_file, df_raw)
    print("      Data sheet updated; calculation sheets preserved.")

    print("Writing sheet 1 (1.Reg by Powertrain)...", flush=True)
    df_ry_all = filter_ry(df_raw)  # includes BEV Major rows needed for sheet 1
    write_sheet1(out_file, df_ry_all, curr_year=curr_year, prev_year=prev_year,
                 curr_month_num=month_num)

    # Verification: independent totals from parquet
    _verify_sheet1(df_ry_all, curr_year=curr_year, prev_year=prev_year,
                   curr_month_num=month_num)

    # Standalone: prove Sheet 1 can be written without the template
    standalone_file = BASE / f"{prefix}_sheet1_standalone.xlsx"
    print("Writing standalone Sheet 1...", flush=True)
    write_sheet1_standalone(standalone_file, df_ry_all, curr_year=curr_year,
                            prev_year=prev_year, curr_month_num=month_num)

    print(f"\nOutput: {out_file.name}")
    print(f"Standalone: {standalone_file.name}")


if __name__ == "__main__":
    main()
