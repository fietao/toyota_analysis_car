#!/usr/bin/env python3
"""
update_raw_data.py — Drop in new government files and update the masters.

Usage:
    python update_raw_data.py <new_fuel_file.xlsx> <new_model_file.xlsx>

What it does:
    1. Copies both files into raw data/ as the new masters.
    2. Auto-classifies any new fuel types into the master powertrain sheet.
    3. Runs the full pipeline (build_cleaned → build_BEV → build_analyst).
    4. Exports dashboard data (export_dashboard.py).
"""

import sys
import shutil
import glob
import subprocess
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE    = Path(__file__).resolve().parent
RAW_DIR = BASE / "raw data"

FUEL_PATTERN  = "รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564*.xlsx"
MODEL_PATTERN = "รถใหม่_ยี่ห้อรถ-รุ่นรถ-จังหวัด ปี 2564*.xlsx"


def find_master(pattern: str) -> Path:
    matches = [p for p in RAW_DIR.glob(pattern) if not p.name.startswith("~$")]
    if not matches:
        print(f"ERROR: No master found: {RAW_DIR / pattern}")
        sys.exit(1)
    return max(matches, key=lambda p: p.stat().st_mtime)


def classify_powertrain(fuel_type: str) -> str:
    if not fuel_type or pd.isna(fuel_type):
        return "N/A"
    f = fuel_type.strip()
    if any(pattern in f for pattern in ["ไม่ระบุ", "ไม่ไฟฟ้าเพลิง", "เพลิงอื่นอื่น ๆ", "ไม่ทราบ", "อื่น ๆ"]):
        return "N/A"
    if "เสียบปลั๊ก" in f or "แบบเสียบปลั๊ก" in f:
        return "PHEV"
    if f == "ไฟฟ้า":
        return "BEV"
    if "ไฟฟ้า" in f:
        return "HEV"
    return "ICE"


def update_master_powertrain(fuel_df: pd.DataFrame):
    """Add any new fuel types from fuel data into the master powertrain sheet."""
    refer_pattern = str(BASE / "*master Model.xlsx")
    matches = [p for p in glob.glob(refer_pattern) if "~$" not in p]
    if not matches:
        print("  WARNING: No master Model file found — skipping master powertrain update")
        return

    refer_path = Path(matches[0])
    wb = load_workbook(str(refer_path))
    ws = wb["master powertrain"]

    existing = set()
    last_row = 7
    for row in ws.iter_rows(min_row=8, max_col=6, values_only=True):
        if row[4] is not None:
            existing.add(str(row[4]).strip())
            last_row = ws.max_row

    new_types = [
        ft for ft in fuel_df["ชนิดเชื้อเพลิง"].dropna().unique()
        if str(ft).strip() not in existing
    ]

    if not new_types:
        print("  master powertrain: no new fuel types found")
        wb.close()
        return

    for ft in sorted(new_types):
        pt = classify_powertrain(ft)
        last_row += 1
        ws.cell(row=last_row, column=5, value=ft)
        ws.cell(row=last_row, column=6, value=pt)
        print(f"  + Added to master powertrain: {ft} → {pt}")

    wb.save(str(refer_path))
    wb.close()


def main():
    if len(sys.argv) == 3:
        new_fuel_path  = Path(sys.argv[1])
        new_model_path = Path(sys.argv[2])

        for p in [new_fuel_path, new_model_path]:
            if not p.exists():
                print(f"ERROR: Not found: {p}")
                sys.exit(1)

        print("[1/4] Copying new files into raw data/...")
        fuel_master = RAW_DIR / new_fuel_path.name
        model_master = RAW_DIR / new_model_path.name

        if new_fuel_path.resolve() != fuel_master.resolve():
            shutil.copy2(new_fuel_path, fuel_master)
        if new_model_path.resolve() != model_master.resolve():
            shutil.copy2(new_model_path, model_master)
        print(f"  Fuel  → {fuel_master.name}")
        print(f"  Model → {model_master.name}")
    elif len(sys.argv) == 1:
        fuel_master  = find_master(FUEL_PATTERN)
        model_master = find_master(MODEL_PATTERN)
        print(f"[1/4] Using existing files in raw data/...")
        print(f"  Fuel  : {fuel_master.name}")
        print(f"  Model : {model_master.name}")
    else:
        print("Usage: python update_raw_data.py [<new_fuel_file.xlsx> <new_model_file.xlsx>]")
        sys.exit(1)

    print("\n[2/4] Checking master powertrain for new fuel types...")
    fuel_df = pd.read_excel(str(fuel_master), header=5, engine="calamine")
    update_master_powertrain(fuel_df)

    print("\n[3/4] Running pipeline...")
    pipeline = BASE / "run_pipeline.py"
    result = subprocess.run(
        [sys.executable, str(pipeline), "--skip-map", "--skip-analyst"],
        cwd=str(BASE),
        env={**os.environ, "PYTHONUTF8": "1"},
    )
    if result.returncode != 0:
        print(f"\nERROR: Pipeline failed (exit {result.returncode})")
        sys.exit(result.returncode)

    print("\n[4/4] Exporting dashboard data...")
    export = BASE / "export_dashboard.py"
    result = subprocess.run(
        [sys.executable, str(export)],
        cwd=str(BASE),
        env={**os.environ, "PYTHONUTF8": "1"},
    )
    if result.returncode != 0:
        print(f"\nERROR: Dashboard export failed (exit {result.returncode})")
        sys.exit(result.returncode)

    print("\n[5/5] Exporting analyst data (non-fatal)...")
    export_analyst = BASE / "export_analyst.py"
    result = subprocess.run(
        [sys.executable, str(export_analyst)],
        cwd=str(BASE),
        env={**os.environ, "PYTHONUTF8": "1"},
    )
    if result.returncode != 0:
        print(f"\nWARNING: Analyst export failed (exit {result.returncode}) — skipping.")

    print("\nAll done. Data updated, pipeline rebuilt, dashboard exported.")


if __name__ == "__main__":
    main()
