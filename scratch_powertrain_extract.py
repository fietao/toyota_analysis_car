"""
Extract Powertrain master data from mastercal workbook.
"""
import openpyxl, time, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILE = r"C:\dev\ai-reading-car-analysis\backend\refer\refer\202605_รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564 - พฤษภาคม 2569(masttercal).xlsx"

t0 = time.time()
wb = openpyxl.load_workbook(FILE, data_only=True, read_only=True)
print(f"Loaded in {time.time()-t0:.1f}s")
print(f"Sheets: {wb.sheetnames}")

# Focus on key sheets: '1.Reg by Powertrain', 'master powertrain', 'Pivot'
target_sheets = ['1.Reg by Powertrain', 'master powertrain', 'Pivot']

for name in target_sheets:
    if name not in wb.sheetnames:
        print(f"\nSheet '{name}' not found, skipping")
        continue
    ws = wb[name]
    print(f"\n{'='*60}")
    print(f"Sheet: '{name}'")
    print(f"{'='*60}")
    row_count = 0
    for row in ws.iter_rows(max_row=50, max_col=20, values_only=False):
        row_count += 1
        vals = []
        for cell in row:
            if cell.value is not None:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                vals.append(f"{col_letter}{cell.row}={repr(cell.value)}")
        if vals:
            print(f"  Row {row_count}: {vals}")

wb.close()
print(f"\nDone in {time.time()-t0:.1f}s")
