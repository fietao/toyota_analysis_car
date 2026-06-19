#!/usr/bin/env python3
"""
update_raw_data.py — Drop in new government files and update the masters.

Usage:
    python update_raw_data.py <new_fuel_file.xlsx> <new_model_file.xlsx>

What it does:
    1. Copies both files into raw data/ as the new masters.
    2. Fills missing ชนิดเชื้อเพลิง in the model file from the fuel file
       (matched on ปี+เดือน+ประเภทรถ+จังหวัด+ยี่ห้อรถ, dominant fuel type wins).
    3. Auto-classifies any new fuel types into the master powertrain sheet.
    4. Runs the full pipeline (build_cleaned → build_BEV → build_analyst).
    5. Exports dashboard data (export_dashboard.py).
"""

import sys
import shutil
import glob
import subprocess
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE    = Path(__file__).resolve().parent
RAW_DIR = BASE / "raw data"

FUEL_PATTERN  = "รถใหม่_ยี่ห้อรถ-ชนิดเชื้อเพลิง-จังหวัด ปี 2564*.xlsx"
MODEL_PATTERN = "รถใหม่_ยี่ห้อรถ-รุ่นรถ-จังหวัด ปี 2564*.xlsx"
JOIN_KEYS     = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด", "ยี่ห้อรถ"]


def find_master(pattern: str) -> Path:
    matches = [p for p in RAW_DIR.glob(pattern) if not p.name.startswith("~$")]
    if not matches:
        print(f"ERROR: No master found: {RAW_DIR / pattern}")
        sys.exit(1)
    return max(matches, key=lambda p: p.stat().st_mtime)


def read_data(path: Path) -> pd.DataFrame:
    return pd.read_excel(str(path), sheet_name="Data", header=5)


def enrich_fuel_type(model_df: pd.DataFrame, fuel_df: pd.DataFrame) -> pd.DataFrame:
    """Add ชนิดเชื้อเพลิง to model rows using the dominant fuel type from fuel data."""
    lookup = (
        fuel_df.groupby(JOIN_KEYS + ["ชนิดเชื้อเพลิง"])["จำนวนรถ"]
        .sum()
        .reset_index()
        .sort_values("จำนวนรถ", ascending=False)
        .drop_duplicates(subset=JOIN_KEYS)
        .set_index(JOIN_KEYS)["ชนิดเชื้อเพลิง"]
    )
    model_df = model_df.copy()
    model_df["ชนิดเชื้อเพลิง"] = (
        model_df[JOIN_KEYS]
        .apply(lambda r: lookup.get(tuple(r)), axis=1)
    )
    filled = model_df["ชนิดเชื้อเพลิง"].notna().sum()
    print(f"  ชนิดเชื้อเพลิง filled: {filled:,} / {len(model_df):,} rows")
    return model_df


def classify_powertrain(fuel_type: str) -> str:
    """Classify a fuel type name into a powertrain category using keyword rules."""
    f = fuel_type.strip()
    if "ไม่ใช้เชื้อเพลิง" in f:
        return "N/A"
    if "ไฟฟ้าแบบเสียบปลั๊ก" in f:
        return "PHEV"
    if f == "ไฟฟ้า":
        return "BEV"
    if "ไฟฟ้า" in f:
        return "HEV"
    return "ICE"


def update_master_powertrain(fuel_df: pd.DataFrame):
    """Add any new fuel types from fuel data into the master powertrain sheet."""
    refer_pattern = str(BASE / "*master Model.xlsx")
    matches = [p for p in glob.glob(refer_pattern) if "~$" not in p]
    if not matches:
        print("  WARNING: No master Model file found — skipping master powertrain update")
        return

    refer_path = Path(matches[0])
    wb = load_workbook(str(refer_path))
    ws = wb["master powertrain"]

    # Read existing mapped fuel types (col E from row 8 onward)
    existing = set()
    last_row = 7
    for row in ws.iter_rows(min_row=8, max_col=6, values_only=True):
        if row[4] is not None:
            existing.add(str(row[4]).strip())
            last_row = ws.max_row

    # Find new fuel types in the incoming data
    new_types = [
        ft for ft in fuel_df["ชนิดเชื้อเพลิง"].dropna().unique()
        if str(ft).strip() not in existing
    ]

    if not new_types:
        print("  master powertrain: no new fuel types found")
        wb.close()
        return

    # Append new rows
    for ft in sorted(new_types):
        pt = classify_powertrain(ft)
        last_row += 1
        ws.cell(row=last_row, column=5, value=ft)
        ws.cell(row=last_row, column=6, value=pt)
        print(f"  + Added to master powertrain: {ft} → {pt}")

    wb.save(str(refer_path))
    wb.close()


def write_back(master_path: Path, new_df: pd.DataFrame):
    """Replace the Data sheet rows while keeping metadata rows 1-5 intact."""
    meta_df = pd.read_excel(str(master_path), sheet_name="Data", nrows=5, header=None)
    with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        meta_df.to_excel(writer, sheet_name="Data", index=False, header=False)
        new_df.to_excel(writer, sheet_name="Data", index=False, startrow=5)
    print(f"  Saved: {master_path.name} ({len(new_df):,} rows)")


def main():
    if len(sys.argv) == 3:
        new_fuel_path  = Path(sys.argv[1])
        new_model_path = Path(sys.argv[2])

        for p in [new_fuel_path, new_model_path]:
            if not p.exists():
                print(f"ERROR: Not found: {p}")
                sys.exit(1)

        fuel_master  = find_master(FUEL_PATTERN)
        model_master = find_master(MODEL_PATTERN)

        # Step 1: Copy new files over the masters
        print("[1/2] Copying new files into raw data/...")
        if new_fuel_path.resolve() != fuel_master.resolve():
            shutil.copy2(new_fuel_path,  fuel_master)
        if new_model_path.resolve() != model_master.resolve():
            shutil.copy2(new_model_path, model_master)
        print(f"  Fuel  → {fuel_master.name}")
        print(f"  Model → {model_master.name}")
    elif len(sys.argv) == 1:
        fuel_master  = find_master(FUEL_PATTERN)
        model_master = find_master(MODEL_PATTERN)
        print(f"[1/2] Using existing files in raw data/...")
        print(f"  Fuel  : {fuel_master.name}")
        print(f"  Model : {model_master.name}")
    else:
        print("Usage: python update_raw_data.py [<new_fuel_file.xlsx> <new_model_file.xlsx>]")
        sys.exit(1)

    # Step 2: Enrich model with ชนิดเชื้อเพลิง and write back
    print("\n[2/3] Enriching model with ชนิดเชื้อเพลิง from fuel data...")
    fuel_df  = read_data(fuel_master)
    model_df = read_data(model_master)
    enriched = enrich_fuel_type(model_df, fuel_df)
    write_back(model_master, enriched)

    # Step 3: Auto-update master powertrain with any new fuel types
    print("\n[3/3] Checking master powertrain for new fuel types...")
    update_master_powertrain(fuel_df)

    # Step 4: Run the full pipeline
    print("\n[4/5] Running pipeline...")
    pipeline = BASE / ".claude" / "scripts" / "run_pipeline.py"
    result = subprocess.run(
        [sys.executable, str(pipeline), "--skip-map"],
        cwd=str(BASE),
        env={**os.environ, "PYTHONUTF8": "1"},
    )
    if result.returncode != 0:
        print(f"\nERROR: Pipeline failed (exit {result.returncode})")
        sys.exit(result.returncode)

    # Step 5: Export dashboard data
    print("\n[5/5] Exporting dashboard data...")
    export = BASE / "export_dashboard.py"
    result = subprocess.run(
        [sys.executable, str(export)],
        cwd=str(BASE),
        env={**os.environ, "PYTHONUTF8": "1"},
    )
    if result.returncode != 0:
        print(f"\nERROR: Dashboard export failed (exit {result.returncode})")
        sys.exit(result.returncode)

    print("\nAll done. Data updated, pipeline rebuilt, dashboard exported.")


if __name__ == "__main__":
    main()
