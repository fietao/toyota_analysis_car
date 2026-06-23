"""Thailand EV Market Registration Dashboard"""
import io
import urllib.request
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────
BASE      = Path(__file__).parent
PARQUET   = BASE / "backend" / "test_model_cleaned.parquet"
FONTS_DIR = BASE / "fonts"
FONT_PATH = FONTS_DIR / "NotoSansThai-Regular.ttf"
FONT_URL  = (
    "https://cdn.jsdelivr.net/gh/googlefonts/noto-fonts@main"
    "/hinted/ttf/NotoSansThai/NotoSansThai-Regular.ttf"
)

THAI_MONTHS = {
    1: "มกราคม",    2: "กุมภาพันธ์", 3: "มีนาคม",    4: "เมษายน",
    5: "พฤษภาคม",  6: "มิถุนายน",   7: "กรกฎาคม",   8: "สิงหาคม",
    9: "กันยายน",  10: "ตุลาคม",    11: "พฤศจิกายน", 12: "ธันวาคม",
}
MONTH_NUM   = {v: k for k, v in THAI_MONTHS.items()}
MONTH_ORDER = [THAI_MONTHS[i] for i in range(1, 13)]
MON_SHORT   = {
    "มกราคม": "Jan",   "กุมภาพันธ์": "Feb", "มีนาคม": "Mar",
    "เมษายน": "Apr",   "พฤษภาคม": "May",    "มิถุนายน": "Jun",
    "กรกฎาคม": "Jul",  "สิงหาคม": "Aug",    "กันยายน": "Sep",
    "ตุลาคม": "Oct",   "พฤศจิกายน": "Nov",  "ธันวาคม": "Dec",
}
PT_ORDER   = ["ICE", "BEV", "HEV", "PHEV"]

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")


# ── Data ─────────────────────────────────────────────────────────────────────
@st.cache_data
def load_data() -> pd.DataFrame:
    df = pd.read_parquet(str(PARQUET))
    df["ประเภทรถ"] = df["ประเภทรถ"].astype(str).str.strip().replace(
        {"nan": pd.NA, "None": pd.NA, "<NA>": pd.NA}
    )
    df["Powertrain"] = df["Powertrain"].astype(str).str.strip().replace(
        {"nan": pd.NA, "None": pd.NA, "<NA>": pd.NA}
    )
    pt_map = {"BEV Major": "BEV", "OTH": None}
    df["Powertrain"] = df["Powertrain"].map(pt_map).fillna(df["Powertrain"])
    df = df[df["Powertrain"].isin(["ICE", "BEV", "HEV", "PHEV"])]
    df["จำนวนรถ"] = pd.to_numeric(df["จำนวนรถ"], errors="coerce").fillna(0).astype(int)
    df["ปี"]      = pd.to_numeric(df["ปี"], errors="coerce").dropna().astype(int)
    return df


def year_info(df: pd.DataFrame):
    curr = int(df["ปี"].max())
    prev = curr - 1
    curr_months = [m for m in MONTH_ORDER if m in df[df["ปี"] == curr]["เดือน"].unique()]
    return curr, prev, curr_months


def ordered_months(available):
    return [m for m in MONTH_ORDER if m in available]


# ── Tab builders ──────────────────────────────────────────────────────────────
def build_powertrain_tables(df: pd.DataFrame):
    curr, prev, curr_months = year_info(df)
    d_curr = df[df["ปี"] == curr]
    d_prev = df[df["ปี"] == prev]

    # Monthly units
    piv = (
        d_curr.groupby(["Powertrain", "เดือน"])["จำนวนรถ"].sum()
        .unstack(fill_value=0)
        .reindex(index=PT_ORDER, columns=curr_months, fill_value=0)
    )
    piv["YTD"] = piv.sum(axis=1)
    piv.loc["Total"] = piv.sum()
    units = piv.rename(columns=MON_SHORT)

    # Share %
    share = piv.drop("Total").copy()
    for col in share.columns:
        tot = piv.loc["Total", col]
        share[col] = (share[col] / tot * 100).round(1) if tot else 0.0
    share = share.rename(columns=MON_SHORT)

    # YTD comparison
    prev_ytd = (
        d_prev[d_prev["เดือน"].isin(curr_months)]
        .groupby("Powertrain")["จำนวนรถ"].sum()
        .reindex(PT_ORDER, fill_value=0)
    )
    prev_fy = d_prev.groupby("Powertrain")["จำนวนรถ"].sum().reindex(PT_ORDER, fill_value=0)
    curr_ytd = d_curr.groupby("Powertrain")["จำนวนรถ"].sum().reindex(PT_ORDER, fill_value=0)

    comp = pd.DataFrame({
        f"{prev} Full Year": prev_fy,
        f"{prev} YTD": prev_ytd,
        f"{curr} YTD": curr_ytd,
    })
    comp["Growth %"] = (comp[f"{curr} YTD"] / comp[f"{prev} YTD"].replace(0, pd.NA) - 1) * 100
    comp = comp.round({"Growth %": 1})
    totals = comp.sum(numeric_only=True)
    py_ytd_tot = totals[f"{prev} YTD"]
    cy_ytd_tot = totals[f"{curr} YTD"]
    totals["Growth %"] = round((cy_ytd_tot / py_ytd_tot - 1) * 100, 1) if py_ytd_tot else None
    comp.loc["Total"] = totals

    return units, share, comp, curr, prev


def build_brand_table(df: pd.DataFrame, pt_filter: str | None = None):
    curr, prev, curr_months = year_info(df)
    sub_c = df[df["ปี"] == curr]
    sub_p = df[(df["ปี"] == prev) & (df["เดือน"].isin(curr_months))]
    if pt_filter:
        sub_c = sub_c[sub_c["Powertrain"] == pt_filter]
        sub_p = sub_p[sub_p["Powertrain"] == pt_filter]

    brand_c = (
        sub_c.groupby(["ยี่ห้อรถ2", "เดือน"])["จำนวนรถ"].sum()
        .unstack(fill_value=0)
        .reindex(columns=curr_months, fill_value=0)
    )
    brand_c["YTD"] = brand_c.sum(axis=1)
    brand_c = brand_c.sort_values("YTD", ascending=False)
    brand_c.insert(0, "Rank", range(1, len(brand_c) + 1))

    prev_rank = {
        b: i + 1
        for i, b in enumerate(
            sub_p.groupby("ยี่ห้อรถ2")["จำนวนรถ"].sum().sort_values(ascending=False).index
        )
    }
    brand_c["Prev"] = brand_c.index.map(lambda b: prev_rank.get(b, 0))
    brand_c["Δ Rank"] = brand_c.apply(
        lambda r: "NEW" if r["Prev"] == 0
        else ("—" if r["Prev"] == r["Rank"]
              else f"+{int(r['Prev'] - r['Rank'])}" if r["Prev"] > r["Rank"]
              else str(int(r["Prev"] - r["Rank"]))),
        axis=1,
    )
    brand_c.drop(columns=["Prev"], inplace=True)
    brand_c.rename(columns=MON_SHORT, inplace=True)
    brand_c.index.name = "Brand"
    return brand_c, curr


def build_compare_monthly(df: pd.DataFrame, selections: list, prev: int, curr: int) -> pd.DataFrame:
    prev_months = [m for m in MONTH_ORDER if m in df[df["ปี"] == prev]["เดือน"].unique()]
    curr_months = [m for m in MONTH_ORDER if m in df[df["ปี"] == curr]["เดือน"].unique()]
    prev_cols = [f"{MON_SHORT[m]} {prev}" for m in prev_months]
    curr_cols = [f"{MON_SHORT[m]} {curr}" for m in curr_months]

    rows = {}
    for label, sub in selections:
        row = {}
        for m, col in zip(prev_months, prev_cols):
            row[col] = int(sub[(sub["ปี"] == prev) & (sub["เดือน"] == m)]["จำนวนรถ"].sum())
        for m, col in zip(curr_months, curr_cols):
            row[col] = int(sub[(sub["ปี"] == curr) & (sub["เดือน"] == m)]["จำนวนรถ"].sum())
        rows[label] = row

    table = pd.DataFrame(rows).T.reindex(columns=prev_cols + curr_cols, fill_value=0)
    table.index.name = "Brand / Model"
    return table


def build_compare_ytd(df: pd.DataFrame, selections: list, prev: int, curr: int, curr_months: list) -> pd.DataFrame:
    rows = {}
    for label, sub in selections:
        prev_ytd = int(sub[(sub["ปี"] == prev) & (sub["เดือน"].isin(curr_months))]["จำนวนรถ"].sum())
        curr_ytd = int(sub[(sub["ปี"] == curr) & (sub["เดือน"].isin(curr_months))]["จำนวนรถ"].sum())
        growth = round((curr_ytd / prev_ytd - 1) * 100, 1) if prev_ytd else None
        rows[label] = {f"{prev} YTD": prev_ytd, f"{curr} YTD": curr_ytd, "Growth %": growth}

    table = pd.DataFrame(rows).T
    table.index.name = "Brand / Model"
    return table


def build_bev_model_table(df: pd.DataFrame):
    curr, _, curr_months = year_info(df)
    bev = df[(df["ปี"] == curr) & (df["Powertrain"] == "BEV")]
    piv = (
        bev.groupby(["ยี่ห้อรถ2", "รุ่นรถ2", "เดือน"])["จำนวนรถ"].sum()
        .unstack(fill_value=0)
        .reindex(columns=curr_months, fill_value=0)
    )
    piv["YTD"] = piv.sum(axis=1)
    piv.rename(columns=MON_SHORT, inplace=True)
    piv = piv.sort_values(["ยี่ห้อรถ2", "YTD"], ascending=[True, False])
    piv.index.names = ["Brand", "Model"]
    return piv, curr


# ── Excel export ──────────────────────────────────────────────────────────────
def _write_df_to_sheet(ws, df: pd.DataFrame, title: str):
    ws.title = title
    idx_levels = df.index.nlevels

    # header
    headers = list(df.index.names) + list(df.columns)
    for col_i, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=str(hdr) if hdr else "")
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_i, (idx, row) in enumerate(df.iterrows(), 2):
        fill = ALT_FILL if row_i % 2 == 0 else None
        # index columns
        if idx_levels == 1:
            idx = (idx,)
        for c, val in enumerate(idx, 1):
            cell = ws.cell(row=row_i, column=c, value=str(val) if val is not None else "")
            if fill:
                cell.fill = fill
        # data columns
        for c, val in enumerate(row, idx_levels + 1):
            try:
                num = float(val)
                cell = ws.cell(row=row_i, column=c, value=num)
            except (ValueError, TypeError):
                cell = ws.cell(row=row_i, column=c, value=str(val) if val is not None else "")
            if fill:
                cell.fill = fill

    for col_i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14


def build_excel(df: pd.DataFrame) -> bytes:
    units, share, comp, curr, prev = build_powertrain_tables(df)
    brands_all, _  = build_brand_table(df)
    brands_bev, _  = build_brand_table(df, pt_filter="BEV")
    bev_models, _  = build_bev_model_table(df)

    wb = Workbook()
    ws_list = [
        wb.active,
        wb.create_sheet(),
        wb.create_sheet(),
        wb.create_sheet(),
    ]

    # Sheet 1: Powertrain — units, then share, then comparison stacked
    ws = ws_list[0]
    ws.title = "Powertrain"
    _write_df_to_sheet(ws, units.reset_index(names=["Powertrain"]).set_index("Powertrain"),
                       "Powertrain")
    # append share and comp below with a gap
    start = len(units) + 3
    ws.cell(row=start, column=1, value="Market Share %").font = Font(bold=True)
    for r, (idx, row) in enumerate(share.iterrows(), start + 1):
        ws.cell(row=r, column=1, value=str(idx))
        for c, val in enumerate(row, 2):
            try:
                ws.cell(row=r, column=c, value=round(float(val), 1))
            except (ValueError, TypeError):
                ws.cell(row=r, column=c, value=str(val))

    start2 = start + len(share) + 3
    ws.cell(row=start2, column=1, value="YTD Comparison").font = Font(bold=True)
    for c_i, col in enumerate(comp.columns, 2):
        ws.cell(row=start2 + 1, column=c_i, value=str(col)).font = HDR_FONT
    ws.cell(row=start2 + 1, column=1, value="Powertrain").font = HDR_FONT
    for r, (idx, row) in enumerate(comp.iterrows(), start2 + 2):
        ws.cell(row=r, column=1, value=str(idx))
        for c, val in enumerate(row, 2):
            try:
                ws.cell(row=r, column=c, value=round(float(val), 1))
            except (ValueError, TypeError):
                ws.cell(row=r, column=c, value=str(val) if pd.notna(val) else "")

    _write_df_to_sheet(ws_list[1], brands_all, "Brand Rankings")
    _write_df_to_sheet(ws_list[2], brands_bev, "BEV by Brand")
    _write_df_to_sheet(ws_list[3], bev_models, "BEV by Model")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF export ────────────────────────────────────────────────────────────────
def _ensure_thai_font() -> str | None:
    """Return path to a Thai-capable TTF, downloading if needed. None if unavailable."""
    FONTS_DIR.mkdir(exist_ok=True)
    if FONT_PATH.exists():
        return str(FONT_PATH)

    # Common Windows Thai font locations
    win_thai = [
        r"C:\Windows\Fonts\THSarabunNew.ttf",
        r"C:\Windows\Fonts\TH Sarabun New.ttf",
        r"C:\Windows\Fonts\Angsana New.ttf",
        r"C:\Windows\Fonts\AngsanaUPC.ttf",
        r"C:\Windows\Fonts\Cordia New.ttf",
        r"C:\Windows\Fonts\CordiaUPC.ttf",
        r"C:\Windows\Fonts\Tahoma.ttf",
    ]
    for path in win_thai:
        from pathlib import Path as P
        if P(path).exists():
            return path

    # Try downloading NotoSansThai
    try:
        urllib.request.urlretrieve(FONT_URL, str(FONT_PATH))
        return str(FONT_PATH)
    except Exception:
        return None


def build_pdf(df: pd.DataFrame) -> bytes | None:
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    font_path = _ensure_thai_font()
    units, share, comp, curr, prev = build_powertrain_tables(df)
    brands_all, _ = build_brand_table(df)
    brands_bev, _ = build_brand_table(df, pt_filter="BEV")
    bev_models, _ = build_bev_model_table(df)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    if font_path:
        pdf.add_font("Thai", "", font_path)
        font_name = "Thai"
    else:
        font_name = "Helvetica"

    def add_table(title: str, table_df: pd.DataFrame, reset_index=True):
        from fpdf.enums import XPos, YPos
        pdf.add_page()
        pdf.set_font(font_name, size=14)
        pdf.cell(0, 10, title, align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(3)

        if reset_index:
            tdf = table_df.reset_index()
        else:
            tdf = table_df.copy()

        cols = list(tdf.columns)
        n = len(cols)
        avail_w = pdf.w - pdf.l_margin - pdf.r_margin
        col_w = avail_w / n

        # Header row
        pdf.set_font(font_name, size=7)
        pdf.set_fill_color(31, 78, 121)
        pdf.set_text_color(255, 255, 255)
        for col in cols:
            pdf.cell(col_w, 7, str(col)[:18], border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_text_color(0, 0, 0)
        for i, (_, row) in enumerate(tdf.iterrows()):
            if i % 2 == 0:
                pdf.set_fill_color(235, 243, 251)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.set_font(font_name, size=7)
            for val in row:
                if isinstance(val, float):
                    text = f"{val:,.1f}" if val != int(val) else f"{int(val):,}"
                elif isinstance(val, int):
                    text = f"{val:,}"
                else:
                    text = str(val) if pd.notna(val) else ""
                pdf.cell(col_w, 6, text[:20], border=1, fill=True, align="R"
                         if isinstance(val, (int, float)) else "L")
            pdf.ln()

    add_table(f"Powertrain Monthly Units — {curr}", units)
    add_table(f"Powertrain Market Share % — {curr}", share)
    add_table(f"YTD Comparison {prev} vs {curr}", comp)
    add_table(f"Brand Rankings (All Powertrain) — {curr}", brands_all.head(40))
    add_table(f"BEV Brand Rankings — {curr}", brands_bev.head(40))
    add_table(f"BEV by Model — {curr}", bev_models.head(60), reset_index=False)

    return bytes(pdf.output())


# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Thailand EV Registration",
    page_icon="🚗",
    layout="wide",
)

df_loaded = load_data()

# Sidebar
with st.sidebar:
    st.title("Thailand EV Market")
    st.caption("Registration Report")
    st.divider()

    if st.button("🔄 Refresh Data", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.divider()
    st.subheader("Export")

    excel_bytes = build_excel(df_loaded)
    st.download_button(
        label="📥 Export Excel",
        data=excel_bytes,
        file_name="thailand_ev_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    if st.button("📄 Export PDF", use_container_width=True):
        with st.spinner("Generating PDF…"):
            pdf_bytes = build_pdf(df_loaded)
        if pdf_bytes:
            st.download_button(
                label="⬇ Download PDF",
                data=pdf_bytes,
                file_name="thailand_ev_report.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.error("fpdf2 not installed. Run: pip install fpdf2")


# Main content
_all_years  = sorted(df_loaded["ปี"].unique().tolist())
_all_brands = sorted(df_loaded["ยี่ห้อรถ2"].dropna().unique().tolist())

_col_year, _col_pt, _col_brand, _col_model = st.columns(4)
with _col_year:
    sel_years  = st.multiselect("Year (BE)", _all_years, default=_all_years, key="g_year")
with _col_pt:
    sel_pt     = st.multiselect("Powertrain", PT_ORDER, default=PT_ORDER, key="g_pt")
with _col_brand:
    sel_brands = st.multiselect("Brand", _all_brands, key="g_brand")
with _col_model:
    _model_opts = sorted(
        df_loaded[df_loaded["ยี่ห้อรถ2"].isin(sel_brands)]["รุ่นรถ2"].dropna().unique().tolist()
    ) if sel_brands else sorted(df_loaded["รุ่นรถ2"].dropna().unique().tolist())
    sel_models = st.multiselect("Model", _model_opts, key="g_model")

df = df_loaded
if sel_years:
    df = df[df["ปี"].isin(sel_years)]
if sel_pt:
    df = df[df["Powertrain"].isin(sel_pt)]
if sel_models:
    df = df[df["รุ่นรถ2"].isin(sel_models)]
elif sel_brands:
    df = df[df["ยี่ห้อรถ2"].isin(sel_brands)]

st.divider()

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Registration by Powertrain",
    "🏆 Brand Rankings",
    "⚡ BEV by Brand",
    "🔋 BEV by Model",
    "🔍 Compare",
])

# ─ Tab 1 ─────────────────────────────────────────────────────────────────────
with tab1:
    units, share, comp, curr, prev = build_powertrain_tables(df_loaded)

    st.subheader(f"Monthly Registrations by Powertrain — {curr}")
    fmt_units = {c: "{:,.0f}" for c in units.columns}
    st.dataframe(
        units.style.format(fmt_units),
        use_container_width=True,
    )

    st.subheader("Market Share % by Powertrain")
    fmt_share = {c: "{:.1f}%" for c in share.columns}
    st.dataframe(
        share.style.format(fmt_share),
        use_container_width=True,
    )

    st.subheader(f"YTD Comparison — {prev} vs {curr}")

    def _growth_color(val):
        if isinstance(val, (int, float)) and not pd.isna(val):
            return "color: green" if val > 0 else "color: red" if val < 0 else ""
        return ""

    fmt_comp = {
        f"{prev} Full Year": "{:,.0f}",
        f"{prev} YTD":       "{:,.0f}",
        f"{curr} YTD":       "{:,.0f}",
        "Growth %":          "{:.1f}%",
    }
    st.dataframe(
        comp.style
            .format(fmt_comp, na_rep="—")
            .map(_growth_color, subset=["Growth %"]),
        use_container_width=True,
    )

# ─ Tab 2 ─────────────────────────────────────────────────────────────────────
with tab2:
    brands_all, yr = build_brand_table(df_loaded)
    st.subheader(f"Brand Rankings — All Powertrain — {yr} YTD")
    num_cols = [c for c in brands_all.columns if c not in ("Rank", "Δ Rank")]
    st.dataframe(
        brands_all.style.format("{:,.0f}", subset=num_cols),
        use_container_width=True,
    )

# ─ Tab 3 ─────────────────────────────────────────────────────────────────────
with tab3:
    brands_bev, yr = build_brand_table(df_loaded, pt_filter="BEV")
    st.subheader(f"BEV Brand Rankings — {yr} YTD")
    num_cols = [c for c in brands_bev.columns if c not in ("Rank", "Δ Rank")]
    st.dataframe(
        brands_bev.style.format("{:,.0f}", subset=num_cols),
        use_container_width=True,
    )

# ─ Tab 4 ─────────────────────────────────────────────────────────────────────
with tab4:
    bev_models, yr = build_bev_model_table(df_loaded)
    st.subheader(f"BEV by Model (Major Brands) — {yr} YTD")
    num_cols = [c for c in bev_models.columns if c != "Δ Rank"]
    st.dataframe(
        bev_models.style.format("{:,.0f}", subset=num_cols),
        use_container_width=True,
    )

# ─ Tab 5 — Compare ────────────────────────────────────────────────────────────
with tab5:
    curr_c, prev_c, curr_months_c = year_info(df)

    all_brands_c = sorted(df["ยี่ห้อรถ2"].dropna().unique().tolist())
    all_models_c = sorted(df["รุ่นรถ2"].dropna().unique().tolist())

    col_b, col_m, col_pt = st.columns(3)
    with col_b:
        sel_brands = st.multiselect("Brand (ยี่ห้อรถ2)", all_brands_c, key="cmp_brand")
    with col_m:
        sel_models = st.multiselect("Model (รุ่นรถ2)", all_models_c, key="cmp_model")
    with col_pt:
        sel_pt = st.multiselect(
            "Powertrain", PT_ORDER, default=PT_ORDER, key="cmp_pt"
        )

    view_mode = st.radio("View", ["Monthly", "YTD"], horizontal=True, key="cmp_view")

    if not sel_brands and not sel_models:
        st.info("Select at least one brand or model above to start comparing.")
    elif not sel_pt:
        st.warning("Select at least one powertrain.")
    else:
        df_pt = df[df["Powertrain"].isin(sel_pt)]

        selections_c = []
        for brand in sel_brands:
            selections_c.append((brand, df_pt[df_pt["ยี่ห้อรถ2"] == brand]))
        for model in sel_models:
            brand_mode = df_pt[df_pt["รุ่นรถ2"] == model]["ยี่ห้อรถ2"].mode()
            brand_label = brand_mode.iloc[0] if len(brand_mode) > 0 else "?"
            selections_c.append((f"{brand_label} / {model}", df_pt[df_pt["รุ่นรถ2"] == model]))

        if view_mode == "Monthly":
            st.subheader(f"Monthly Registrations — {prev_c} vs {curr_c}")
            tbl = build_compare_monthly(df, selections_c, prev_c, curr_c)
            st.dataframe(tbl.style.format("{:,.0f}"), use_container_width=True)
        else:
            period = f"{MON_SHORT[curr_months_c[0]]}–{MON_SHORT[curr_months_c[-1]]}" if curr_months_c else "YTD"
            st.subheader(f"YTD Comparison ({period}) — {prev_c} vs {curr_c}")
            tbl = build_compare_ytd(df, selections_c, prev_c, curr_c, curr_months_c)
            fmt_ytd = {f"{prev_c} YTD": "{:,.0f}", f"{curr_c} YTD": "{:,.0f}", "Growth %": "{:.1f}%"}
            st.dataframe(
                tbl.style.format(fmt_ytd, na_rep="—").map(_growth_color, subset=["Growth %"]),
                use_container_width=True,
            )
